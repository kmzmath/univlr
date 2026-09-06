import argparse
import json
import time
import random
import re
import shutil
import sys
import unicodedata
from copy import copy
from pathlib import Path
from typing import Any, Dict, Iterable, List, Set, Tuple, Optional
from collections import defaultdict, deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from urllib.parse import quote
from datetime import datetime, timedelta, timezone

import requests
from openpyxl import load_workbook

import config_hub


# =========================
# CONFIG
# =========================
# A chave NUNCA fica no codigo: este repositorio e publico. Ela vem da
# variavel de ambiente RIOT_API_KEY ou do .env da pasta de trabalho.
RIOT_API_KEY = config_hub.riot_api_key()

VAL_REGION = "br"
RIOT_ROUTING = "americas"

# Input file
# PLAYERS_XLSX_PATH aponta para a planilha do site (repo "Hub Uni"), que e a
# fonte unica desde 03/09/2026: e la que moram as colunas "puuid 2"/"puuid 3"
# das contas alternativas e o cadastro que o build do site consome. Deixe None
# para voltar ao players.xlsx da pasta do script.
# ATENCAO: o script ESCREVE nesta planilha (PUUID resolvido e Riot ID atual) e
# grava um players.backup_<data>.xlsx ao lado antes de salvar.
PLAYERS_XLSX_PATH = str(config_hub.PLAYERS_XLSX)
PLAYERS_XLSX_NAME = "players.xlsx"
WRITE_RESOLVED_PUUIDS_TO_XLSX = True
WRITE_CURRENT_RIOT_IDS_TO_XLSX = True
BACKUP_PLAYERS_XLSX_BEFORE_WRITE = True

# Verificação/atualização via Account-V1 para quem JÁ tem PUUID salvo:
# - True: comportamento original — 1 chamada por jogador para anotar o Riot ID
#   atual nas colunas nick N e detectar PUUID inválido na hora (~560 chamadas/execução).
# - False: usa o PUUID salvo direto, sem nenhuma chamada. PUUIDs inválidos (raros)
#   são detectados na busca de matchlist (HTTP 400/404) e re-resolvidos pelo Riot ID
#   automaticamente na mesma execução.
# Ligue de vez em quando (ex.: 1x por semana) para atualizar os nicks na planilha.
REFRESH_CURRENT_RIOT_IDS = False

# Histórico/dedupe em Finalizados:
# Sempre considera TODOS os .json dentro de Finalizados, recursivamente.
# Isso inclui:
#   - Finalizados/*.json
#   - Finalizados/Válidas/**/*.json
#   - Finalizados/Inválidas/**/*.json
#   - qualquer subpasta futura
HISTORY_RECURSIVE_ALL_JSONS = True


# Depois da introdução de current_team:
# - True: somente jogadores com current_team preenchido participam da descoberta/validação de partidas.
# - False: todos os jogadores com PUUID/nick válido têm matchlist buscada, mesmo sem time atual.
# Para rosters ativos, mantenha True. Para coleta histórica, use uma tabela de passagens com datas.
ONLY_CURRENT_TEAM_PLAYERS_FOR_MATCH_DISCOVERY = True

# Duplicatas em out_tournament/matches
# O script considera matchInfo.matchId como chave única.
# "move" remove da pasta matches e preserva backup em out_tournament/duplicates_removed.
# "delete" apaga definitivamente os arquivos duplicados.
DUPLICATE_MATCH_ACTION = "move"  # "move" ou "delete"
DUPLICATE_MATCHES_DIR_NAME = "duplicates_removed"

# Filtro de data
# Formato: "YYYY-MM-DD"
# None = sem limite
FILTER_START_DATE = "2026-05-01"   # inclusivo
FILTER_END_DATE = "2026-09-06"     # inclusivo

# Aceita apenas partidas custom no modo padrão de Spike/Bomb.
# Rejeita Premier/Ranked/Matchmaking e modos custom alternativos como Retake/FortCollins.
REQUIRE_CUSTOM_BOMB_GAME = True
REQUIRED_PROVISIONING_FLOW_ID = "CustomGame"
REQUIRED_GAME_MODE = "/Game/GameModes/Bomb/BombGameMode.BombGameMode_C"

# Rejeita partidas incompletas/abandonadas.
# Além de matchInfo.isCompleted=True, exige pelo menos 5 jogadores ativos em cada lado
# e vencedor com placar mínimo de partida padrão (13+ rounds).
REQUIRE_COMPLETED_MATCH = True
MIN_ACTIVE_PLAYERS_PER_SIDE = 5
REQUIRE_WINNER_MIN_ROUNDS = True
MIN_WINNING_ROUNDS = 13

BRAZIL_TZ = timezone(timedelta(hours=-3))

# Match discovery heuristic
MIN_PLAYERS_IN_MATCHLIST = 6          # candidate threshold (matchId appears in >=6 players' matchlists)
MIN_PLAYERS_IN_MATCH_PAYLOAD = 6      # confirm threshold in /matches response
MIN_PLAYERS_PER_SIDE_IN_PAYLOAD = 2   # reject 5x1/6x0 known-player splits; require >=2 known players on Blue and Red

# Team detection heuristic
MIN_TEAM_PLAYERS_TO_IDENTIFY = 3

# How many matchlist entries per player to consider
LOOKBACK_PER_PLAYER = 120

# Safety: cap how many matchIds we will download per run
MAX_MATCHES_TO_FETCH = 800

# Parallelism
# O teto de app (100 req / 2 min) equivale a ~0,83 req/s sustentado, então
# concorrência alta não aumenta a vazão: o limiter apenas enfileira as threads.
# 8 workers aproveitam a rajada inicial de 20 req/s sem segurar conexões ociosas.
MAX_WORKERS = 8
HTTP_TIMEOUT_S = 10.0
MAX_RETRIES = 5

# =========================
# Rate limits da Riot
# =========================
# A Riot aplica DUAS camadas de limite ao mesmo tempo:
#
#   1) Limite de APP: vale para a soma de TODAS as rotas da chave.
#      Chave de desenvolvimento: 20 req/1s E 100 req/2min.
#      É o teto mais apertado e o que realmente manda aqui.
#
#   2) Limite de MÉTODO: por rota, bem mais folgado.
#      account-v1 by-riot-id / by-puuid: 1000 req/1min
#      val-match-v1 matches e matchlists: 800 req/10s
#
# Cada tupla é (max_chamadas, janela_em_segundos); todas as janelas de uma
# camada são respeitadas simultaneamente.
#
# Ao trocar por uma chave de produção, atualize APP_RATE_LIMITS com os limites
# da nova chave; o resto do script se ajusta sozinho.
APP_RATE_LIMITS = [
    (20, 1.0),
    (100, 120.0),
]

ACCOUNT_METHOD_LIMITS = [(1000, 60.0)]
VAL_METHOD_LIMITS = [(800, 10.0)]

# Margem de segurança sobre os limites acima (0.9 = usa 90% do teto).
# Protege contra relógio dessincronizado e requisições em voo que a Riot conta
# num instante diferente do nosso.
RATE_LIMIT_SAFETY = 0.85


# =========================
# Filesystem helpers
# =========================
def _base_dir() -> Path:
    """
    Pasta de trabalho: Finalizados, out_tournament, matchlists.

    Deixou de ser "a pasta do script" quando os scripts vieram para dentro do
    repo. O acervo bruto continua fora - sao centenas de MB que nao precisam de
    historico de versao e que duplicariam o campeonatos/ do site.
    """
    return config_hub.TRABALHO_DIR


def ensure_dirs(base: Path):
    (base / "matchlists").mkdir(parents=True, exist_ok=True)
    (base / "matches").mkdir(parents=True, exist_ok=True)




# =========================
# Seleção de partidas em Finalizados
# =========================
# Opções: "validas", "invalidas", "ambas" ou "raiz".
# - validas: usa Finalizados/Válidas ou Finalizados/Validas
# - invalidas: usa Finalizados/Inválidas ou Finalizados/Invalidas
# - ambas: usa Válidas + Inválidas
# - raiz: usa diretamente a pasta informada, sem entrar nessas subpastas

def _fold_ascii(value: Any) -> str:
    s = str(value or "").strip().casefold()
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[\s_\-]+", "-", s).strip("-")
    return s


def normalize_match_source(value: Any) -> str:
    folded = _fold_ascii(value)
    aliases = {
        "validas": "validas",
        "valida": "validas",
        "validos": "validas",
        "valido": "validas",
        "invalidas": "invalidas",
        "invalida": "invalidas",
        "invalidos": "invalidas",
        "invalido": "invalidas",
        "ambas": "ambas",
        "ambos": "ambas",
        "todas": "ambas",
        "todos": "ambas",
        "both": "ambas",
        "all": "ambas",
        "raiz": "raiz",
        "root": "raiz",
        "base": "raiz",
    }
    if folded not in aliases:
        raise ValueError("match_source inválido. Use: validas, invalidas, ambas ou raiz.")
    return aliases[folded]


# A pasta no disco pode estar no masculino ou no feminino. Procurar só por uma
# grafia fazia o resolvedor não achar "Válidos", cair no caminho inventado
# "Válidas", que não existe, e devolver zero arquivo em silêncio.
GRAFIAS_DA_FONTE = {
    "validas": ("validas", "validos", "valida", "valido"),
    "invalidas": ("invalidas", "invalidos", "invalida", "invalido"),
}


def _find_named_child_dir(base_dir: Path, folded_name: str) -> Path:
    aceitas = GRAFIAS_DA_FONTE.get(folded_name, (folded_name,))
    if base_dir.exists():
        for child in base_dir.iterdir():
            if child.is_dir() and _fold_ascii(child.name) in aceitas:
                return child

    if folded_name == "validas":
        return base_dir / "Válidas"
    if folded_name == "invalidas":
        return base_dir / "Inválidas"
    return base_dir / folded_name


def resolve_match_source_dirs(base_dir: Path, match_source: str) -> List[Path]:
    choice = normalize_match_source(match_source)
    base_dir = Path(base_dir)

    if choice == "raiz":
        return [base_dir]

    wanted = ["validas", "invalidas"] if choice == "ambas" else [choice]
    dirs = [_find_named_child_dir(base_dir, item) for item in wanted]

    if choice == "ambas" and not any(d.exists() and d.is_dir() for d in dirs):
        return [base_dir]

    return dirs


def iter_json_files_from_match_source(base_dir: Path, recursive: bool, match_source: str) -> Iterable[Path]:
    # Em modo raiz, lê apenas a raiz antiga, não Válidas/Inválidas por recursão.
    source_choice = normalize_match_source(match_source)
    pattern = "*.json" if source_choice == "raiz" else ("**/*.json" if recursive else "*.json")
    seen: Set[Path] = set()

    for source_dir in resolve_match_source_dirs(base_dir, match_source):
        if not source_dir.exists() or not source_dir.is_dir():
            continue
        for fp in sorted(source_dir.glob(pattern)):
            if not fp.is_file():
                continue
            key = fp.resolve()
            if key in seen:
                continue
            seen.add(key)
            yield fp


def load_existing_match_index_from_sources(base_dir: Path, recursive: bool, match_source: str) -> Dict[str, str]:
    """Lê matchInfo.matchId dos JSONs nas fontes escolhidas dentro de Finalizados."""
    index: Dict[str, str] = {}
    if not base_dir.exists():
        return index

    for fp in iter_json_files_from_match_source(base_dir, recursive=recursive, match_source=match_source):
        mid = read_match_id_from_json_file(fp)
        if not mid:
            continue
        try:
            label = str(fp.relative_to(base_dir))
        except ValueError:
            label = fp.name
        existing = index.get(mid)
        if existing is None or label < existing:
            index[mid] = label

    return index


def iter_all_history_json_files(base_dir: Path) -> Iterable[Path]:
    """
    Itera todos os .json dentro de Finalizados, recursivamente.

    Não há mais seleção de Válidas/Inválidas/raiz para o PegaCustomMatch:
    o histórico usado para dedupe deve considerar todos os jogos já arquivados.
    """
    if not base_dir.exists() or not base_dir.is_dir():
        return

    seen: Set[Path] = set()
    for fp in sorted(base_dir.rglob("*.json")):
        if not fp.is_file():
            continue
        key = fp.resolve()
        if key in seen:
            continue
        seen.add(key)
        yield fp


def load_existing_match_index_all_history(base_dir: Path) -> Dict[str, str]:
    """Lê matchInfo.matchId de todos os JSONs em Finalizados, incluindo subpastas."""
    index: Dict[str, str] = {}
    if not base_dir.exists():
        return index

    for fp in iter_all_history_json_files(base_dir):
        mid = read_match_id_from_json_file(fp)
        if not mid:
            continue
        try:
            label = str(fp.relative_to(base_dir))
        except ValueError:
            label = fp.name
        existing = index.get(mid)
        if existing is None or label < existing:
            index[mid] = label

    return index

def save_json(p: Path, obj: Any):
    p.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def read_match_id_from_json_file(fp: Path) -> Optional[str]:
    """Lê matchInfo.matchId de um JSON de partida. Retorna None se não for um JSON válido de match."""
    try:
        obj = json.loads(fp.read_text(encoding="utf-8"))
    except Exception:
        return None

    if not isinstance(obj, dict):
        return None

    mi = obj.get("matchInfo") or {}
    if not isinstance(mi, dict):
        return None

    mid = mi.get("matchId")
    return mid if isinstance(mid, str) and mid else None


def load_existing_match_files_by_id(matches_dir: Path) -> Dict[str, List[str]]:
    """
    Lê todos os *.json em matches_dir e agrupa por matchInfo.matchId:
      matchId -> [filename1, filename2, ...]

    Isso permite detectar duplicatas mesmo quando os arquivos já foram renomeados.
    """
    grouped: Dict[str, List[str]] = defaultdict(list)
    if not matches_dir.exists():
        return {}

    for fp in matches_dir.glob("*.json"):
        mid = read_match_id_from_json_file(fp)
        if mid:
            grouped[mid].append(fp.name)

    return dict(grouped)


def choose_primary_match_file(matches_dir: Path, file_names: List[str]) -> Optional[str]:
    """
    Escolhe qual arquivo preservar quando há duplicatas do mesmo matchId.

    Critério prático:
      1) maior tamanho em bytes;
      2) modificação mais recente;
      3) nome lexicograficamente menor para desempate estável.
    """
    candidates: List[Tuple[int, float, str]] = []
    for name in file_names:
        fp = matches_dir / name
        if not fp.exists():
            continue
        try:
            st = fp.stat()
        except OSError:
            continue
        candidates.append((int(st.st_size), float(st.st_mtime), name))

    if not candidates:
        return None

    candidates.sort(key=lambda item: (-item[0], -item[1], item[2]))
    return candidates[0][2]


def remove_duplicate_file(fp: Path, backup_dir: Path, reason: str) -> Dict[str, Any]:
    """Remove um arquivo duplicado de matches/. Por padrão move para backup em vez de apagar."""
    report = {
        "fileName": fp.name,
        "path": str(fp),
        "reason": reason,
        "action": DUPLICATE_MATCH_ACTION,
    }

    if not fp.exists():
        report["removed"] = False
        report["error"] = "file_not_found"
        return report

    try:
        if DUPLICATE_MATCH_ACTION == "delete":
            fp.unlink()
            report["removed"] = True
        else:
            backup_dir.mkdir(parents=True, exist_ok=True)
            target = backup_dir / fp.name
            if target.exists():
                stem = target.stem
                suffix = target.suffix
                counter = 2
                while True:
                    candidate = backup_dir / f"{stem}_{counter}{suffix}"
                    if not candidate.exists():
                        target = candidate
                        break
                    counter += 1
            shutil.move(str(fp), str(target))
            report["removed"] = True
            report["backupPath"] = str(target)
    except Exception as e:
        report["removed"] = False
        report["error"] = str(e)

    return report


def dedupe_matches_dir_by_match_id(matches_dir: Path, backup_dir: Path) -> Tuple[Dict[str, str], List[Dict[str, Any]]]:
    """
    Remove duplicatas em out_tournament/matches usando matchInfo.matchId como chave única.

    Retorna:
      - índice limpo: matchId -> filename preservado
      - relatório de arquivos duplicados removidos/movidos
    """
    grouped = load_existing_match_files_by_id(matches_dir)
    clean_index: Dict[str, str] = {}
    duplicate_report: List[Dict[str, Any]] = []

    for mid, file_names in grouped.items():
        primary = choose_primary_match_file(matches_dir, file_names)
        if not primary:
            continue

        clean_index[mid] = primary

        for name in sorted(set(file_names)):
            if name == primary:
                continue
            duplicate_report.append(
                remove_duplicate_file(
                    matches_dir / name,
                    backup_dir,
                    reason=f"duplicate_matchId({mid}); kept({primary})",
                )
            )

    return clean_index, duplicate_report


def load_existing_match_index(matches_dir: Path) -> Dict[str, str]:
    """
    Lê todos os *.json em matches_dir e retorna matchId -> filename preservado.
    Não remove arquivos; use dedupe_matches_dir_by_match_id para limpar duplicatas.
    """
    grouped = load_existing_match_files_by_id(matches_dir)
    index: Dict[str, str] = {}
    for mid, file_names in grouped.items():
        primary = choose_primary_match_file(matches_dir, file_names)
        if primary:
            index[mid] = primary
    return index

def _date_to_epoch_ms(date_str: Optional[str], end_of_day: bool = False) -> Optional[int]:
    if not date_str:
        return None

    dt = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=BRAZIL_TZ)
    if end_of_day:
        dt += timedelta(days=1)  # fim exclusivo do dia
    return int(dt.timestamp() * 1000)


FILTER_START_MS = _date_to_epoch_ms(FILTER_START_DATE, end_of_day=False)
FILTER_END_MS_EXCLUSIVE = _date_to_epoch_ms(FILTER_END_DATE, end_of_day=True)


def is_in_date_range(game_start_ms: Any) -> bool:
    if FILTER_START_MS is None and FILTER_END_MS_EXCLUSIVE is None:
        return True

    if not isinstance(game_start_ms, (int, float)):
        return False

    game_start_ms = int(game_start_ms)

    if FILTER_START_MS is not None and game_start_ms < FILTER_START_MS:
        return False

    if FILTER_END_MS_EXCLUSIVE is not None and game_start_ms >= FILTER_END_MS_EXCLUSIVE:
        return False

    return True


def match_type_info(match: Dict[str, Any]) -> Dict[str, str]:
    mi = match.get("matchInfo") or {}
    if not isinstance(mi, dict):
        mi = {}

    return {
        "provisioningFlowId": str(mi.get("provisioningFlowId") or ""),
        "gameMode": str(mi.get("gameMode") or ""),
        "queueId": str(mi.get("queueId") or ""),
        "customGameName": str(mi.get("customGameName") or ""),
    }


def is_custom_bomb_game(match: Dict[str, Any]) -> bool:
    info = match_type_info(match)
    return (
        info["provisioningFlowId"] == REQUIRED_PROVISIONING_FLOW_ID
        and info["gameMode"] == REQUIRED_GAME_MODE
    )


def reject_non_custom_bomb_record(
    mid: str,
    match: Dict[str, Any],
    source: str,
    current_file_name: Optional[str] = None,
) -> Dict[str, Any]:
    info = match_type_info(match)
    return {
        "matchId": mid,
        "saved": False,
        "source": source,
        "currentFileName": current_file_name,
        "reason": (
            "not_custom_bomb_game("
            f"provisioningFlowId={info['provisioningFlowId']!r},"
            f"gameMode={info['gameMode']!r})"
        ),
        **info,
    }


def get_rounds_won_by_side(match: Dict[str, Any]) -> Dict[str, int]:
    """Extrai rounds vencidos por lado a partir do payload teams[]."""
    rounds_by_side: Dict[str, int] = {}
    teams_payload = match.get("teams") or []
    if not isinstance(teams_payload, list):
        return rounds_by_side

    for team in teams_payload:
        if not isinstance(team, dict):
            continue

        side = team.get("teamId")
        if side not in ("Blue", "Red"):
            continue

        score: Optional[int] = None
        for key in ("roundsWon", "numPoints", "score", "wonRounds"):
            value = team.get(key)
            if isinstance(value, int):
                score = value
                break

        if score is not None:
            rounds_by_side[str(side)] = int(score)

    return rounds_by_side


def get_winning_rounds(match: Dict[str, Any]) -> int:
    """Retorna o maior placar de rounds entre Blue/Red; 0 se não houver placar confiável."""
    rounds_by_side = get_rounds_won_by_side(match)
    if not rounds_by_side:
        return 0
    return max(rounds_by_side.values())

# =========================
# Team helpers
# =========================
NO_CURRENT_TEAM_VALUES = {
    "-",
    "sem_time",
    "sem time",
    "sem equipe",
    "sem_team",
    "free_agent",
    "free agent",
    "none",
    "null",
    "n/a",
    "na",
}


def normalize_team(value: Any) -> Optional[str]:
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    normalized = re.sub(r"\s+", " ", s).strip().lower()
    normalized = normalized.replace("-", "_")
    normalized = normalized.replace(" ", "_")

    if normalized in {v.replace(" ", "_").replace("-", "_") for v in NO_CURRENT_TEAM_VALUES}:
        return None

    return s


def get_last_historical_team(team_values: List[Any]) -> Optional[str]:
    """
    Retorna o último time preenchido no histórico team N.

    Importante: com a coluna current_team, este valor NÃO deve mais ser usado
    automaticamente como time atual. Ele é apenas o último time conhecido no histórico.
    """
    for value in reversed(team_values):
        team = normalize_team(value)
        if team is not None:
            return team
    return None


def resolve_current_team_value(current_team_value: Any, team_values: List[Any], has_current_team_column: bool) -> Optional[str]:
    """
    Define o time atual do jogador.

    Regra nova:
      - se existir coluna current_team, ela manda; vazio/sem_time = sem time atual;
      - se não existir current_team, usa fallback legado: último team N preenchido.
    """
    if has_current_team_column:
        return normalize_team(current_team_value)

    return get_last_historical_team(team_values)


# =========================
# Filename helpers
# =========================
def sanitize_filename_part(value: Any) -> str:
    s = str(value or "").strip()
    if not s:
        return "unknown"

    s = re.sub(r'[<>:"/\\|?*]+', "", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s)
    s = s.strip("._- ")

    return s or "unknown"


def normalize_map_name(map_value: Any) -> str:
    """
    Traduz mapId/path interno para nome amigável de mapa.
    O retorno é em minúsculas para manter o padrão do nome do arquivo.
    """
    s = str(map_value or "").strip()
    if not s:
        return "unknown"

    s = s.replace("\\", "/").lower()
    parts = [p for p in s.split("/") if p]
    token = parts[-1] if parts else s

    if "." in token:
        token = token.rsplit(".", 1)[0]

    token = token.strip().lower()

    map_aliases = {
        # aliases internos do VALORANT
        "ascent": "ascent",
        "infinity": "abyss",
        "duality": "bind",
        "foxtrot": "breeze",
        "rook": "corrode",
        "canyon": "fracture",
        "triad": "haven",
        "port": "icebox",
        "jam": "lotus",
        "pitt": "pearl",
        "bonsai": "split",
        "juliett": "sunset",
    }

    return map_aliases.get(token, token if token else "unknown")


def format_match_date_for_filename(game_start_ms: Any) -> str:
    """
    Converte matchInfo.gameStartMillis para DD-MM-AA no fuso do Brasil.

    Exemplo:
      2026-04-24 -> 24-04-26
    """
    if isinstance(game_start_ms, (int, float)):
        try:
            dt = datetime.fromtimestamp(int(game_start_ms) / 1000, tz=BRAZIL_TZ)
            return dt.strftime("%d-%m-%y")
        except (OSError, OverflowError, ValueError):
            pass

    return "data_desconhecida"


def build_match_filename(
    order_num: int,
    game_start_ms: Any,
    detected_teams: List[Dict[str, Any]],
    score_a: Optional[int],
    score_b: Optional[int],
    map_id: Any,
    match_id: Any,
) -> str:
    """
    Formato:
      <ordem>_<data-DD-MM-AA>_<teamA>_<placarA>_x_<placarB>_<teamB>_<mapa>.json

    Exemplo:
      1_24-04-26_caap_hellhounds_11_x_13_ufu_saints_lotus.json
    """
    team_names: List[str] = []

    for team in detected_teams:
        raw_name = team.get("teamName")
        if not raw_name:
            continue

        clean_name = sanitize_filename_part(raw_name).lower()
        if clean_name != "unknown" and clean_name not in team_names:
            team_names.append(clean_name)

    team_names = team_names[:2]

    if len(team_names) >= 2:
        team_a = team_names[0]
        team_b = team_names[1]
    elif len(team_names) == 1:
        team_a = team_names[0]
        team_b = "unknown"
    else:
        fallback = sanitize_filename_part(match_id).lower()
        team_a = fallback
        team_b = "unknown"

    score_a_part = str(score_a) if isinstance(score_a, int) else "0"
    score_b_part = str(score_b) if isinstance(score_b, int) else "0"
    match_date = format_match_date_for_filename(game_start_ms)
    map_name = normalize_map_name(map_id)

    return f"{order_num}_{match_date}_{team_a}_{score_a_part}_x_{score_b_part}_{team_b}_{map_name}.json"


def extract_match_score(
    match: Dict[str, Any],
    detected_teams: List[Dict[str, Any]],
) -> Tuple[Optional[int], Optional[int]]:
    """
    Tenta extrair o placar do payload do match.

    Prioriza associar score por teamId, usando a mesma ordem de detected_teams.
    """
    teams_payload = match.get("teams") or []
    if not isinstance(teams_payload, list):
        return None, None

    score_by_team_id: Dict[str, int] = {}

    for team in teams_payload:
        if not isinstance(team, dict):
            continue

        team_id = team.get("teamId")
        if not isinstance(team_id, str):
            continue

        score = None
        for key in ("roundsWon", "wonRounds", "score"):
            value = team.get(key)
            if isinstance(value, int):
                score = value
                break

        if score is not None:
            score_by_team_id[team_id] = score

    if len(detected_teams) >= 2:
        team_a_id = detected_teams[0].get("teamId")
        team_b_id = detected_teams[1].get("teamId")

        score_a = score_by_team_id.get(team_a_id) if isinstance(team_a_id, str) else None
        score_b = score_by_team_id.get(team_b_id) if isinstance(team_b_id, str) else None

        if isinstance(score_a, int) and isinstance(score_b, int):
            return score_a, score_b

    fallback_scores = [v for v in score_by_team_id.values() if isinstance(v, int)]
    fallback_scores.sort(reverse=True)

    if len(fallback_scores) >= 2:
        return fallback_scores[0], fallback_scores[1]

    return None, None


# =========================
# Excel parsing (players.xlsx)
# =========================
def parse_riot_id(value: str) -> Tuple[str, str]:
    """
    Parses 'gameName#tagLine' splitting on the LAST '#'.
    Supports spaces and non-ascii in gameName.
    """
    s = str(value).strip()
    if not s:
        raise ValueError("empty riot id")
    if "#" not in s:
        raise ValueError(f"riot id missing '#': {s}")
    game_name, tag_line = s.rsplit("#", 1)
    game_name = game_name.strip()
    tag_line = tag_line.strip()
    if not game_name or not tag_line:
        raise ValueError(f"invalid riot id (empty part): {s}")
    return game_name, tag_line


def normalize_puuid(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == "-":
        return None
    return s


def is_probably_puuid(value: Any) -> bool:
    """Validação estrutural simples para evitar chamar Account-V1 com Riot ID na coluna puuid."""
    s = normalize_puuid(value)
    if not s:
        return False
    return len(s) == 78 and re.fullmatch(r"[A-Za-z0-9_-]+", s) is not None


def normalize_header_key(value: Any) -> Optional[str]:
    """
    Normaliza cabeçalhos antigos e novos para chaves internas estáveis.

    Aceita, por exemplo:
      - "Jogador"
      - "team1", "team 1", "team 2", "team 3", ...
      - "time1", "time 1", "equipe 1", ...
      - "puuid"
      - "nick1", "nick 1", "Nick 1", ...

    Internamente o script usa:
      jogador, team1, team2, team3, ..., puuid, nick1, nick2, ...
    """
    if value is None:
        return None

    s = str(value).strip().lower()
    if not s:
        return None

    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    compact = re.sub(r"\s+", "", s)

    aliases = {
        "jogador": "jogador",
        "player": "jogador",
        "nome": "jogador",
        "name": "jogador",
        "puuid": "puuid",
        "currentteam": "current_team",
        "teamatual": "current_team",
        "timeatual": "current_team",
        "equipeatual": "current_team",
    }
    if compact in aliases:
        return aliases[compact]

    m = re.fullmatch(r"(?:team|time|equipe)(\d*)", compact)
    if m:
        raw_idx = m.group(1)
        idx = int(raw_idx) if raw_idx else 1
        if idx <= 0:
            raise RuntimeError(f"Cabeçalho de equipe inválido: {value}")
        return f"team{idx}"

    m = re.fullmatch(r"nick(\d*)", compact)
    if m:
        raw_idx = m.group(1)
        idx = int(raw_idx) if raw_idx else 1
        if idx <= 0:
            raise RuntimeError(f"Cabeçalho de nick inválido: {value}")
        return f"nick{idx}"

    return compact


def canonical_header_name(key: str) -> str:
    if key == "jogador":
        return "Jogador"
    if key == "puuid":
        return "puuid"
    if key == "current_team":
        return "current_team"

    m = re.fullmatch(r"team(\d+)", key)
    if m:
        return f"team {int(m.group(1))}"

    m = re.fullmatch(r"nick(\d+)", key)
    if m:
        return f"nick {int(m.group(1))}"

    m = re.fullmatch(r"puuid(\d+)", key)
    if m:
        return f"puuid {int(m.group(1))}"

    return key


def build_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_header_key(ws.cell(row=1, column=col).value)
        if not key:
            continue
        if key in headers:
            raise RuntimeError(
                f"Coluna duplicada no players.xlsx: '{canonical_header_name(key)}' "
                f"aparece nas colunas {headers[key]} e {col}."
            )
        headers[key] = col
    return headers


def ensure_column(ws, headers: Dict[str, int], key: str) -> Tuple[int, bool]:
    existing_col = headers.get(key)
    if existing_col is not None:
        return existing_col, False

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = canonical_header_name(key)
    headers[key] = col
    return col, True


def ensure_puuid_column(ws, headers: Dict[str, int]) -> Tuple[int, bool]:
    return ensure_column(ws, headers, "puuid")


def normalize_riot_id_for_compare(value: Any) -> Optional[str]:
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    try:
        gn, tg = parse_riot_id(s)
    except ValueError:
        return None

    return f"{gn.lower()}#{tg.lower()}"


def account_to_riot_id(account: Any) -> Optional[str]:
    if not isinstance(account, dict):
        return None

    game_name = account.get("gameName")
    tag_line = account.get("tagLine")

    if game_name is None or tag_line is None:
        return None

    game_name = str(game_name).strip()
    tag_line = str(tag_line).strip()

    if not game_name or not tag_line:
        return None

    return f"{game_name}#{tag_line}"


def alt_puuid_index_from_key(key: str) -> Optional[int]:
    """
    Indice das colunas de conta alternativa: "puuid 2" -> 2, "puuid 3" -> 3.

    A coluna "puuid" (conta principal) NAO entra aqui: ela e a que vira o id do
    jogador e a unica que o script escreve de volta.
    """
    m = re.fullmatch(r"puuid(\d+)", key)
    if not m:
        return None
    idx = int(m.group(1))
    return idx if idx >= 2 else None


def get_alt_puuid_columns(headers: Dict[str, int]) -> List[Tuple[int, str]]:
    cols = [(c, h) for h, c in headers.items() if alt_puuid_index_from_key(h) is not None]
    cols.sort(key=lambda x: alt_puuid_index_from_key(x[1]) or 0)
    return cols


def team_index_from_key(key: str) -> Optional[int]:
    m = re.fullmatch(r"team(\d+)", key)
    if not m:
        return None
    return int(m.group(1))


def get_team_columns(headers: Dict[str, int]) -> List[Tuple[int, str]]:
    team_cols: List[Tuple[int, str]] = []
    for h, c in headers.items():
        if team_index_from_key(h) is not None:
            team_cols.append((c, h))
    team_cols.sort(key=lambda x: team_index_from_key(x[1]) or 0)
    return team_cols


def nick_index_from_key(key: str) -> Optional[int]:
    m = re.fullmatch(r"nick(\d+)", key)
    if not m:
        return None
    return int(m.group(1))


def get_nick_columns(headers: Dict[str, int]) -> List[Tuple[int, str]]:
    nick_cols: List[Tuple[int, str]] = []
    for h, c in headers.items():
        if nick_index_from_key(h) is not None:
            nick_cols.append((c, h))
    nick_cols.sort(key=lambda x: nick_index_from_key(x[1]) or 0)
    return nick_cols


def next_nick_header_name(headers: Dict[str, int]) -> str:
    max_idx = 0
    for h in headers:
        idx = nick_index_from_key(h)
        if idx is not None:
            max_idx = max(max_idx, idx)

    return f"nick{max_idx + 1 if max_idx > 0 else 1}"


def create_next_nick_column(ws, headers: Dict[str, int]) -> Tuple[int, str]:
    header_key = next_nick_header_name(headers)
    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = canonical_header_name(header_key)
    headers[header_key] = col
    return col, header_key


def normalize_players_sheet_layout(ws, headers: Dict[str, int]) -> Tuple[Dict[str, int], Dict[str, Any]]:
    """
    Padroniza o players.xlsx para a ordem:
      Jogador | team 1 | ... | current_team | puuid | puuid 2 | ... | nick 1 | nick 2 | ...

    Não existe limite fixo de team nem de nick. O script usa todas as colunas team N
    existentes e cria a próxima coluna nick N sempre que precisar gravar um Riot ID
    novo e todas as colunas existentes da linha estiverem cheias.
    """
    created_base_columns: List[str] = []

    for key in ("jogador", "team1", "puuid"):
        _col, created = ensure_column(ws, headers, key)
        if created:
            created_base_columns.append(canonical_header_name(key))

    if not get_nick_columns(headers):
        _col, header_key = create_next_nick_column(ws, headers)
        created_base_columns.append(canonical_header_name(header_key))

    team_keys = sorted(
        [key for key in headers if team_index_from_key(key) is not None],
        key=lambda key: team_index_from_key(key) or 0,
    )

    nick_keys = sorted(
        [key for key in headers if nick_index_from_key(key) is not None],
        key=lambda key: nick_index_from_key(key) or 0,
    )

    current_team_keys = ["current_team"] if "current_team" in headers else []

    alt_puuid_keys = [key for _col, key in get_alt_puuid_columns(headers)]

    # `completes` anda junto do current_team: as duas dizem de que equipe a
    # pessoa é, uma como elenco e a outra como reforço.
    completes_keys = ["completes"] if "completes" in headers else []

    fixed_keys = ["jogador"] + team_keys + current_team_keys + completes_keys + ["puuid"] + alt_puuid_keys
    known_keys = set(fixed_keys + nick_keys)
    other_keys = [
        key for key, _col in sorted(headers.items(), key=lambda item: item[1])
        if key not in known_keys
    ]
    desired_keys = fixed_keys + nick_keys + other_keys

    current_keys = []
    for col in range(1, ws.max_column + 1):
        key = normalize_header_key(ws.cell(row=1, column=col).value)
        if key:
            current_keys.append(key)

    canonical_names_needed = [canonical_header_name(key) for key in desired_keys]
    current_names = [str(ws.cell(row=1, column=headers[key]).value).strip() for key in desired_keys if key in headers]

    order_changed = current_keys[:len(desired_keys)] != desired_keys
    names_changed = current_names != canonical_names_needed

    if not order_changed and not names_changed:
        return headers, {
            "layoutChanged": False,
            "createdBaseColumns": created_base_columns,
            "canonicalOrder": canonical_names_needed,
        }

    max_row = ws.max_row
    max_col = ws.max_column

    snapshots: Dict[str, List[Dict[str, Any]]] = {}
    for key in desired_keys:
        source_col = headers[key]
        snapshots[key] = []
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=source_col)
            snapshots[key].append({
                "value": cell.value,
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            })

    # Limpa a área antiga antes de reescrever a ordem canônica.
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    for new_col, key in enumerate(desired_keys, start=1):
        for row, snapshot in enumerate(snapshots[key], start=1):
            cell = ws.cell(row=row, column=new_col)
            cell.value = snapshot["value"]
            cell._style = copy(snapshot["style"])
            cell.number_format = snapshot["number_format"]
            cell.font = copy(snapshot["font"])
            cell.fill = copy(snapshot["fill"])
            cell.border = copy(snapshot["border"])
            cell.alignment = copy(snapshot["alignment"])
            cell.protection = copy(snapshot["protection"])

        ws.cell(row=1, column=new_col).value = canonical_header_name(key)

    new_headers = {key: idx for idx, key in enumerate(desired_keys, start=1)}
    return new_headers, {
        "layoutChanged": True,
        "createdBaseColumns": created_base_columns,
        "canonicalOrder": canonical_names_needed,
    }

def backup_xlsx(xlsx_path: Path) -> Optional[Path]:
    if not xlsx_path.exists():
        return None
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = xlsx_path.with_name(f"{xlsx_path.stem}.backup_{stamp}{xlsx_path.suffix}")
    shutil.copy2(xlsx_path, backup_path)
    return backup_path


def write_resolved_puuids_to_xlsx(
    xlsx_path: Path,
    player_key_to_row: Dict[str, int],
    player_key_to_puuid: Dict[str, str],
    player_key_to_current_riotid: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    if not WRITE_RESOLVED_PUUIDS_TO_XLSX and not WRITE_CURRENT_RIOT_IDS_TO_XLSX:
        return {
            "enabled": False,
            "updatedPuuidRows": 0,
            "updatedNickRows": 0,
            "createdPuuidColumn": False,
            "createdNickColumns": [],
            "puuidConflicts": [],
            "nickConflicts": [],
        }

    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = build_headers(ws)

    puuid_col: Optional[int] = None
    created_puuid_col = False
    if WRITE_RESOLVED_PUUIDS_TO_XLSX:
        puuid_col, created_puuid_col = ensure_puuid_column(ws, headers)

    puuid_conflicts: List[Dict[str, Any]] = []
    nick_conflicts: List[Dict[str, Any]] = []
    created_nick_columns: List[str] = []
    updated_puuid_rows = 0
    updated_nick_rows = 0

    for player_key, puuid in player_key_to_puuid.items():
        row = player_key_to_row.get(player_key)
        if row is None or puuid_col is None:
            continue

        cell = ws.cell(row=row, column=puuid_col)
        existing = normalize_puuid(cell.value)

        if existing and existing != puuid:
            puuid_conflicts.append({
                "playerKey": player_key,
                "row": row,
                "existingPuuid": existing,
                "resolvedPuuid": puuid,
            })
            continue

        if not existing:
            cell.value = puuid
            updated_puuid_rows += 1

    if WRITE_CURRENT_RIOT_IDS_TO_XLSX and player_key_to_current_riotid:
        nick_cols = get_nick_columns(headers)
        if not nick_cols:
            col, header_name = create_next_nick_column(ws, headers)
            nick_cols = [(col, header_name)]
            created_nick_columns.append(header_name)

        # Map existing Riot IDs in the whole sheet. This prevents writing the same nick into two players.
        riotid_to_locations: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
        for row in range(2, ws.max_row + 1):
            for col, header_name in nick_cols:
                norm = normalize_riot_id_for_compare(ws.cell(row=row, column=col).value)
                if norm:
                    riotid_to_locations[norm].append({"row": row, "column": col, "header": header_name})

        for player_key, riot_id in player_key_to_current_riotid.items():
            row = player_key_to_row.get(player_key)
            if row is None:
                continue

            norm = normalize_riot_id_for_compare(riot_id)
            if not norm:
                continue

            existing_locations = riotid_to_locations.get(norm, [])
            if any(loc["row"] == row for loc in existing_locations):
                continue

            if existing_locations:
                nick_conflicts.append({
                    "playerKey": player_key,
                    "row": row,
                    "riotId": riot_id,
                    "existingLocations": existing_locations,
                })
                continue

            target_cell = None
            for col, _header_name in nick_cols:
                cell = ws.cell(row=row, column=col)
                if cell.value is None or str(cell.value).strip() == "":
                    target_cell = cell
                    break

            if target_cell is None:
                col, header_name = create_next_nick_column(ws, headers)
                nick_cols.append((col, header_name))
                created_nick_columns.append(header_name)
                target_cell = ws.cell(row=row, column=col)

            target_cell.value = riot_id
            riotid_to_locations[norm].append({
                "row": row,
                "column": target_cell.column,
                "header": ws.cell(row=1, column=target_cell.column).value,
            })
            updated_nick_rows += 1

    headers, layout_report = normalize_players_sheet_layout(ws, build_headers(ws))

    backup_path = None
    should_save = (
        updated_puuid_rows > 0
        or updated_nick_rows > 0
        or created_puuid_col
        or len(created_nick_columns) > 0
        or bool(layout_report.get("layoutChanged"))
        or bool(layout_report.get("createdBaseColumns"))
    )

    if should_save:
        if BACKUP_PLAYERS_XLSX_BEFORE_WRITE:
            backup_path = backup_xlsx(xlsx_path)
        wb.save(xlsx_path)

    return {
        "enabled": True,
        "updatedPuuidRows": updated_puuid_rows,
        "updatedNickRows": updated_nick_rows,
        "createdPuuidColumn": created_puuid_col,
        "createdNickColumns": [canonical_header_name(h) for h in created_nick_columns],
        "layoutChanged": bool(layout_report.get("layoutChanged")),
        "createdBaseColumns": layout_report.get("createdBaseColumns", []),
        "canonicalOrder": layout_report.get("canonicalOrder", []),
        "backupPath": str(backup_path) if backup_path else None,
        "puuidConflicts": puuid_conflicts,
        "nickConflicts": nick_conflicts,
        # Backward-compatible aliases for older print/report code.
        "updatedRows": updated_puuid_rows,
        "conflicts": puuid_conflicts,
    }


def load_players_from_xlsx(
    xlsx_path: Path
) -> Tuple[
    Dict[str, List[str]],
    Dict[str, List[str]],
    Dict[str, Optional[str]],
    Dict[str, Optional[str]],
    Dict[str, int],
    Dict[str, List[str]],
]:
    if not xlsx_path.exists():
        raise RuntimeError(f"Arquivo não encontrado: {xlsx_path}")

    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = build_headers(ws)

    team_cols = get_team_columns(headers)
    if not team_cols:
        raise RuntimeError("players.xlsx precisa ter pelo menos uma coluna 'team N', 'time N' ou 'equipe N'.")

    jogador_col = headers.get("jogador")
    current_team_col = headers.get("current_team")
    puuid_col = headers.get("puuid")

    nick_cols = get_nick_columns(headers)
    alt_puuid_cols = get_alt_puuid_columns(headers)

    if not nick_cols and puuid_col is None:
        raise RuntimeError("players.xlsx precisa ter pelo menos uma coluna 'nick N' ou a coluna 'puuid'.")

    teams_to_players: Dict[str, List[str]] = defaultdict(list)
    player_to_riotids: Dict[str, List[str]] = {}
    player_to_current_team: Dict[str, Optional[str]] = {}
    player_to_existing_puuid: Dict[str, Optional[str]] = {}
    player_to_alt_puuids: Dict[str, List[str]] = {}
    player_key_to_row: Dict[str, int] = {}

    seen_riotids: Set[str] = set()
    seen_puuids: Dict[str, str] = {}

    has_current_team_column = current_team_col is not None

    for row in range(2, ws.max_row + 1):
        team_values = [ws.cell(row=row, column=col).value for col, _header_name in team_cols]
        raw_current_team = ws.cell(row=row, column=current_team_col).value if current_team_col is not None else None
        current_team = resolve_current_team_value(
            current_team_value=raw_current_team,
            team_values=team_values,
            has_current_team_column=has_current_team_column,
        )
        last_historical_team = get_last_historical_team(team_values)

        if has_current_team_column and current_team is not None and last_historical_team is not None:
            if current_team != last_historical_team:
                print(
                    f"[WARNING] current_team diferente do último team N na linha {row}: "
                    f"current_team={current_team!r}, último histórico={last_historical_team!r}. "
                    f"O script vai usar current_team."
                )

        jogador = None
        if jogador_col is not None:
            jv = ws.cell(row=row, column=jogador_col).value
            if jv is not None and str(jv).strip() != "":
                jogador = str(jv).strip()

        player_label = jogador if jogador else "sem_nome"
        player_key = f"{player_label}|row{row}"

        existing_puuid = None
        misplaced_riotid_from_puuid_col: Optional[str] = None
        if puuid_col is not None:
            raw_puuid_cell = normalize_puuid(ws.cell(row=row, column=puuid_col).value)
            if raw_puuid_cell:
                # Às vezes a coluna puuid vem preenchida com Riot ID por engano, ex.: nome#tag.
                # Nesse caso, tratamos o valor como nick/Riot ID e deixamos o script resolver o PUUID correto.
                if normalize_riot_id_for_compare(raw_puuid_cell):
                    gn, tg = parse_riot_id(raw_puuid_cell)
                    misplaced_riotid_from_puuid_col = f"{gn}#{tg}"
                elif is_probably_puuid(raw_puuid_cell):
                    existing_puuid = raw_puuid_cell
                else:
                    print(
                        f"[WARNING] Valor ignorado na coluna puuid para {player_key}: "
                        f"{raw_puuid_cell}. O formato não parece PUUID nem Riot ID."
                    )

        # Contas alternativas do mesmo dono ("puuid 2", "puuid 3", ...). Elas nao
        # viram id de jogador nem sao reescritas pelo script: servem so para
        # reconhecer o dono quando ele entra de conta secundaria.
        alt_puuids: List[str] = []
        for alt_col, alt_header_key in alt_puuid_cols:
            raw_alt = normalize_puuid(ws.cell(row=row, column=alt_col).value)
            if not raw_alt:
                continue
            if not is_probably_puuid(raw_alt):
                print(
                    f"[WARNING] Valor ignorado na coluna {canonical_header_name(alt_header_key)} "
                    f"para {player_key}: {raw_alt}. O formato nao parece PUUID."
                )
                continue
            alt_puuids.append(raw_alt)

        riotids: List[str] = []

        def add_riotid(raw_riotid: str):
            gn, tg = parse_riot_id(raw_riotid)
            riotid_norm = f"{gn}#{tg}"

            riotid_compare = normalize_riot_id_for_compare(riotid_norm)
            if riotid_compare in seen_riotids:
                raise RuntimeError(f"Riot ID duplicado no players.xlsx: {riotid_norm}")
            if riotid_compare:
                seen_riotids.add(riotid_compare)

            riotids.append(riotid_norm)

        if misplaced_riotid_from_puuid_col:
            add_riotid(misplaced_riotid_from_puuid_col)

        for col, _hname in nick_cols:
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is None:
                continue

            s = str(cell_val).strip()
            if not s:
                continue

            add_riotid(s)

        if not existing_puuid and not riotids:
            continue

        if existing_puuid:
            previous_player = seen_puuids.get(existing_puuid)
            if previous_player is not None:
                raise RuntimeError(
                    f"PUUID duplicado no players.xlsx: {existing_puuid} em {previous_player} e {player_key}"
                )
            seen_puuids[existing_puuid] = player_key

        for alt_puuid in alt_puuids:
            previous_player = seen_puuids.get(alt_puuid)
            if previous_player is not None:
                raise RuntimeError(
                    f"PUUID duplicado no players.xlsx: {alt_puuid} em {previous_player} e {player_key}"
                )
            seen_puuids[alt_puuid] = player_key

        player_to_alt_puuids[player_key] = alt_puuids
        player_to_riotids[player_key] = riotids
        player_to_current_team[player_key] = current_team
        player_to_existing_puuid[player_key] = existing_puuid
        player_key_to_row[player_key] = row

        if current_team is not None:
            teams_to_players[current_team].append(player_key)

    if not player_to_existing_puuid:
        raise RuntimeError("players.xlsx não possui linhas válidas com PUUID ou nick(s).")

    return (
        dict(teams_to_players),
        player_to_riotids,
        player_to_current_team,
        player_to_existing_puuid,
        player_key_to_row,
        player_to_alt_puuids,
    )

# =========================
# Rate limiter
# =========================
class MultiWindowRateLimiter:
    """
    Limiter que respeita VÁRIAS janelas ao mesmo tempo.

    Uma requisição só passa quando cabe em todas as janelas configuradas.
    Ex.: com [(20, 1.0), (100, 120.0)], a 21ª chamada dentro de 1s espera, e a
    101ª dentro de 2min também, mesmo que a janela de 1s esteja livre.

    Checagem e registro ficam sob o mesmo lock: sem isso, com vários workers,
    duas threads passavam na checagem ao mesmo tempo e estouravam o teto.
    """

    def __init__(self, limits: List[Tuple[int, float]], name: str):
        self.name = name
        self.windows: List[Dict[str, Any]] = []
        for max_calls, window_s in limits:
            self.windows.append({
                "max": max(1, int(max_calls * RATE_LIMIT_SAFETY)),
                "window": float(window_s),
                "calls": deque(),
            })
        self._lock = Lock()
        self._penalty_until = 0.0
        self._in_flight = 0
        self.waited_s = 0.0
        self.wait_count = 0
        self.sync_ajustes = 0

    def describe(self) -> str:
        return ", ".join(f"{w['max']}/{w['window']:g}s" for w in self.windows)

    def begin_request(self):
        """Marca uma requisicao em voo: saiu daqui, mas a Riot ainda nao contou."""
        with self._lock:
            self._in_flight += 1

    def end_request(self):
        with self._lock:
            self._in_flight = max(0, self._in_flight - 1)

    def observe(self, counts: Dict[float, int]):
        """
        Alinha o contador local ao que a Riot REALMENTE contabilizou.

        Motivo: nos contamos quando a requisicao SAI; a Riot conta quando ela
        CHEGA. Com latencia e concorrencia, chamadas que saem em janelas
        diferentes caem na mesma janela da Riot. Medimos 90 locais contra 93 na
        Riot - a diferenca comia a margem e virava 429 em cascata.

        Sincroniza nos dois sentidos:
          - Riot acima do local  -> completa (fica mais conservador, evita 429);
          - Riot abaixo do local -> apara (a janela fixa da Riot virou, e sem
            isso ficariamos travados por ate 120s sem necessidade).

        `_in_flight` entra na conta porque essas ja sairam mas ainda nao
        apareceram no contador da Riot.
        """
        if not counts:
            return

        with self._lock:
            now = time.monotonic()
            for w in self.windows:
                riot = counts.get(w["window"])
                if riot is None:
                    continue

                calls = w["calls"]
                while calls and (now - calls[0]) >= w["window"]:
                    calls.popleft()

                alvo = int(riot) + self._in_flight
                if len(calls) != alvo:
                    self.sync_ajustes += 1
                while len(calls) < alvo:
                    calls.append(now)
                while len(calls) > alvo and calls:
                    calls.popleft()

    def penalize(self, seconds: float):
        """
        Após um 429, segura TODAS as threads deste limiter, não só a que levou o
        erro. Sem isso os outros workers seguiam martelando a API durante a
        punição e renovavam o bloqueio.
        """
        with self._lock:
            self._penalty_until = max(self._penalty_until, time.monotonic() + float(seconds))

    def acquire(self):
        while True:
            with self._lock:
                now = time.monotonic()
                wait_s = max(0.0, self._penalty_until - now)

                if wait_s <= 0:
                    for w in self.windows:
                        calls = w["calls"]
                        while calls and (now - calls[0]) >= w["window"]:
                            calls.popleft()
                        if len(calls) >= w["max"]:
                            wait_s = max(wait_s, (calls[0] + w["window"]) - now)

                if wait_s <= 0:
                    for w in self.windows:
                        w["calls"].append(now)
                    return

            sleep_s = min(max(wait_s, 0.01), 5.0)
            time.sleep(sleep_s)

            with self._lock:
                self.wait_count += 1
                self.waited_s += sleep_s


# Teto de APP: soma de TODAS as rotas. É o limite que realmente aperta.
app_limiter = MultiWindowRateLimiter(APP_RATE_LIMITS, "app")

# Tetos por método, bem mais folgados; raramente bloqueiam.
account_limiter = MultiWindowRateLimiter(ACCOUNT_METHOD_LIMITS, "account-v1")
val_limiter = MultiWindowRateLimiter(VAL_METHOD_LIMITS, "val-match-v1")

# Telemetria de rate limit, para o resumo no fim da execução.
rate_limit_stats: Dict[str, int] = defaultdict(int)
rate_limit_stats_lock = Lock()


def parse_rate_limit_header(value: Any) -> Dict[float, int]:
    """Converte o formato da Riot "100:120,20:1" em {120.0: 100, 1.0: 20}."""
    out: Dict[float, int] = {}
    for parte in str(value or "").split(","):
        parte = parte.strip()
        if not parte or ":" not in parte:
            continue
        n, janela = parte.split(":", 1)
        try:
            out[float(janela)] = int(n)
        except ValueError:
            continue
    return out


def rate_limit_report() -> Dict[str, Any]:
    with rate_limit_stats_lock:
        stats = dict(rate_limit_stats)
    return {
        "appLimits": app_limiter.describe(),
        "requests": stats.get("requests", 0),
        "http429": stats.get("429", 0),
        "throttleWaitSeconds": round(
            app_limiter.waited_s + account_limiter.waited_s + val_limiter.waited_s, 1
        ),
        "byScope": {k.split(":", 1)[1]: v for k, v in stats.items() if k.startswith("429:")},
        "syncAjustes": app_limiter.sync_ajustes,
    }


# =========================
# Riot client
# =========================
class RiotHttpError(RuntimeError):
    def __init__(self, status_code: int, body: str):
        super().__init__(f"HTTP {status_code}: {body}")
        self.status_code = status_code
        self.body = body


class RiotClient:
    def __init__(self, api_key: str, val_region: str, riot_routing: str):
        if not api_key or not api_key.startswith("RGAPI-"):
            raise RuntimeError(
                "RIOT_API_KEY inválida ou não configurada.\n"
                "Defina a variável de ambiente RIOT_API_KEY, ou escreva\n"
                f"RIOT_API_KEY=RGAPI-... em {config_hub.TRABALHO_DIR / '.env'}"
            )
        self.base_val = f"https://{val_region}.api.riotgames.com"
        self.base_riot = f"https://{riot_routing}.api.riotgames.com"
        self.session = requests.Session()
        self.session.headers.update({"X-Riot-Token": api_key})
        # O pool padrão do requests guarda só 10 conexões por host; com mais
        # workers, conexões extras eram descartadas e reabertas a cada chamada.
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=4,
            pool_maxsize=max(MAX_WORKERS, 10),
        )
        self.session.mount("https://", adapter)

    def _limiter_for_url(self, url: str) -> "MultiWindowRateLimiter":
        return account_limiter if url.startswith(self.base_riot) else val_limiter

    def get_json(self, url: str) -> Any:
        method_limiter = self._limiter_for_url(url)

        for attempt in range(MAX_RETRIES + 1):
            # Método primeiro (quase nunca bloqueia) e app por último, para o
            # slot do teto mais apertado ser consumido imediatamente antes do GET.
            method_limiter.acquire()
            app_limiter.acquire()

            with rate_limit_stats_lock:
                rate_limit_stats["requests"] += 1

            app_limiter.begin_request()
            method_limiter.begin_request()
            try:
                resp = self.session.get(url, timeout=HTTP_TIMEOUT_S)
            finally:
                app_limiter.end_request()
                method_limiter.end_request()

            try:
                # A Riot devolve a contagem real em TODA resposta, inclusive 429.
                # Sincronizar aqui elimina a deriva entre o nosso relogio (saida)
                # e o dela (chegada), que era a origem dos 429.
                app_limiter.observe(parse_rate_limit_header(resp.headers.get("X-App-Rate-Limit-Count")))
                method_limiter.observe(parse_rate_limit_header(resp.headers.get("X-Method-Rate-Limit-Count")))

                text = resp.text

                if 200 <= resp.status_code < 300:
                    return resp.json()

                if (resp.status_code == 429 or (500 <= resp.status_code <= 504)) and attempt < MAX_RETRIES:
                    ra = str(resp.headers.get("Retry-After") or "").strip()
                    if ra.isdigit():
                        wait_s = int(ra)
                    else:
                        wait_s = min(2 ** attempt, 15) + random.random()

                    if resp.status_code == 429:
                        # 429 = o teto já estourou. Segura todas as threads pelo
                        # tempo pedido, senão os outros workers renovam o bloqueio.
                        scope = str(resp.headers.get("X-Rate-Limit-Type") or "desconhecido")
                        with rate_limit_stats_lock:
                            rate_limit_stats["429"] += 1
                            rate_limit_stats["429:" + scope] += 1
                        app_limiter.penalize(wait_s)
                        method_limiter.penalize(wait_s)

                    time.sleep(wait_s)
                    continue

                raise RiotHttpError(resp.status_code, text)

            except (requests.Timeout, requests.ConnectionError) as e:
                if attempt < MAX_RETRIES:
                    wait_s = min(0.5 * (2 ** attempt), 5.0) + random.random() * 0.25
                    time.sleep(wait_s)
                    continue
                raise e

        raise RuntimeError("Unreachable")

    def get_account_by_riot_id(self, game_name: str, tag_line: str) -> Any:
        gn = quote(game_name, safe="")
        tg = quote(tag_line, safe="")
        return self.get_json(f"{self.base_riot}/riot/account/v1/accounts/by-riot-id/{gn}/{tg}")

    def get_account_by_puuid(self, puuid: str) -> Any:
        return self.get_json(f"{self.base_riot}/riot/account/v1/accounts/by-puuid/{puuid}")

    def get_matchlist_by_puuid(self, puuid: str) -> Any:
        return self.get_json(f"{self.base_val}/val/match/v1/matchlists/by-puuid/{puuid}")

    def get_match(self, match_id: str) -> Any:
        return self.get_json(f"{self.base_val}/val/match/v1/matches/{match_id}")


# =========================
# Main
# =========================

# Sem argumentos para histórico: PegaCustomMatch sempre usa todos os jogos finalizados.

def imprime_cadastrados_sem_puuid(unresolved_players: List[Dict[str, Any]]) -> None:
    """
    Bloco proprio para quem esta na planilha e nao tem PUUID resolvido.

    Antes cada um saia como um [WARNING] solto no meio da execucao, misturado
    com erro de verdade. Quase sempre e reserva que nunca entrou em quadra: o
    nick cadastrado nao existe mais, e nao ha partida de onde tirar o PUUID.
    Juntos e no fim, viram uma lista de conferencia em vez de ruido.
    """
    if not unresolved_players:
        print("Cadastrados sem PUUID: nenhum")
        return

    com_time = [p for p in unresolved_players if p.get("currentTeam")]
    sem_time = [p for p in unresolved_players if not p.get("currentTeam")]

    print()
    print("-" * 72)
    print(f"CADASTRADOS SEM PUUID: {len(unresolved_players)}")
    print("Nenhum Riot ID salvo resolveu na Riot. Quase sempre e reserva que nunca")
    print("entrou em quadra, ou conta renomeada. Se a pessoa jogou alguma partida")
    print("registrada, o PUUID esta no JSON dela e basta copiar para a planilha.")
    for rotulo, grupo in (("com current_team", com_time), ("sem current_team", sem_time)):
        if not grupo:
            continue
        print()
        print(f"  {rotulo} ({len(grupo)}):")
        for p in sorted(grupo, key=lambda x: (str(x.get("currentTeam") or ""), x["playerKey"])):
            nome = p["playerKey"].split("|")[0]
            nicks = ", ".join(p.get("riotIdsTried") or []) or "(sem nick cadastrado)"
            print(f"    {nome:22} {str(p.get('currentTeam') or '-'):24} {nicks}")
    print("-" * 72)
    print()


def main():
    base_dir = _base_dir()
    output_dir = base_dir / "out_tournament"

    ensure_dirs(output_dir)

    xlsx_path = Path(PLAYERS_XLSX_PATH) if PLAYERS_XLSX_PATH else base_dir / PLAYERS_XLSX_NAME
    print("players.xlsx:", xlsx_path)

    (
        teams_to_players,
        player_to_riotids,
        player_to_current_team,
        player_to_existing_puuid,
        player_key_to_row,
        player_to_alt_puuids,
    ) = load_players_from_xlsx(xlsx_path)

    history_dir = base_dir / "Finalizados"
    matches_dir = output_dir / "matches"

    history_dir.mkdir(parents=True, exist_ok=True)

    # Índices por matchId lido de dentro dos JSONs.
    # - history_index: partidas já arquivadas fora da pasta de saída atual.
    # - local_match_index: partidas já existentes em out_tournament/matches, mesmo que já estejam renomeadas.
    #
    # Antes de rodar a coleta, limpamos duplicatas locais baseadas em matchInfo.matchId.
    history_index = load_existing_match_index_all_history(history_dir)
    duplicate_backup_dir = output_dir / DUPLICATE_MATCHES_DIR_NAME
    local_match_index, initial_duplicate_report = dedupe_matches_dir_by_match_id(matches_dir, duplicate_backup_dir)
    existing_lock = Lock()

    print("Base dir:", base_dir)
    print("History dir:", history_dir)
    print("History source: todos os .json recursivamente em Finalizados")
    print("Já existentes em /Finalizados:", len(history_index))
    print("Já existentes em /out_tournament/matches:", len(local_match_index))
    print("Duplicatas removidas de /out_tournament/matches:", len(initial_duplicate_report))
    if REQUIRE_CUSTOM_BOMB_GAME:
        print("Filtro de tipo de partida: provisioningFlowId=CustomGame + gameMode=Bomb")

    client = RiotClient(RIOT_API_KEY, VAL_REGION, RIOT_ROUTING)

    print("Times carregados do Excel:", len(teams_to_players))
    print("Jogadores carregados do Excel:", len(player_to_riotids))

    # 1) Resolve all players -> puuids (try nick1, then nick2, etc.)
    player_key_to_puuid: Dict[str, str] = {}
    player_key_to_current_riotid: Dict[str, str] = {}
    puuid_to_team: Dict[str, str] = {}
    unresolved_players: List[Dict[str, Any]] = []

    def resolve_player(player_key: str) -> Tuple[Optional[str], str, Optional[str], Optional[str], Optional[str]]:
        current_team = player_to_current_team[player_key]

        existing_puuid = player_to_existing_puuid.get(player_key)
        invalid_existing_puuid_warning: Optional[str] = None

        if existing_puuid and not REFRESH_CURRENT_RIOT_IDS:
            # Usa o PUUID salvo sem gastar Account-V1. Se estiver inválido (raro),
            # a busca de matchlist detecta o 400/404 e re-resolve pelo Riot ID.
            return current_team, player_key, existing_puuid, None, None

        if existing_puuid:
            try:
                acc = client.get_account_by_puuid(existing_puuid)
                current_riot_id = account_to_riot_id(acc)
                return current_team, player_key, existing_puuid, current_riot_id, None
            except RiotHttpError as e:
                # 400 costuma significar PUUID salvo inválido/não confiável para a rota Account-V1
                # (ex.: valor colado errado, ID de outro tipo, célula corrompida).
                # 404 significa que a Riot não encontrou aquele PUUID.
                # Em ambos os casos, não derrubamos a execução: tentamos resolver pelo Riot ID salvo.
                if e.status_code in (400, 404):
                    invalid_existing_puuid_warning = (
                        f"[WARNING] PUUID salvo inválido/não encontrado para {player_key}: "
                        f"{existing_puuid}. Vou tentar resolver pelo Riot ID."
                    )
                else:
                    raise

        tried: List[str] = []

        for riotid in player_to_riotids.get(player_key, []):
            tried.append(riotid)
            gn, tg = parse_riot_id(riotid)
            try:
                acc = client.get_account_by_riot_id(gn, tg)
                puuid = normalize_puuid(acc.get("puuid"))
                current_riot_id = account_to_riot_id(acc) or riotid
                if puuid:
                    if invalid_existing_puuid_warning:
                        warning = (
                            f"{invalid_existing_puuid_warning} Resolvido pelo Riot ID "
                            f"{riotid}; o PUUID será atualizado no players.xlsx."
                        )
                        return current_team, player_key, puuid, current_riot_id, warning
                    return current_team, player_key, puuid, current_riot_id, None
            except RiotHttpError as e:
                if e.status_code == 404:
                    continue
                raise

        if invalid_existing_puuid_warning:
            warning = (
                f"{invalid_existing_puuid_warning} Não encontrei novo PUUID. "
                f"Riot IDs tentados: {tried}"
            )
        else:
            warning = f"[WARNING] Não encontrei PUUID para {player_key}. Riot IDs tentados: {tried}"

        return current_team, player_key, None, None, warning

    all_player_keys = list(player_to_riotids.keys())

    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(all_player_keys))) as ex:
        futures = [ex.submit(resolve_player, pk) for pk in all_player_keys]
        for f in as_completed(futures):
            current_team, pk, puuid, current_riot_id, warning = f.result()

            # Jogador sem PUUID nao vira WARNING solto no meio da execucao:
            # quase sempre e reserva que nunca entrou em quadra, e misturado com
            # erro de verdade vira ruido. Vai todo mundo para o bloco do fim.
            if puuid is None:
                unresolved_players.append({
                    "playerKey": pk,
                    "riotIdsTried": player_to_riotids.get(pk, []),
                    "currentTeam": current_team,
                    "detalhe": warning or "sem detalhe",
                })
                continue

            # PUUID salvo invalido que foi recuperado pelo Riot ID: isso sim e
            # acionavel na hora, porque a planilha vai ser corrigida.
            if warning:
                print(warning)

            player_key_to_puuid[pk] = puuid
            if current_riot_id:
                player_key_to_current_riotid[pk] = current_riot_id
            if current_team is not None:
                puuid_to_team[puuid] = current_team

    puuid_to_first_player: Dict[str, str] = {}
    duplicate_resolved_puuids: List[Dict[str, str]] = []
    for pk, puuid in player_key_to_puuid.items():
        first = puuid_to_first_player.get(puuid)
        if first is not None:
            duplicate_resolved_puuids.append({"puuid": puuid, "firstPlayer": first, "duplicatePlayer": pk})
        else:
            puuid_to_first_player[puuid] = pk

    if duplicate_resolved_puuids:
        save_json(output_dir / "duplicate_resolved_puuids.json", duplicate_resolved_puuids)
        raise RuntimeError(
            "PUUID duplicado após resolução por nick. Veja out_tournament/duplicate_resolved_puuids.json"
        )

    puuid_write_report = write_resolved_puuids_to_xlsx(
        xlsx_path=xlsx_path,
        player_key_to_row=player_key_to_row,
        player_key_to_puuid=player_key_to_puuid,
        player_key_to_current_riotid=player_key_to_current_riotid,
    )

    print("Jogadores com PUUID resolvido:", len(player_key_to_puuid))
    imprime_cadastrados_sem_puuid(unresolved_players)
    print("PUUIDs anotados no players.xlsx:", puuid_write_report.get("updatedPuuidRows", 0))
    print("Riot IDs atuais anotados no players.xlsx:", puuid_write_report.get("updatedNickRows", 0))
    if puuid_write_report.get("createdPuuidColumn"):
        print("Coluna 'puuid' criada no players.xlsx.")
    if puuid_write_report.get("createdNickColumns"):
        print("Colunas de nick criadas no players.xlsx:", ", ".join(puuid_write_report["createdNickColumns"]))
    if puuid_write_report.get("layoutChanged"):
        print("Layout do players.xlsx padronizado: Jogador | team 1 | team 2 | team 3 | ... | current_team | puuid | nick 1 | nick 2 | ...")
    if puuid_write_report.get("backupPath"):
        print("Backup do players.xlsx:", puuid_write_report["backupPath"])
    if puuid_write_report.get("puuidConflicts"):
        print("Conflitos de PUUID no Excel:", len(puuid_write_report["puuidConflicts"]))
    if puuid_write_report.get("nickConflicts"):
        print("Conflitos de Riot ID no Excel:", len(puuid_write_report["nickConflicts"]))

    if not player_key_to_puuid:
        raise RuntimeError("Nenhum jogador teve PUUID resolvido. Não é possível continuar.")

    puuid_to_player_key: Dict[str, str] = {v: k for k, v in player_key_to_puuid.items()}

    save_json(output_dir / "resolved_players.json", {
        "teams_to_players": teams_to_players,
        "player_to_riotids": player_to_riotids,
        "player_to_current_team": player_to_current_team,
        "player_to_existing_puuid": player_to_existing_puuid,
        "player_key_to_puuid": player_key_to_puuid,
        "player_key_to_current_riotid": player_key_to_current_riotid,
        "puuid_write_report": puuid_write_report,
        "unresolved_players": unresolved_players,
    })

    if ONLY_CURRENT_TEAM_PLAYERS_FOR_MATCH_DISCOVERY:
        active_player_key_to_puuid = {
            pk: puuid
            for pk, puuid in player_key_to_puuid.items()
            if player_to_current_team.get(pk) is not None
        }
    else:
        active_player_key_to_puuid = dict(player_key_to_puuid)

    all_puuids = list(active_player_key_to_puuid.values())
    all_puuid_set: Set[str] = set(all_puuids)

    if not all_puuids:
        raise RuntimeError(
            "Nenhum jogador elegível para busca de partidas. "
            "Verifique current_team ou desative ONLY_CURRENT_TEAM_PLAYERS_FOR_MATCH_DISCOVERY."
        )

    print("Jogadores elegíveis para busca de partidas:", len(all_puuids))

    # As contas alternativas entram no RECONHECIMENTO, nao na busca: a matchlist
    # continua sendo puxada so pela conta principal, entao isto nao custa
    # chamada nenhuma a mais na Riot. Sem isto, quem entra de conta secundaria
    # nao conta para MIN_TEAM_PLAYERS_TO_IDENTIFY e a equipe dele sai como
    # "unknown" no nome do arquivo - foi o que aconteceu com a A2E UFF no JUBS
    # presencial (3 dos 5 jogadores estavam de conta nova).
    contas_alternativas = 0
    for pk, alt_puuids in player_to_alt_puuids.items():
        if pk not in active_player_key_to_puuid:
            continue
        time_atual = player_to_current_team.get(pk)
        for alt_puuid in alt_puuids:
            if alt_puuid in all_puuid_set:
                continue
            all_puuid_set.add(alt_puuid)
            puuid_to_player_key.setdefault(alt_puuid, pk)
            if time_atual is not None:
                puuid_to_team.setdefault(alt_puuid, time_atual)
            contas_alternativas += 1

    print("Contas alternativas reconhecidas (colunas puuid N):", contas_alternativas)

    save_json(output_dir / "active_players_for_discovery.json", {
        pk: {
            "puuid": puuid,
            "currentTeam": player_to_current_team.get(pk),
        }
        for pk, puuid in active_player_key_to_puuid.items()
    })

    # 2) Fetch matchlists for all eligible players
    match_to_players_seen: Dict[str, Set[str]] = defaultdict(set)

    def refetch_puuid_for_player(pk: str) -> Optional[str]:
        """Re-resolve o PUUID de um jogador pelos Riot IDs salvos na planilha."""
        for riotid in player_to_riotids.get(pk, []):
            gn, tg = parse_riot_id(riotid)
            try:
                acc = client.get_account_by_riot_id(gn, tg)
            except RiotHttpError as e:
                if e.status_code == 404:
                    continue
                raise
            new_puuid = normalize_puuid(acc.get("puuid"))
            if new_puuid:
                return new_puuid
        return None

    def fetch_matchlist(puuid: str) -> Tuple[str, str, List[Dict[str, Any]], Optional[str]]:
        """
        Busca a matchlist e retorna (puuid_original, puuid_efetivo, history, warning).

        Com REFRESH_CURRENT_RIOT_IDS=False os PUUIDs salvos não passam pela Account-V1;
        se algum estiver inválido (400/404 aqui), re-resolve pelo Riot ID e segue com o novo.
        """
        effective_puuid = puuid
        warning: Optional[str] = None

        try:
            ml = client.get_matchlist_by_puuid(puuid)
        except RiotHttpError as e:
            if e.status_code not in (400, 404):
                raise

            pk = puuid_to_player_key.get(puuid)
            new_puuid = refetch_puuid_for_player(pk) if pk else None

            if not new_puuid or new_puuid == puuid:
                warning = (
                    f"[WARNING] Matchlist indisponível (HTTP {e.status_code}) para {pk or puuid} "
                    f"e não consegui re-resolver pelo Riot ID. Jogador ignorado na descoberta."
                )
                return puuid, puuid, [], warning

            warning = (
                f"[WARNING] PUUID salvo inválido para {pk}; re-resolvido pelo Riot ID nesta execução. "
                f"Corrija a coluna puuid no players.xlsx (novo: {new_puuid})."
            )
            ml = client.get_matchlist_by_puuid(new_puuid)
            effective_puuid = new_puuid

        hist = ml.get("history") or []
        if not isinstance(hist, list):
            hist = []

        hist = [h for h in hist if is_in_date_range(h.get("gameStartTimeMillis"))]

        return puuid, effective_puuid, hist[:LOOKBACK_PER_PLAYER], warning

    puuid_corrections: List[Dict[str, str]] = []

    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(all_puuids))) as ex:
        futures = [ex.submit(fetch_matchlist, p) for p in all_puuids]
        for f in as_completed(futures):
            original_puuid, puuid, hist, warning = f.result()
            if warning:
                print(warning)
            if puuid != original_puuid:
                puuid_corrections.append({"oldPuuid": original_puuid, "newPuuid": puuid})
            save_json(output_dir / "matchlists" / f"{puuid}.json", {"puuid": puuid, "history": hist})
            for h in hist:
                mid = h.get("matchId")
                if isinstance(mid, str) and mid:
                    match_to_players_seen[mid].add(puuid)

    # Aplica as correções de PUUID nos índices usados pela validação/detecção de times.
    # Feito após o pool encerrar para não mutar os dicionários com threads ativas.
    for corr in puuid_corrections:
        old, new = corr["oldPuuid"], corr["newPuuid"]
        pk = puuid_to_player_key.pop(old, None)
        if pk is None:
            continue
        corr["playerKey"] = pk
        if new in puuid_to_player_key:
            print(f"[WARNING] PUUID corrigido de {pk} colide com outro jogador; {pk} fica fora da detecção.")
            all_puuid_set.discard(old)
            puuid_to_team.pop(old, None)
            continue
        puuid_to_player_key[new] = pk
        player_key_to_puuid[pk] = new
        all_puuid_set.discard(old)
        all_puuid_set.add(new)
        team = puuid_to_team.pop(old, None)
        if team is not None:
            puuid_to_team[new] = team

    if puuid_corrections:
        save_json(output_dir / "puuid_corrections.json", puuid_corrections)

    # 3) Candidate matches by matchlist overlap
    candidates = sorted(
        [(mid, len(puuids)) for mid, puuids in match_to_players_seen.items() if len(puuids) >= MIN_PLAYERS_IN_MATCHLIST],
        key=lambda x: x[1],
        reverse=True,
    )

    candidate_match_ids = [mid for mid, _ in candidates][:MAX_MATCHES_TO_FETCH]
    # Remove apenas o que já existe no histórico externo.
    # Se também existir em out_tournament/matches, mantém o candidato para permitir renomeação local.
    candidate_match_ids = [
        mid for mid in candidate_match_ids
        if mid not in history_index or mid in local_match_index
    ]

    save_json(
        output_dir / "candidates_by_matchlist_overlap.json",
        [{"matchId": mid, "matchlistOverlapPlayers": cnt} for mid, cnt in candidates],
    )

    print("Players in tournament/discovery pool:", len(all_puuids))
    print(f"Candidates (matchlist overlap >= {MIN_PLAYERS_IN_MATCHLIST}):", len(candidate_match_ids))
    print("Candidates após dedupe:", len(candidate_match_ids))

    # 4) Download each candidate match, confirm in payload, detect teams, and save
    report: List[Dict[str, Any]] = []
    rejected: List[Dict[str, Any]] = []

    def detect_teams_in_match(match: Dict[str, Any]) -> List[Dict[str, Any]]:
        players = match.get("players") or []
        if not isinstance(players, list):
            return []

        active_players = [
            p for p in players
            if isinstance(p, dict) and not bool(p.get("isObserver", False))
        ]

        group: Dict[Tuple[str, str], List[str]] = defaultdict(list)
        for p in active_players:
            puuid = p.get("puuid")
            team_id = p.get("teamId")

            if not isinstance(puuid, str) or puuid not in all_puuid_set:
                continue
            if not isinstance(team_id, str):
                continue

            tname = puuid_to_team.get(puuid)
            if not tname:
                continue

            group[(tname, team_id)].append(puuid)

        detected: List[Dict[str, Any]] = []
        for (tname, team_id), puuids in group.items():
            unique = sorted(set(puuids))
            if len(unique) >= MIN_TEAM_PLAYERS_TO_IDENTIFY:
                players_readable = [puuid_to_player_key.get(p, p) for p in unique]
                detected.append({
                    "teamName": tname,
                    "teamId": team_id,
                    "players": sorted(players_readable),
                    "playerCount": len(unique),
                })

        return sorted(detected, key=lambda x: (x["teamName"], -x["playerCount"]))

    def build_match_record_from_payload(
        mid: str,
        match: Dict[str, Any],
        source: str,
        current_file_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Valida uma partida e monta o registro usado no report/renomeação.
        Serve tanto para partidas recém-baixadas quanto para partidas já presentes em out_tournament/matches.
        """
        mi = match.get("matchInfo") or {}
        if not is_in_date_range(mi.get("gameStartMillis")):
            return {
                "matchId": mid,
                "saved": False,
                "source": source,
                "currentFileName": current_file_name,
                "reason": "outside_date_range",
                **match_type_info(match),
            }

        if REQUIRE_CUSTOM_BOMB_GAME and not is_custom_bomb_game(match):
            return reject_non_custom_bomb_record(
                mid=mid,
                match=match,
                source=source,
                current_file_name=current_file_name,
            )

        players = match.get("players") or []
        if not isinstance(players, list):
            players = []

        active = [
            p for p in players
            if isinstance(p, dict) and not bool(p.get("isObserver", False))
        ]

        active_players_by_side = {"Blue": 0, "Red": 0}
        for p in active:
            side = p.get("teamId")
            if side in active_players_by_side:
                active_players_by_side[str(side)] += 1

        is_completed = bool(mi.get("isCompleted", False))
        blue_active_count = active_players_by_side["Blue"]
        red_active_count = active_players_by_side["Red"]

        if (
            REQUIRE_COMPLETED_MATCH
            and (
                not is_completed
                or blue_active_count < MIN_ACTIVE_PLAYERS_PER_SIDE
                or red_active_count < MIN_ACTIVE_PLAYERS_PER_SIDE
            )
        ):
            return {
                "matchId": mid,
                "saved": False,
                "source": source,
                "currentFileName": current_file_name,
                "reason": (
                    f"incomplete_match("
                    f"isCompleted={is_completed},"
                    f"Blue={blue_active_count},Red={red_active_count},"
                    f"minimum={MIN_ACTIVE_PLAYERS_PER_SIDE})"
                ),
                "isCompleted": mi.get("isCompleted"),
                "activePlayersBySide": {
                    "Blue": blue_active_count,
                    "Red": red_active_count,
                },
                **match_type_info(match),
            }

        rounds_won_by_side = get_rounds_won_by_side(match)
        winning_rounds = get_winning_rounds(match)
        if REQUIRE_WINNER_MIN_ROUNDS and winning_rounds < MIN_WINNING_ROUNDS:
            return {
                "matchId": mid,
                "saved": False,
                "source": source,
                "currentFileName": current_file_name,
                "reason": (
                    f"winner_below_min_rounds("
                    f"winnerRounds={winning_rounds},"
                    f"minimum={MIN_WINNING_ROUNDS},"
                    f"roundsBySide={rounds_won_by_side})"
                ),
                "roundsWonBySide": rounds_won_by_side,
                "winningRounds": winning_rounds,
                "minimumWinningRounds": MIN_WINNING_ROUNDS,
                **match_type_info(match),
            }

        tournament_puuids_in_match = {
            p.get("puuid") for p in active
            if isinstance(p.get("puuid"), str) and p.get("puuid") in all_puuid_set
        }
        in_match_count = len(tournament_puuids_in_match)

        if in_match_count < MIN_PLAYERS_IN_MATCH_PAYLOAD:
            return {
                "matchId": mid,
                "saved": False,
                "source": source,
                "currentFileName": current_file_name,
                "reason": f"payload_has_{in_match_count}_tournament_players",
            }

        known_players_by_side: Dict[str, Set[str]] = defaultdict(set)
        for p in active:
            puuid = p.get("puuid")
            team_id = p.get("teamId")

            if not isinstance(puuid, str) or puuid not in all_puuid_set:
                continue
            if team_id not in ("Blue", "Red"):
                continue

            known_players_by_side[str(team_id)].add(puuid)

        known_blue_count = len(known_players_by_side.get("Blue", set()))
        known_red_count = len(known_players_by_side.get("Red", set()))

        if known_blue_count < MIN_PLAYERS_PER_SIDE_IN_PAYLOAD or known_red_count < MIN_PLAYERS_PER_SIDE_IN_PAYLOAD:
            return {
                "matchId": mid,
                "saved": False,
                "source": source,
                "currentFileName": current_file_name,
                "reason": (
                    f"payload_side_split_below_minimum("
                    f"Blue={known_blue_count},Red={known_red_count},"
                    f"minimum={MIN_PLAYERS_PER_SIDE_IN_PAYLOAD})"
                ),
                "tournamentPlayersInMatch": in_match_count,
                "knownPlayersBySide": {
                    "Blue": known_blue_count,
                    "Red": known_red_count,
                },
                "activePlayersBySide": {
                    "Blue": blue_active_count,
                    "Red": red_active_count,
                },
            }

        detected_teams = detect_teams_in_match(match)
        if not detected_teams:
            return {
                "matchId": mid,
                "saved": False,
                "source": source,
                "currentFileName": current_file_name,
                "reason": "no teams detected with >=3 players",
            }

        # Placar extraído aqui, com o payload já em memória, para a etapa de
        # renomeação não precisar reler o arquivo do disco.
        score_a, score_b = extract_match_score(match=match, detected_teams=detected_teams)

        return {
            "matchId": mid,
            "saved": True,
            "scoreA": score_a,
            "scoreB": score_b,
            "source": source,
            "currentFileName": current_file_name,
            "tournamentPlayersInMatch": in_match_count,
            "knownPlayersBySide": {
                "Blue": known_blue_count,
                "Red": known_red_count,
            },
            "activePlayersBySide": {
                "Blue": blue_active_count,
                "Red": red_active_count,
            },
            "gameStartMillis": mi.get("gameStartMillis"),
            "mapId": mi.get("mapId"),
            "provisioningFlowId": mi.get("provisioningFlowId"),
            "queueId": mi.get("queueId"),
            "gameMode": mi.get("gameMode"),
            "isCompleted": mi.get("isCompleted"),
            "customGameName": mi.get("customGameName"),
            "roundsWonBySide": rounds_won_by_side,
            "winningRounds": winning_rounds,
            "detectedTeams": detected_teams,
        }

    def process_match(mid: str) -> Dict[str, Any]:
        with existing_lock:
            local_file_name = local_match_index.get(mid)

        if local_file_name:
            local_path = matches_dir / local_file_name
            try:
                local_match = json.loads(local_path.read_text(encoding="utf-8"))
            except Exception as e:
                return {
                    "matchId": mid,
                    "saved": False,
                    "reason": f"failed_to_read_existing_local_file({local_file_name}): {e}",
                }

            return build_match_record_from_payload(
                mid=mid,
                match=local_match,
                source="existing_local",
                current_file_name=local_file_name,
            )

        with existing_lock:
            if mid in history_index:
                return {
                    "matchId": mid,
                    "saved": False,
                    "skipped": True,
                    "reason": f"already_in_history({history_index[mid]})",
                }

        out_path = matches_dir / f"{mid}.json"
        if out_path.exists():
            try:
                existing_match = json.loads(out_path.read_text(encoding="utf-8"))
            except Exception as e:
                return {
                    "matchId": mid,
                    "saved": False,
                    "reason": f"failed_to_read_existing_matchid_file({out_path.name}): {e}",
                }

            with existing_lock:
                local_match_index[mid] = out_path.name

            return build_match_record_from_payload(
                mid=mid,
                match=existing_match,
                source="existing_local",
                current_file_name=out_path.name,
            )

        match = client.get_match(mid)
        rec = build_match_record_from_payload(
            mid=mid,
            match=match,
            source="downloaded_new",
            current_file_name=out_path.name,
        )

        if rec.get("saved"):
            save_json(out_path, match)
            with existing_lock:
                local_match_index[mid] = out_path.name

        return rec

    with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, max(1, len(candidate_match_ids)))) as ex:
        futures = [ex.submit(process_match, mid) for mid in candidate_match_ids]
        for f in as_completed(futures):
            rec = f.result()
            if rec.get("skipped"):
                continue
            if rec.get("saved"):
                report.append(rec)
            else:
                rejected.append(rec)

    # mais antigo -> mais recente
    report_sorted = sorted(report, key=lambda r: (r.get("gameStartMillis") or 0))

    # 5) Renomeia os arquivos aceitos em matches/
    # Inclui tanto arquivos recém-baixados quanto arquivos já existentes localmente.
    rename_report: List[Dict[str, Any]] = []

    for idx, rec in enumerate(report_sorted, start=1):
        mid = rec.get("matchId")
        if not mid:
            continue

        current_file_name = rec.get("currentFileName") or f"{mid}.json"
        old_path = matches_dir / str(current_file_name)
        if not old_path.exists():
            rejected.append({
                "matchId": mid,
                "saved": False,
                "reason": f"file_to_rename_not_found({current_file_name})",
            })
            continue

        score_a = rec.get("scoreA")
        score_b = rec.get("scoreB")

        new_name = build_match_filename(
            order_num=idx,
            game_start_ms=rec.get("gameStartMillis"),
            detected_teams=rec.get("detectedTeams") or [],
            score_a=score_a,
            score_b=score_b,
            map_id=rec.get("mapId"),
            match_id=mid,
        )
        new_path = matches_dir / new_name

        if new_path.exists() and new_path.resolve() != old_path.resolve():
            existing_mid_at_target = read_match_id_from_json_file(new_path)

            if existing_mid_at_target == mid:
                # Mesmo matchId no destino: é duplicata. Remove o destino e preserva/move o arquivo atual.
                initial_duplicate_report.append(
                    remove_duplicate_file(
                        new_path,
                        duplicate_backup_dir,
                        reason=f"duplicate_matchId({mid}) during_rename; kept({old_path.name})",
                    )
                )
            else:
                # Colisão de nome com outro matchId. Não sobrescreve; cria sufixo para preservar ambos.
                stem = new_path.stem
                suffix = new_path.suffix
                counter = 2
                while True:
                    candidate = matches_dir / f"{stem}_{counter}{suffix}"
                    if not candidate.exists():
                        new_path = candidate
                        break
                    counter += 1

        if new_path.resolve() != old_path.resolve():
            shutil.move(str(old_path), str(new_path))

        with existing_lock:
            local_match_index[str(mid)] = new_path.name

        rec["fileName"] = new_path.name
        rec["matchDate"] = format_match_date_for_filename(rec.get("gameStartMillis"))
        rec["scoreA"] = score_a
        rec["scoreB"] = score_b

        rename_report.append({
            "matchId": mid,
            "source": rec.get("source"),
            "oldFileName": old_path.name,
            "newFileName": new_path.name,
            "matchDate": format_match_date_for_filename(rec.get("gameStartMillis")),
            "scoreA": score_a,
            "scoreB": score_b,
        })

    save_json(output_dir / "report.json", report_sorted)
    save_json(output_dir / "rejected.json", rejected)
    save_json(output_dir / "renamed_matches.json", rename_report)
    save_json(output_dir / "duplicate_matches_removed.json", initial_duplicate_report)

    rejected_not_custom_bomb = sum(
        1 for r in rejected
        if str(r.get("reason") or "").startswith("not_custom_bomb_game")
    )
    rejected_incomplete = sum(
        1 for r in rejected
        if str(r.get("reason") or "").startswith("incomplete_match")
    )
    rejected_winner_below_min_rounds = sum(
        1 for r in rejected
        if str(r.get("reason") or "").startswith("winner_below_min_rounds")
    )

    print("Collected matches:", len(report_sorted))
    print("Rejected candidates:", len(rejected))
    print("Rejected non-custom/non-bomb:", rejected_not_custom_bomb)
    print("Rejected incomplete/under-5v5:", rejected_incomplete)
    print("Rejected winner below 13 rounds:", rejected_winner_below_min_rounds)
    print("Renamed matches:", len(rename_report))
    print("Duplicate files removed from matches:", len(initial_duplicate_report))

    rl = rate_limit_report()
    save_json(output_dir / "rate_limit_report.json", rl)
    print(
        f"Rate limit: {rl['requests']} requisições | "
        f"{rl['http429']} x HTTP 429 | "
        f"{rl['throttleWaitSeconds']}s aguardando o limiter (teto de app: {rl['appLimits']})"
    )
    if rl["http429"]:
        print(
            "[WARNING] Houve 429. Se persistir, reduza RATE_LIMIT_SAFETY "
            "ou confirme se APP_RATE_LIMITS bate com o teto da sua chave."
        )

    print("Output:", str(output_dir))


if __name__ == "__main__":
    main()