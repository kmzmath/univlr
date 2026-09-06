#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Audita nicks/contas que aparecem em JSONs de partidas, mas não estão cadastrados
em players.xlsx e/ou não pertencem ao roster atual em teams.xlsx.

Entrada esperada, por padrão, no mesmo diretório do script:
  - players.xlsx
  - teams.xlsx
  - Finalizados/**/*.json

Saída:
  - out_auditoria_nicks/nicks_fora_do_cadastro_resumo.csv
  - out_auditoria_nicks/nicks_fora_do_cadastro_detalhado.csv
  - out_auditoria_nicks/nicks_fora_do_cadastro_resumo.json
  - out_auditoria_nicks/team_detection_debug.json

Uso:
  python AuditaNicksPartidas.py

Ou informando caminhos:
  python AuditaNicksPartidas.py --players players.xlsx --teams teams.xlsx --matches-dir Finalizados
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import load_workbook

import config_hub


# =========================
# Configurações
# =========================

MIN_TEAM_PLAYERS_TO_IDENTIFY = 3

# A auditoria compara a equipe da partida com TODO o histórico do jogador
# (colunas team 1, team 2, ... mais o current_team), não só com o current_team.
# Jogar por uma equipe que já está no próprio histórico não é erro: é o caso de
# quem defende uma equipe da organização num campeonato e outra no resto do ano.
#
# Os dois achados que importam não têm flag para desligar:
#   nunca_cadastrado         - jogou e não tem linha no players.xlsx
#   equipe_fora_do_historico - jogou por equipe que não está em nenhum team N
#                              nem no current_team (transferência não registrada,
#                              ou estamos colocando a pessoa na equipe errada)

# Jogador com linha no players.xlsx mas com current_team vazio.
REPORT_SEM_TIME_ATUAL = True

# current_team bate com a equipe da partida, mas o nome não está no roster
# daquela equipe no teams.xlsx (as duas planilhas discordando entre si).
REPORT_FORA_DO_ROSTER_TEAMS = True

# Jogou por uma equipe do próprio histórico que não é o current_team. É ruído
# conhecido e fica desligado; ligue para revisar o histórico de propósito.
REPORT_EQUIPE_DO_HISTORICO_NAO_ATUAL = False

# Completou por outra equipe: entrou para fechar a line, sem fazer parte do
# elenco. Fica registrado na coluna `completes` do players.xlsx e NÃO entra no
# histórico - do contrário inflaria a profundidade de elenco daquela equipe no
# ranking e registraria uma line que nunca existiu. Desligado porque é fato
# conhecido e conferido, não pendência.
REPORT_COMPLETE_POR_OUTRA_EQUIPE = False

# Pessoa que está no elenco de duas equipes ao mesmo tempo - o caso da line
# inclusiva, em que a jogadora defende a line principal da organização no
# circuito normal e a inclusiva no campeonato feminino. O elenco duplo mora no
# teams.xlsx, que já é a lista de quem está em cada equipe; o current_team
# continua único, porque é ele que o resto do pipeline lê como chave. Desligado
# porque é arranjo conhecido, não pendência.
REPORT_ELENCO_DE_DUAS_EQUIPES = False

# Severidade de cada achado, na ordem em que o relatório é ordenado.
SEVERIDADE_POR_ISSUE = {
    "nunca_cadastrado": "ALTA",
    "equipe_fora_do_historico": "ALTA",
    "sem_time_atual": "MEDIA",
    "fora_do_roster_do_teams": "MEDIA",
    "equipe_do_historico_nao_atual": "BAIXA",
    "complete_por_outra_equipe": "BAIXA",
    "elenco_de_duas_equipes": "BAIXA",
}
ORDEM_SEVERIDADE = {"ALTA": 0, "MEDIA": 1, "BAIXA": 2}

DESCRICAO_POR_ISSUE = {
    "nunca_cadastrado": "jogou e não tem linha no players.xlsx",
    "equipe_fora_do_historico": "jogou por equipe que não está no histórico dele",
    "sem_time_atual": "tem linha no players.xlsx, mas sem current_team",
    "fora_do_roster_do_teams": "current_team bate, mas o nome não está no roster do teams.xlsx",
    "equipe_do_historico_nao_atual": "jogou por equipe do histórico que não é a atual",
    "complete_por_outra_equipe": "completou por outra equipe, registrado em `completes`",
    "elenco_de_duas_equipes": "está no elenco desta equipe e de outra no teams.xlsx",
}


def severidade_do_conjunto(issues) -> str:
    """A pior severidade entre os achados da conta."""
    return min(
        (SEVERIDADE_POR_ISSUE.get(i, "BAIXA") for i in issues),
        key=lambda s: ORDEM_SEVERIDADE.get(s, 9),
        default="BAIXA",
    )

BRAZIL_TZ = timezone(timedelta(hours=-3))

NO_TEAM_VALUES = {
    "",
    "-",
    "none",
    "null",
    "sem_time",
    "sem time",
    "sem-time",
    "free_agent",
    "free agent",
    "no_team",
    "no team",
}

DEFAULT_MATCH_SOURCE = "validas"

# Se True, ao rodar direto no VS Code o programa pergunta qual fonte usar.
# Se False, usa DEFAULT_MATCH_SOURCE sem perguntar.
#
# Fica em False desde 03/09/2026: a auditoria considera só Finalizados/Válidos.
# Antes o mecanismo de fonte existia mas nunca era chamado - a varredura pegava
# a pasta inteira, os 93 arquivos de Inválidos entravam junto, e o relatório
# saía com 176 contas em vez de 78. Mais da metade era ruído de partida que o
# próprio dono já tinha descartado.
ASK_MATCH_SOURCE_AT_RUNTIME = False


def choose_match_source_runtime(default_source: str = DEFAULT_MATCH_SOURCE) -> str:
    """Escolhe a fonte de partidas sem exigir parâmetros de linha de comando."""
    default_source = normalize_match_source(default_source)
    if not ASK_MATCH_SOURCE_AT_RUNTIME:
        return default_source

    options = {
        "1": "validas",
        "2": "invalidas",
        "3": "ambas",
        "4": "raiz",
        "validas": "validas",
        "v": "validas",
        "invalidas": "invalidas",
        "i": "invalidas",
        "ambas": "ambas",
        "a": "ambas",
        "raiz": "raiz",
        "r": "raiz",
    }

    print()
    print("Escolha a fonte de partidas:")
    print("  1) Válidas")
    print("  2) Inválidas")
    print("  3) Ambas")
    print("  4) Raiz antiga")
    print(f"Pressione Enter para usar o padrão: {default_source}")

    while True:
        try:
            raw = input("Fonte [1/2/3/4]: ").strip()
        except EOFError:
            return default_source

        if not raw:
            return default_source

        folded = _fold_ascii(raw)
        if folded in options:
            return options[folded]

        try:
            return normalize_match_source(raw)
        except ValueError:
            print("Opção inválida. Use 1, 2, 3, 4, validas, invalidas, ambas ou raiz.")

DEFAULT_RECURSIVE = True


# =========================
# Leitura recursiva de partidas em Finalizados
# =========================
# Opções: "validas", "invalidas", "ambas" ou "raiz".
# - validas: usa Finalizados/Válidos ou Finalizados/Validas
# - invalidas: usa Finalizados/Inválidas ou Finalizados/Invalidas
# - ambas: usa Válidas + Inválidas
# - raiz: usa diretamente a pasta informada, sem entrar nessas subpastas

def _fold_ascii(value: Any) -> str:
    s = str(value or "").strip().casefold()
    s = "".join(ch for ch in unicodedata.normalize("NFD", s) if unicodedata.category(ch) != "Mn")
    s = re.sub(r"[\s_\-]+", "-", s).strip("-")
    return s


def normalize_match_source(value: Any) -> str:
    folded = _fold_ascii(value)
    aliases = {
        "validas": "validas",
        "valida": "validas",
        "validos": "validas",
        "valido": "validas",
        "invalidas": "invalidas",
        "invalida": "invalidas",
        "invalidos": "invalidas",
        "invalido": "invalidas",
        "ambas": "ambas",
        "ambos": "ambas",
        "todas": "ambas",
        "todos": "ambas",
        "both": "ambas",
        "all": "ambas",
        "raiz": "raiz",
        "root": "raiz",
        "base": "raiz",
    }
    if folded not in aliases:
        raise ValueError("match_source inválido. Use: validas, invalidas, ambas ou raiz.")
    return aliases[folded]


# A pasta no disco pode estar no masculino ou no feminino. Procurar só por uma
# grafia fazia o resolvedor não achar "Válidos", cair no caminho inventado
# "Válidas", que não existe, e devolver zero arquivo em silêncio.
GRAFIAS_DA_FONTE = {
    "validas": ("validas", "validos", "valida", "valido"),
    "invalidas": ("invalidas", "invalidos", "invalida", "invalido"),
}


def _find_named_child_dir(base_dir: Path, folded_name: str) -> Path:
    aceitas = GRAFIAS_DA_FONTE.get(folded_name, (folded_name,))
    if base_dir.exists():
        for child in base_dir.iterdir():
            if child.is_dir() and _fold_ascii(child.name) in aceitas:
                return child

    if folded_name == "validas":
        return base_dir / "Válidas"
    if folded_name == "invalidas":
        return base_dir / "Inválidas"
    return base_dir / folded_name


def resolve_match_source_dirs(base_dir: Path, match_source: str) -> List[Path]:
    choice = normalize_match_source(match_source)
    base_dir = Path(base_dir)

    if choice == "raiz":
        return [base_dir]

    wanted = ["validas", "invalidas"] if choice == "ambas" else [choice]
    dirs = [_find_named_child_dir(base_dir, item) for item in wanted]

    if choice == "ambas" and not any(d.exists() and d.is_dir() for d in dirs):
        return [base_dir]

    return dirs


# =========================
# Normalização
# =========================

def clean_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def norm_key(value: Any) -> str:
    """Normalização conservadora para comparação de nomes, times e Riot IDs."""
    return clean_str(value).casefold()


def normalize_header_key(value: Any) -> Optional[str]:
    if value is None:
        return None

    s = clean_str(value).lower()
    if not s:
        return None

    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    compact = re.sub(r"\s+", "", s)

    aliases = {
        "jogador": "jogador",
        "player": "jogador",
        "nome": "jogador",
        "name": "jogador",
        "puuid": "puuid",
        "currentteam": "current_team",
        "timeatual": "current_team",
        "equipeatual": "current_team",
        "teamatual": "current_team",
        "current_team": "current_team",
        "team": "team",
    }
    if compact in aliases:
        return aliases[compact]

    m = re.fullmatch(r"(?:team|time|equipe)(\d*)", compact)
    if m:
        raw_idx = m.group(1)
        idx = int(raw_idx) if raw_idx else 1
        return f"team{idx}"

    m = re.fullmatch(r"nick(\d*)", compact)
    if m:
        raw_idx = m.group(1)
        idx = int(raw_idx) if raw_idx else 1
        return f"nick{idx}"

    # teams.xlsx: player1, player 1, player_1...
    m = re.fullmatch(r"player(\d*)", compact)
    if m:
        raw_idx = m.group(1)
        idx = int(raw_idx) if raw_idx else 1
        return f"player{idx}"

    if compact in {"team", "time", "equipe"}:
        return "team"

    return compact



def build_headers_from_values(header_values: Iterable[Any]) -> Dict[str, int]:
    """Retorna headers 1-based a partir da primeira linha de valores."""
    headers: Dict[str, int] = {}
    for idx, value in enumerate(header_values, start=1):
        key = normalize_header_key(value)
        if not key:
            continue
        if key in headers:
            raise RuntimeError(f"Coluna duplicada: {key} nas colunas {headers[key]} e {idx}")
        headers[key] = idx
    return headers

def build_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        key = normalize_header_key(ws.cell(row=1, column=col).value)
        if not key:
            continue
        if key in headers:
            raise RuntimeError(f"Coluna duplicada: {key} nas colunas {headers[key]} e {col}")
        headers[key] = col
    return headers


def normalize_current_team(value: Any) -> Optional[str]:
    s = clean_str(value)
    if norm_key(s) in NO_TEAM_VALUES:
        return None
    return s


def parse_riot_id(value: Any) -> Optional[str]:
    s = clean_str(value)
    if not s or "#" not in s:
        return None
    game_name, tag_line = s.rsplit("#", 1)
    game_name = game_name.strip()
    tag_line = tag_line.strip()
    if not game_name or not tag_line:
        return None
    return f"{game_name}#{tag_line}"


def riot_id_from_match_player(player: Dict[str, Any]) -> Optional[str]:
    game_name = clean_str(player.get("gameName"))
    tag_line = clean_str(player.get("tagLine"))
    if not game_name or not tag_line:
        return None
    return f"{game_name}#{tag_line}"


def is_probably_puuid(value: Any) -> bool:
    s = clean_str(value)
    return bool(s) and len(s) == 78 and re.fullmatch(r"[A-Za-z0-9_-]+", s) is not None


# =========================
# Dados de players/teams
# =========================

@dataclass
class PlayerRecord:
    row: int
    jogador: str
    current_team: Optional[str]
    puuid: Optional[str]
    riot_ids: List[str] = field(default_factory=list)
    # Colunas team 1, team 2, ... na ordem da planilha. O current_team NÃO entra
    # aqui: quem quiser as duas coisas usa equipes_conhecidas().
    team_history: List[str] = field(default_factory=list)
    # Coluna `completes`: equipes pelas quais a pessoa completou sem fazer parte
    # do elenco. Fica fora de equipes_conhecidas() de propósito - completar não
    # é passagem pela equipe.
    completes: List[str] = field(default_factory=list)

    def equipes_conhecidas(self) -> Set[str]:
        """Toda equipe pela qual há registro de passagem, normalizada."""
        equipes = {norm_key(t) for t in self.team_history if t}
        if self.current_team:
            equipes.add(norm_key(self.current_team))
        equipes.discard("")
        return equipes


@dataclass
class TeamData:
    team_names: Set[str]
    team_name_by_norm: Dict[str, str]
    roster_by_team: Dict[str, Set[str]]
    roster_norm_by_team: Dict[str, Set[str]]
    player_norm_to_teams: Dict[str, Set[str]]


@dataclass
class PlayerData:
    records: List[PlayerRecord]
    by_puuid: Dict[str, PlayerRecord]
    by_riot_id_norm: Dict[str, PlayerRecord]
    by_player_name_norm: Dict[str, PlayerRecord]


def load_players_xlsx(path: Path) -> PlayerData:
    if not path.exists():
        raise RuntimeError(f"players.xlsx não encontrado: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_values = next(rows_iter)
    except StopIteration:
        raise RuntimeError("players.xlsx está vazio.")

    headers = build_headers_from_values(header_values)

    jogador_col = headers.get("jogador")
    current_team_col = headers.get("current_team")
    puuid_col = headers.get("puuid")
    completes_col = headers.get("completes")
    # team 1, team 2, ...: histórico de passagens. É o que evita acusar como erro
    # quem joga por uma equipe pela qual já passou.
    team_cols = sorted(
        [(key, col) for key, col in headers.items() if re.fullmatch(r"team\d+", key)],
        key=lambda item: int(re.search(r"\d+", item[0]).group(0)),
    )
    # "puuid 2", "puuid 3"...: contas alternativas do mesmo dono. Sem le-las, a
    # auditoria acusa conta cadastrada como se fosse nick desconhecido.
    alt_puuid_cols = sorted(
        [(key, col) for key, col in headers.items() if re.fullmatch(r"puuid\d+", key) and int(re.search(r"\d+", key).group(0)) >= 2],
        key=lambda item: int(re.search(r"\d+", item[0]).group(0)),
    )
    nick_cols = sorted(
        [(key, col) for key, col in headers.items() if re.fullmatch(r"nick\d+", key)],
        key=lambda item: int(re.search(r"\d+", item[0]).group(0)),
    )

    if jogador_col is None:
        raise RuntimeError("players.xlsx precisa ter coluna 'Jogador'.")
    if current_team_col is None:
        raise RuntimeError("players.xlsx precisa ter coluna 'current_team'.")
    if puuid_col is None and not nick_cols:
        raise RuntimeError("players.xlsx precisa ter coluna 'puuid' e/ou colunas 'nick N'.")

    def get_cell(row_values: Tuple[Any, ...], one_based_col: Optional[int]) -> Any:
        if one_based_col is None:
            return None
        idx = one_based_col - 1
        if idx < 0 or idx >= len(row_values):
            return None
        return row_values[idx]

    records: List[PlayerRecord] = []
    by_puuid: Dict[str, PlayerRecord] = {}
    by_riot_id_norm: Dict[str, PlayerRecord] = {}
    by_player_name_norm: Dict[str, PlayerRecord] = {}

    for row_num, row_values in enumerate(rows_iter, start=2):
        jogador = clean_str(get_cell(row_values, jogador_col))
        if not jogador:
            continue

        current_team = normalize_current_team(get_cell(row_values, current_team_col))

        puuid = None
        raw_puuid = clean_str(get_cell(row_values, puuid_col))
        if is_probably_puuid(raw_puuid):
            puuid = raw_puuid

        riot_ids: List[str] = []
        seen_riot_norms: Set[str] = set()

        # Se alguém ainda colocou Riot ID na coluna puuid, aproveita como nick.
        possible_riot = parse_riot_id(raw_puuid)
        if possible_riot:
            seen_riot_norms.add(norm_key(possible_riot))
            riot_ids.append(possible_riot)

        for _key, col in nick_cols:
            riot_id = parse_riot_id(get_cell(row_values, col))
            if not riot_id:
                continue
            nk = norm_key(riot_id)
            if nk in seen_riot_norms:
                continue
            seen_riot_norms.add(nk)
            riot_ids.append(riot_id)

        team_history: List[str] = []
        for _team_key, team_col in team_cols:
            equipe = normalize_current_team(get_cell(row_values, team_col))
            if equipe and equipe not in team_history:
                team_history.append(equipe)

        completes: List[str] = []
        if completes_col is not None:
            bruto = clean_str(get_cell(row_values, completes_col))
            for pedaco in bruto.split(";"):
                equipe = normalize_current_team(pedaco)
                if equipe and equipe not in completes:
                    completes.append(equipe)

        rec = PlayerRecord(
            row=row_num,
            jogador=jogador,
            current_team=current_team,
            puuid=puuid,
            riot_ids=riot_ids,
            team_history=team_history,
            completes=completes,
        )
        records.append(rec)

        name_key = norm_key(jogador)
        if name_key in by_player_name_norm:
            raise RuntimeError(f"Jogador duplicado em players.xlsx: {jogador}")
        by_player_name_norm[name_key] = rec

        if puuid:
            if puuid in by_puuid:
                other = by_puuid[puuid]
                raise RuntimeError(
                    f"PUUID duplicado em players.xlsx: {puuid} em {other.jogador} e {jogador}"
                )
            by_puuid[puuid] = rec

        for _alt_key, alt_col in alt_puuid_cols:
            alt_puuid = clean_str(get_cell(row_values, alt_col))
            if not is_probably_puuid(alt_puuid):
                continue
            other = by_puuid.get(alt_puuid)
            if other is not None and other is not rec:
                raise RuntimeError(
                    f"PUUID duplicado em players.xlsx: {alt_puuid} em {other.jogador} e {jogador}"
                )
            by_puuid[alt_puuid] = rec

        for riot_id in riot_ids:
            rk = norm_key(riot_id)
            if rk in by_riot_id_norm:
                other = by_riot_id_norm[rk]
                raise RuntimeError(
                    f"Riot ID duplicado em players.xlsx: {riot_id} em {other.jogador} e {jogador}"
                )
            by_riot_id_norm[rk] = rec

    return PlayerData(
        records=records,
        by_puuid=by_puuid,
        by_riot_id_norm=by_riot_id_norm,
        by_player_name_norm=by_player_name_norm,
    )

def load_teams_xlsx(path: Path) -> TeamData:
    if not path.exists():
        raise RuntimeError(f"teams.xlsx não encontrado: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_values = next(rows_iter)
    except StopIteration:
        raise RuntimeError("teams.xlsx está vazio.")

    headers = build_headers_from_values(header_values)

    team_col = headers.get("team")
    if team_col is None:
        raise RuntimeError("teams.xlsx precisa ter coluna 'team'.")

    player_cols = sorted(
        [(key, col) for key, col in headers.items() if re.fullmatch(r"player\d+", key)],
        key=lambda item: int(re.search(r"\d+", item[0]).group(0)),
    )
    if not player_cols:
        raise RuntimeError("teams.xlsx precisa ter colunas 'player1', 'player2', ...")

    def get_cell(row_values: Tuple[Any, ...], one_based_col: Optional[int]) -> Any:
        if one_based_col is None:
            return None
        idx = one_based_col - 1
        if idx < 0 or idx >= len(row_values):
            return None
        return row_values[idx]

    team_names: Set[str] = set()
    team_name_by_norm: Dict[str, str] = {}
    roster_by_team: Dict[str, Set[str]] = defaultdict(set)
    roster_norm_by_team: Dict[str, Set[str]] = defaultdict(set)
    player_norm_to_teams: Dict[str, Set[str]] = defaultdict(set)

    for row_num, row_values in enumerate(rows_iter, start=2):
        team = clean_str(get_cell(row_values, team_col))
        if not team:
            continue

        team_norm = norm_key(team)
        if team_norm in team_name_by_norm:
            raise RuntimeError(f"Time duplicado em teams.xlsx: {team}")

        team_names.add(team)
        team_name_by_norm[team_norm] = team

        for _key, col in player_cols:
            player_name = clean_str(get_cell(row_values, col))
            if not player_name:
                continue

            player_norm = norm_key(player_name)
            roster_by_team[team].add(player_name)
            roster_norm_by_team[team].add(player_norm)
            player_norm_to_teams[player_norm].add(team)

    return TeamData(
        team_names=team_names,
        team_name_by_norm=team_name_by_norm,
        roster_by_team=dict(roster_by_team),
        roster_norm_by_team=dict(roster_norm_by_team),
        player_norm_to_teams=dict(player_norm_to_teams),
    )


# =========================
# JSON / detecção de time
# =========================

def read_match_json(path: Path) -> Optional[Dict[str, Any]]:
    try:
        obj = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return obj if isinstance(obj, dict) else None


def active_players(match: Dict[str, Any]) -> List[Dict[str, Any]]:
    players = match.get("players") or []
    if not isinstance(players, list):
        return []

    result = []
    for player in players:
        if not isinstance(player, dict):
            continue
        if bool(player.get("isObserver", False)):
            continue
        if player.get("teamId") not in ("Blue", "Red"):
            continue
        result.append(player)
    return result


def find_player_record(player: Dict[str, Any], pdata: PlayerData) -> Optional[PlayerRecord]:
    puuid = clean_str(player.get("puuid"))
    if puuid and puuid in pdata.by_puuid:
        return pdata.by_puuid[puuid]

    riot_id = riot_id_from_match_player(player)
    if riot_id:
        return pdata.by_riot_id_norm.get(norm_key(riot_id))

    return None


def get_score_by_side(match: Dict[str, Any]) -> Dict[str, int]:
    score_by_side: Dict[str, int] = {}
    teams_payload = match.get("teams") or []
    if not isinstance(teams_payload, list):
        return score_by_side

    for team in teams_payload:
        if not isinstance(team, dict):
            continue
        side = team.get("teamId")
        if side not in ("Blue", "Red"):
            continue

        score = None
        for key in ("roundsWon", "wonRounds", "score", "numPoints"):
            value = team.get(key)
            if isinstance(value, int):
                score = value
                break

        if isinstance(score, int):
            score_by_side[side] = score

    return score_by_side


def detect_sides_by_known_players(
    match: Dict[str, Any],
    pdata: PlayerData,
    min_players: int,
) -> Tuple[Dict[str, str], Dict[str, Dict[str, int]], Dict[str, Dict[str, int]], Set[str]]:
    """
    Descobre a equipe de cada lado pelo voto dos jogadores conhecidos.

    Duas passadas. A primeira só conta quem tem `current_team`, que é o dado
    mais confiável. A segunda, feita apenas para o lado que sobrou sem
    resposta, deixa votar também quem está sem `current_team`, usando o último
    time do histórico.

    A segunda passada existe porque a versão anterior era circular: jogador sem
    `current_team` não votava, então o lado inteiro de uma equipe com cadastro
    incompleto ficava sem apuração - e aí a auditoria não conseguia medir
    justamente o caso que ela existe para reportar.

    Os dois resultados voltam SEPARADOS de propósito. O voto por histórico é o
    sinal mais fraco dos três e precisa entrar depois do nome do arquivo: numa
    primeira versão ele vinha junto com o voto normal, ganhava do nome, e quatro
    partidas com "fei_whiteowls" escrito no arquivo foram atribuídas à
    fei_darkowls porque o histórico dos jogadores apontava para lá.
    """
    votos_atual: Dict[str, Counter] = {"Blue": Counter(), "Red": Counter()}
    votos_com_historico: Dict[str, Counter] = {"Blue": Counter(), "Red": Counter()}

    for player in active_players(match):
        side = str(player.get("teamId"))
        if side not in votos_atual:
            continue
        rec = find_player_record(player, pdata)
        if not rec:
            continue
        if rec.current_team:
            votos_atual[side][rec.current_team] += 1
            votos_com_historico[side][rec.current_team] += 1
        elif rec.team_history:
            votos_com_historico[side][rec.team_history[-1]] += 1

    def decidir(counter: Counter) -> Optional[str]:
        if not counter:
            return None
        mais_votados = counter.most_common()
        equipe, votos = mais_votados[0]
        # Empate no primeiro lugar não decide nada.
        if len(mais_votados) >= 2 and mais_votados[1][1] == votos:
            return None
        return equipe if votos >= min_players else None

    side_to_team: Dict[str, str] = {}
    por_historico: Dict[str, str] = {}

    for side in ("Blue", "Red"):
        equipe = decidir(votos_atual[side])
        if equipe is not None:
            side_to_team[side] = equipe
            continue
        equipe = decidir(votos_com_historico[side])
        if equipe is not None:
            por_historico[side] = equipe

    debug_counts = {side: dict(votos_atual[side]) for side in votos_atual}
    debug_counts_historico = {side: dict(votos_com_historico[side]) for side in votos_com_historico}
    return side_to_team, debug_counts, debug_counts_historico, por_historico


# Nome compacto do RivvalsGG: Rivals_001_a_puccXazr.json. Ele não traz placar,
# só o par de equipes, e por isso não serve para mapear lado por placar - serve
# para COMPLETAR por eliminação: sabendo uma ponta, a outra é a que sobrou.
# A tabela espelha FILENAME_TEAM_ALIASES do app.js; se mudar lá, mude aqui.
ALIASES_NOME_COMPACTO = {
    "azr": "azure_bears",
    "caap": "caap_hellhounds",
    "fei": "fei_darkowls",
    "mackred": "macklogic_red",
    "pucc": "pucc_cardinals",
    "tritons": "unicamp_tritons_red",
    "uspstars": "usp_stars",
    "axissupernova": "axis_anteaters",
}


def mesma_familia_de_equipe(a: Optional[str], b: Optional[str]) -> bool:
    """
    Duas referências apontam para a mesma equipe ou para a mesma organização?

    "azr" vira "azure_bears", que não é equipe de verdade: é o prefixo comum de
    azure_bears_black, _golden, _silver. Por isso a comparação aceita prefixo.
    """
    if not a or not b:
        return False
    na, nb = norm_key(a), norm_key(b)
    return na == nb or na.startswith(nb + "_") or nb.startswith(na + "_")


def parse_compact_pair_filename(file_name: str) -> Optional[Tuple[str, str]]:
    """
    Extrai o par de equipes de um nome no formato Rivals_<n>_<sufixo>_<aXb>.

    O separador é a letra x, que também aparece dentro de nome de equipe, então
    testamos toda posição de x e aceitamos só a divisão em que as DUAS metades
    são apelidos conhecidos - sem chute.
    """
    # Um arquivo do acervo está gravado como "....json.json"; o .stem tira só
    # uma extensão e deixaria ".json" colado no nome da equipe.
    stem = re.sub(r"(?:\.json)+$", "", str(file_name), flags=re.IGNORECASE)
    partes = Path(stem).name.split("_")
    if len(partes) < 2 or norm_key(partes[0]) != "rivals":
        return None

    par = partes[-1]
    alvo = par.lower()
    for i, caractere in enumerate(alvo):
        if caractere != "x":
            continue
        esquerda, direita = alvo[:i], alvo[i + 1:]
        eq_a = ALIASES_NOME_COMPACTO.get(esquerda)
        eq_b = ALIASES_NOME_COMPACTO.get(direita)
        if eq_a and eq_b:
            return eq_a, eq_b
    return None


def completar_lados_pelo_par_do_nome(
    side_to_team: Dict[str, str],
    par: Optional[Tuple[str, str]],
    team_data: TeamData,
    votos_por_lado: Optional[Dict[str, Dict[str, int]]] = None,
) -> Dict[str, str]:
    """
    Preenche lado em aberto quando o nome do arquivo nomeia as duas equipes.

    Com um lado já resolvido, é eliminação pura: a outra ponta é a que sobrou.

    Com os DOIS em aberto, o nome já diz quais são as duas equipes e falta só
    orientar - e para isso o voto mais forte basta, mesmo abaixo do mínimo de
    jogadores, porque a resposta só tem duas possibilidades. Sem esse ramo, uma
    transferência que tira um voto de um lado derruba a partida inteira: foi o
    que aconteceu com o Rivals_006 quando a Luqueta saiu da caap_hellhounds.

    Só atribui equipe que exista no teams.xlsx.
    """
    if not par:
        return {}

    oficial = {equipe: team_data.team_name_by_norm.get(norm_key(equipe)) for equipe in par}
    faltando = [lado for lado in ("Blue", "Red") if lado not in side_to_team]
    resolvidos = [lado for lado in ("Blue", "Red") if lado in side_to_team]

    if len(faltando) == 1 and len(resolvidos) == 1:
        equipe_conhecida = side_to_team[resolvidos[0]]
        if mesma_familia_de_equipe(equipe_conhecida, par[0]):
            candidata = par[1]
        elif mesma_familia_de_equipe(equipe_conhecida, par[1]):
            candidata = par[0]
        else:
            # O lado resolvido não é nenhuma das duas do nome: alguma das fontes
            # está errada, e completar por eliminação seria chutar.
            return {}
        return {faltando[0]: oficial[candidata]} if oficial[candidata] else {}

    if len(faltando) == 2 and votos_por_lado and all(oficial.values()):
        def votos(lado: str, equipe: str) -> int:
            contagem = votos_por_lado.get(lado) or {}
            return sum(n for votada, n in contagem.items() if mesma_familia_de_equipe(votada, equipe))

        direto = votos("Blue", par[0]) + votos("Red", par[1])
        invertido = votos("Blue", par[1]) + votos("Red", par[0])
        if direto == invertido:
            return {}
        if direto > invertido:
            return {"Blue": oficial[par[0]], "Red": oficial[par[1]]}
        return {"Blue": oficial[par[1]], "Red": oficial[par[0]]}

    return {}


def parse_filename_matchup(
    file_name: str,
    team_data: TeamData,
) -> Dict[str, Any]:
    """
    Extrai teamA, scoreA, scoreB, teamB do nome do arquivo.

    Suporta formatos como:
      AOC_008a_04-06-26_azure_bears_golden_11_x_13_ceub_octopus_breeze.json
      CIA_002a_caap_hellhounds_13_x_6_unifacens_fodens_pearl.json
      1_24-04-26_caap_hellhounds_11_x_13_ufu_saints_lotus.json

    A extração de time usa a lista oficial do teams.xlsx:
      - teamA = maior time conhecido que é sufixo da parte antes do placar
      - teamB = maior time conhecido que é prefixo da parte depois do placar
    """
    stem = Path(file_name).stem
    stem_norm = norm_key(stem)
    tokens = stem_norm.split("_")

    known_team_norms = sorted(team_data.team_name_by_norm.keys(), key=len, reverse=True)

    def canonical(team_norm: Optional[str]) -> Optional[str]:
        if not team_norm:
            return None
        return team_data.team_name_by_norm.get(team_norm)

    def longest_team_suffix(text: str) -> Optional[str]:
        for team_norm in known_team_norms:
            if text == team_norm or text.endswith("_" + team_norm):
                return team_norm
        return None

    def longest_team_prefix(text: str) -> Optional[str]:
        for team_norm in known_team_norms:
            if text == team_norm or text.startswith(team_norm + "_"):
                return team_norm
        return None

    for i in range(0, max(0, len(tokens) - 2)):
        if tokens[i].isdigit() and tokens[i + 1] == "x" and tokens[i + 2].isdigit():
            pre = "_".join(tokens[:i])
            post = "_".join(tokens[i + 3:])

            team_a_norm = longest_team_suffix(pre)
            team_b_norm = longest_team_prefix(post)

            return {
                "parsed": bool(team_a_norm or team_b_norm),
                "teamA": canonical(team_a_norm),
                "scoreA": int(tokens[i]),
                "scoreB": int(tokens[i + 2]),
                "teamB": canonical(team_b_norm),
                "rawPreScore": pre,
                "rawPostScore": post,
            }

    # Fallback: acha os dois primeiros times conhecidos que aparecem no nome, priorizando posição.
    found: List[Tuple[int, int, str]] = []
    for team_norm in known_team_norms:
        idx = stem_norm.find(team_norm)
        if idx >= 0:
            found.append((idx, -len(team_norm), team_norm))
    found.sort()

    teams_found = []
    seen = set()
    for _idx, _neg_len, team_norm in found:
        team = canonical(team_norm)
        if team and team not in seen:
            teams_found.append(team)
            seen.add(team)
        if len(teams_found) == 2:
            break

    return {
        "parsed": bool(teams_found),
        "teamA": teams_found[0] if len(teams_found) >= 1 else None,
        "scoreA": None,
        "scoreB": None,
        "teamB": teams_found[1] if len(teams_found) >= 2 else None,
        "rawPreScore": None,
        "rawPostScore": None,
    }


def map_filename_teams_to_sides(
    filename_info: Dict[str, Any],
    score_by_side: Dict[str, int],
) -> Dict[str, str]:
    team_a = filename_info.get("teamA")
    team_b = filename_info.get("teamB")
    score_a = filename_info.get("scoreA")
    score_b = filename_info.get("scoreB")

    if not team_a and not team_b:
        return {}

    # Se não temos placar confiável, não dá para saber qual time foi Blue/Red só pelo nome.
    if not isinstance(score_a, int) or not isinstance(score_b, int) or not score_by_side:
        return {}

    # Placar igual não deveria ocorrer em VALORANT, mas evita mapeamento ambíguo.
    if score_a == score_b:
        return {}

    side_to_team: Dict[str, str] = {}
    for side, side_score in score_by_side.items():
        if side_score == score_a and team_a:
            side_to_team[side] = team_a
        elif side_score == score_b and team_b:
            side_to_team[side] = team_b

    return side_to_team


def combine_side_detection(
    by_players: Dict[str, str],
    by_filename: Dict[str, str],
) -> Tuple[Dict[str, str], List[Dict[str, str]], Dict[str, str]]:
    """
    Prioridade:
      1) nome do arquivo + placar;
      2) detecção por >=3 jogadores com current_team.

    O nome do arquivo vem primeiro porque ele registra **por que equipe a pessoa
    jogava na data da partida**, que é a pergunta que a auditoria faz. O
    `current_team` é o estado de hoje: para partida antiga ele erra sempre que
    alguém trocou de equipe depois - e trocar de equipe é a regra, não a
    exceção, num circuito universitário. Foi por isso que 29 lados saíam
    atribuídos à equipe nova da pessoa em vez da equipe que disputou o jogo.

    A troca põe peso no nome do arquivo, então nome errado passa a corromper a
    atribuição direto. Os conflitos continuam sendo registrados: eles são a
    forma de pegar nome errado - e foi assim que apareceram os dois JUBS_016
    gravados em duplicata com o placar invertido.

    Retorna:
      - side_to_team final
      - conflitos entre as duas fontes
      - source_by_side
    """
    final: Dict[str, str] = {}
    conflicts: List[Dict[str, str]] = []
    source_by_side: Dict[str, str] = {}

    for side in ("Blue", "Red"):
        team_players = by_players.get(side)
        team_file = by_filename.get(side)

        if team_file and team_players and norm_key(team_file) != norm_key(team_players):
            conflicts.append({
                "side": side,
                "teamByPlayers": team_players,
                "teamByFilename": team_file,
            })

        if team_file:
            final[side] = team_file
            source_by_side[side] = "filename_score"
        elif team_players:
            final[side] = team_players
            source_by_side[side] = "players_3plus"

    return final, conflicts, source_by_side


# =========================
# Auditoria
# =========================

@dataclass
class SummaryItem:
    puuid: Optional[str]
    display_riot_id: str
    issue_types: Set[str] = field(default_factory=set)
    registered_player: Optional[str] = None
    registered_current_team: Optional[str] = None
    registered_team_history: List[str] = field(default_factory=list)
    registered_completes: List[str] = field(default_factory=list)
    total_matches: Set[str] = field(default_factory=set)
    teams_to_matches: Dict[str, Set[str]] = field(default_factory=lambda: defaultdict(set))
    sides: Counter = field(default_factory=Counter)
    files: Set[str] = field(default_factory=set)
    sample_riot_ids: Set[str] = field(default_factory=set)


def should_report_player(
    rec: Optional[PlayerRecord],
    inferred_team: Optional[str],
    team_data: TeamData,
) -> Tuple[bool, List[str]]:
    issues: List[str] = []

    # Jogou e não existe linha para ele. É o achado mais importante da auditoria.
    if rec is None:
        return True, ["nunca_cadastrado"]

    sem_time_atual = not rec.current_team

    # Sem equipe inferida para o lado não há com o que comparar: só o cadastro
    # incompleto continua valendo como achado.
    if not inferred_team:
        if sem_time_atual and REPORT_SEM_TIME_ATUAL:
            issues.append("sem_time_atual")
        return bool(issues), issues

    inferida = norm_key(inferred_team)
    atual = norm_key(rec.current_team) if rec.current_team else ""

    # Caso normal: jogou pela equipe atual. Sobra conferir se as duas planilhas
    # concordam, ou seja, se ele está no roster daquela equipe no teams.xlsx.
    if atual and inferida == atual:
        if REPORT_FORA_DO_ROSTER_TEAMS and not esta_no_roster(rec, rec.current_team, team_data):
            issues.append("fora_do_roster_do_teams")
        return bool(issues), issues

    # Equipe do próprio histórico: não é erro de atribuição. No máximo o cadastro
    # está desatualizado, e isso já é dito por sem_time_atual.
    if inferida in rec.equipes_conhecidas():
        if sem_time_atual and REPORT_SEM_TIME_ATUAL:
            issues.append("sem_time_atual")
        elif REPORT_EQUIPE_DO_HISTORICO_NAO_ATUAL:
            issues.append("equipe_do_historico_nao_atual")
        return bool(issues), issues

    # Completou por essa equipe: fato conhecido, não pendência.
    if inferida in {norm_key(equipe) for equipe in rec.completes}:
        if sem_time_atual and REPORT_SEM_TIME_ATUAL:
            issues.append("sem_time_atual")
        elif REPORT_COMPLETE_POR_OUTRA_EQUIPE:
            issues.append("complete_por_outra_equipe")
        return bool(issues), issues

    # Está no elenco daquela equipe no teams.xlsx, mesmo que o current_team seja
    # outro: é participação simultânea, não erro.
    if esta_no_roster(rec, inferred_team, team_data):
        if sem_time_atual and REPORT_SEM_TIME_ATUAL:
            issues.append("sem_time_atual")
        elif REPORT_ELENCO_DE_DUAS_EQUIPES:
            issues.append("elenco_de_duas_equipes")
        return bool(issues), issues

    # Nem a atual, nem nenhuma do histórico, nem complete, nem elenco: é
    # transferência não registrada, ou a pessoa está na equipe errada.
    issues.append("equipe_fora_do_historico")
    if sem_time_atual and REPORT_SEM_TIME_ATUAL:
        issues.append("sem_time_atual")
    return True, issues


def esta_no_roster(rec: PlayerRecord, equipe: str, team_data: TeamData) -> bool:
    """
    O jogador aparece no roster daquela equipe no teams.xlsx?

    A busca do nome da equipe é normalizada de propósito: casar a chave crua
    faria toda a equipe cair como "fora do roster" por uma diferença de caixa
    ou de espaço entre as duas planilhas.
    """
    nome_oficial = team_data.team_name_by_norm.get(norm_key(equipe))
    if nome_oficial is None:
        return False
    return norm_key(rec.jogador) in team_data.roster_norm_by_team.get(nome_oficial, set())



def linhas_seguras(linhas: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    O CSV carrega nick da Riot e nome de arquivo, os dois escolhidos por
    terceiro. Aberto no Excel, celula que comeca com = vira formula.
    """
    return [{k: config_hub.texto_seguro(v) for k, v in linha.items()} for linha in linhas]


def iter_json_files(
    matches_dir: Path,
    recursive: bool = True,
    match_source: str = DEFAULT_MATCH_SOURCE,
) -> Iterable[Path]:
    """
    Lê os JSONs das subpastas escolhidas por `match_source`.

    Em modo "raiz" a leitura fica na pasta informada e não desce para
    Válidos/Inválidos. Nas demais, cada pasta resolvida é varrida e o mesmo
    arquivo nunca sai duas vezes, mesmo que dois caminhos apontem para ele.
    """
    escolha = normalize_match_source(match_source)
    pattern = "*.json" if escolha == "raiz" else ("**/*.json" if recursive else "*.json")
    vistos: Set[Path] = set()

    for source_dir in resolve_match_source_dirs(matches_dir, match_source):
        if not source_dir.exists() or not source_dir.is_dir():
            continue
        for fp in sorted(source_dir.glob(pattern)):
            if not fp.is_file():
                continue
            chave = fp.resolve()
            if chave in vistos:
                continue
            vistos.add(chave)
            yield fp

def nome_bate_com_o_placar(
    file_name: str,
    match: Dict[str, Any],
    team_data: TeamData,
    pdata: PlayerData,
    min_players: int,
) -> bool:
    """
    O nome do arquivo dá a cada equipe o placar que ela realmente fez?

    É o critério para escolher entre duas cópias da mesma partida gravadas com
    nomes diferentes: fica a que concorda com o próprio conteúdo.

    Comparar o par de placares não resolve: "13_x_5" e "5_x_13" têm o mesmo par,
    e é exatamente aí que as duas cópias diferem. Então o lado de cada equipe vem
    do voto dos jogadores - que aqui é a fonte independente - e o teste é se o
    placar que o nome dá àquela equipe bate com o placar do lado dela.
    """
    info = parse_filename_matchup(file_name, team_data)
    if not info.get("parsed"):
        return False

    placar_do_lado = get_score_by_side(match)
    votos, _contagem, _hist, _por_hist = detect_sides_by_known_players(match, pdata, min_players)
    if not votos:
        return False

    placar_pelo_nome: Dict[str, int] = {}
    if info.get("teamA") and isinstance(info.get("scoreA"), int):
        placar_pelo_nome[norm_key(info["teamA"])] = info["scoreA"]
    if info.get("teamB") and isinstance(info.get("scoreB"), int):
        placar_pelo_nome[norm_key(info["teamB"])] = info["scoreB"]

    conferidos = 0
    for lado, equipe in votos.items():
        esperado = placar_pelo_nome.get(norm_key(equipe))
        real = placar_do_lado.get(lado)
        if esperado is None or real is None:
            continue
        if esperado != real:
            return False
        conferidos += 1

    return conferidos > 0


def audit_matches(
    players_path: Path,
    teams_path: Path,
    matches_dir: Path,
    output_dir: Path,
    min_team_players: int = MIN_TEAM_PLAYERS_TO_IDENTIFY,
    recursive: bool = True,
    match_source: str = DEFAULT_MATCH_SOURCE,
) -> Dict[str, Any]:
    pdata = load_players_xlsx(players_path)
    tdata = load_teams_xlsx(teams_path)

    if not matches_dir.exists():
        raise RuntimeError(f"Pasta de JSONs não encontrada: {matches_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    fonte = normalize_match_source(match_source)
    pastas = [d for d in resolve_match_source_dirs(matches_dir, fonte) if d.exists() and d.is_dir()]
    json_files = list(iter_json_files(matches_dir, recursive=recursive, match_source=fonte))
    if not json_files:
        alvos = ", ".join(str(d) for d in pastas) or str(matches_dir)
        raise RuntimeError(f"Nenhum .json encontrado em: {alvos}")

    summary: Dict[str, SummaryItem] = {}
    detail_rows: List[Dict[str, Any]] = []
    debug_rows: List[Dict[str, Any]] = []

    # Linha do tempo, montada para TODA partida e TODA conta - inclusive as que
    # não viram achado. Sem isso não dá para responder "a equipe seguiu jogando
    # sem ele?", que é a pergunta que separa quem saiu de quem só está com o
    # cadastro incompleto.
    partidas_por_equipe: Dict[str, Dict[str, str]] = defaultdict(dict)
    historico_da_conta: Dict[str, Dict[str, Tuple[str, str]]] = defaultdict(dict)

    processed_files = 0
    invalid_files = 0
    arquivos_sem_jogadores = 0
    total_active_players_seen = 0
    reported_occurrences = 0

    # A mesma partida pode estar gravada duas vezes com nomes diferentes - e com
    # o nome do arquivo mandando na atribuição de equipe, a cópia errada
    # corrompe o resultado. Fica a cópia cujo nome concorda com o placar de
    # dentro do JSON; as outras entram no relatório como duplicata.
    vistos_por_match_id: Dict[str, Path] = {}
    duplicatas: List[Dict[str, str]] = []
    escolhidos: List[Tuple[Path, Dict[str, Any]]] = []

    for fp in json_files:
        match = read_match_json(fp)
        if not match:
            invalid_files += 1
            continue
        if not match.get("players"):
            arquivos_sem_jogadores += 1
            continue

        match_id = clean_str((match.get("matchInfo") or {}).get("matchId"))
        if not match_id:
            escolhidos.append((fp, match))
            continue

        anterior = vistos_por_match_id.get(match_id)
        if anterior is None:
            vistos_por_match_id[match_id] = fp
            escolhidos.append((fp, match))
            continue

        indice = next(i for i, (caminho, _m) in enumerate(escolhidos) if caminho == anterior)
        nova_confere = nome_bate_com_o_placar(fp.name, match, tdata, pdata, min_team_players)
        antiga_confere = nome_bate_com_o_placar(anterior.name, escolhidos[indice][1], tdata, pdata, min_team_players)

        if nova_confere and not antiga_confere:
            escolhidos[indice] = (fp, match)
            vistos_por_match_id[match_id] = fp
            descartado, mantido = anterior.name, fp.name
        else:
            descartado, mantido = fp.name, anterior.name

        duplicatas.append({
            "matchId": match_id,
            "mantido": mantido,
            "descartado": descartado,
            "criterio": "placar do nome bate com o do JSON" if (nova_confere != antiga_confere) else "primeiro encontrado",
        })

    for fp, match in escolhidos:

        processed_files += 1
        match_info = match.get("matchInfo") or {}
        match_id = clean_str(match_info.get("matchId")) or fp.stem
        game_start_ms = match_info.get("gameStartMillis")

        match_date = ""
        if isinstance(game_start_ms, (int, float)):
            try:
                dt = datetime.fromtimestamp(int(game_start_ms) / 1000, tz=BRAZIL_TZ)
                match_date = dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                match_date = ""

        (
            players_by_detection,
            counts_by_side,
            counts_by_side_historico,
            players_por_historico,
        ) = detect_sides_by_known_players(
            match=match,
            pdata=pdata,
            min_players=min_team_players,
        )

        filename_info = parse_filename_matchup(fp.name, tdata)
        score_by_side = get_score_by_side(match)
        filename_by_side = map_filename_teams_to_sides(filename_info, score_by_side)

        side_to_team, conflicts, source_by_side = combine_side_detection(
            by_players=players_by_detection,
            by_filename=filename_by_side,
        )

        # Precedência: voto por current_team, depois nome do arquivo, e só então
        # voto por histórico. Ele preenche o que sobrou, nunca sobrepõe.
        for lado, equipe in players_por_historico.items():
            if lado not in side_to_team:
                side_to_team[lado] = equipe
                source_by_side[lado] = "players_3plus_historico"

        par_do_nome = parse_compact_pair_filename(fp.name)
        for lado_completado, equipe_completada in completar_lados_pelo_par_do_nome(
            side_to_team, par_do_nome, tdata, counts_by_side_historico
        ).items():
            side_to_team[lado_completado] = equipe_completada
            source_by_side[lado_completado] = "par_do_nome_compacto"

        for _lado, _equipe in side_to_team.items():
            if _equipe:
                partidas_por_equipe[_equipe][match_id] = match_date

        debug_rows.append({
            "fileName": fp.name,
            "matchId": match_id,
            "matchDate": match_date,
            "scoreBySide": score_by_side,
            "filenameInfo": filename_info,
            "teamByPlayers3Plus": players_by_detection,
            "teamByFilenameScore": filename_by_side,
            "finalSideToTeam": side_to_team,
            "sourceBySide": source_by_side,
            "playerTeamCountsBySide": counts_by_side,
            "playerTeamCountsBySideComHistorico": counts_by_side_historico,
            "sidesResolvidosPeloHistorico": {
                lado: equipe for lado, equipe in players_por_historico.items()
                if source_by_side.get(lado) == "players_3plus_historico"
            },
            "parDoNomeCompacto": list(par_do_nome) if par_do_nome else None,
            "conflicts": conflicts,
        })

        for player in active_players(match):
            total_active_players_seen += 1

            riot_id = riot_id_from_match_player(player) or ""
            puuid = clean_str(player.get("puuid")) or None
            side = clean_str(player.get("teamId"))
            inferred_team = side_to_team.get(side)

            rec = find_player_record(player, pdata)

            chave = f"puuid:{puuid}" if puuid else f"riot:{norm_key(riot_id)}"
            historico_da_conta[chave][match_id] = (match_date, inferred_team or "")

            report_it, issues = should_report_player(rec, inferred_team, tdata)

            if not report_it:
                continue

            reported_occurrences += 1

            # Chave de agregação: PUUID quando existe; senão Riot ID normalizado.
            agg_key = f"puuid:{puuid}" if puuid else f"riot:{norm_key(riot_id)}"
            if agg_key not in summary:
                summary[agg_key] = SummaryItem(
                    puuid=puuid,
                    display_riot_id=riot_id or "(sem Riot ID)",
                    registered_player=rec.jogador if rec else None,
                    registered_current_team=rec.current_team if rec else None,
                    registered_team_history=list(rec.team_history) if rec else [],
                    registered_completes=list(rec.completes) if rec else [],
                )

            item = summary[agg_key]
            if riot_id:
                item.sample_riot_ids.add(riot_id)
                # Mantém o Riot ID mais recente visto no JSON como display.
                item.display_riot_id = riot_id

            if rec:
                item.registered_player = rec.jogador
                item.registered_current_team = rec.current_team
                item.registered_team_history = list(rec.team_history)
                item.registered_completes = list(rec.completes)

            item.issue_types.update(issues)
            item.total_matches.add(match_id)
            item.sides[side] += 1
            item.files.add(fp.name)
            item.teams_to_matches[inferred_team or "unknown_team"].add(match_id)

            detail_rows.append({
                "riot_id": riot_id,
                "puuid": puuid or "",
                "issues": "; ".join(issues),
                "associated_team": inferred_team or "unknown_team",
                "side": side,
                "file_name": fp.name,
                "match_id": match_id,
                "match_date_brt": match_date,
                "registered_player": rec.jogador if rec else "",
                "registered_current_team": rec.current_team if rec and rec.current_team else "",
                "registered_team_history": "; ".join(rec.team_history) if rec else "",
                "source_for_associated_team": source_by_side.get(side, ""),
                "team_by_players_3plus": players_by_detection.get(side, ""),
                "team_by_filename_score": filename_by_side.get(side, ""),
                "filename_teamA": filename_info.get("teamA") or "",
                "filename_scoreA": filename_info.get("scoreA") if filename_info.get("scoreA") is not None else "",
                "filename_scoreB": filename_info.get("scoreB") if filename_info.get("scoreB") is not None else "",
                "filename_teamB": filename_info.get("teamB") or "",
            })

    def metricas_de_atividade(chave: str, item: SummaryItem) -> Dict[str, Any]:
        """
        Datas das partidas da conta e o que a equipe de referência fez depois.

        A "equipe de referência" é onde a pessoa deveria estar: o current_team
        quando existe, senão a equipe pela qual ela mais jogou. Comparar a última
        partida dela com a última da equipe é o que distingue quem saiu do time
        de quem continua nele com o cadastro pela metade.
        """
        proprias = historico_da_conta.get(chave, {})
        datas_proprias = sorted(d for d, _e in proprias.values() if d)

        equipe_ref = item.registered_current_team
        if not equipe_ref:
            # "unknown_team" sai ANTES do desempate: ele é o rótulo das partidas
            # em que não deu para apurar o lado, e como vinha por último no
            # alfabeto vencia todo empate e zerava as métricas da conta.
            candidatas = {
                equipe: ids
                for equipe, ids in item.teams_to_matches.items()
                if equipe and equipe != "unknown_team"
            }
            if candidatas:
                equipe_ref = max(candidatas.items(), key=lambda kv: (len(kv[1]), kv[0]))[0]
            elif item.registered_team_history:
                # Nenhuma partida com lado apurado: o último time do histórico é
                # a melhor aposta de onde a pessoa deveria estar.
                equipe_ref = item.registered_team_history[-1]
            else:
                equipe_ref = ""

        info = {
            "primeira_partida": datas_proprias[0][:10] if datas_proprias else "",
            "ultima_partida": datas_proprias[-1][:10] if datas_proprias else "",
            "equipe_de_referencia": equipe_ref,
            "ultima_partida_dele_na_equipe_ref": "",
            "ultima_partida_da_equipe_ref": "",
            "partidas_da_equipe_ref_apos_a_dele": "",
            "partidas_da_equipe_ref_sem_ele": "",
            "total_partidas_da_equipe_ref": "",
            # Partidas em que a auditoria nao conseguiu apurar o lado. Enquanto
            # esse numero for alto, as metricas acima medem pouco: as partidas
            # dele nao entram na linha do tempo da equipe.
            "partidas_sem_equipe_apurada": len(item.teams_to_matches.get("unknown_team", set())),
        }
        if not equipe_ref:
            return info

        da_equipe = partidas_por_equipe.get(equipe_ref, {})
        info["total_partidas_da_equipe_ref"] = len(da_equipe)
        datas_equipe = sorted(d for d in da_equipe.values() if d)
        info["ultima_partida_da_equipe_ref"] = datas_equipe[-1][:10] if datas_equipe else ""

        dele_na_equipe = sorted(
            d for d, equipe in proprias.values() if d and equipe == equipe_ref
        )
        info["partidas_da_equipe_ref_sem_ele"] = len(da_equipe) - len(
            [1 for mid in da_equipe if mid in proprias]
        )
        if not dele_na_equipe:
            info["ultima_partida_dele_na_equipe_ref"] = "nunca"
            info["partidas_da_equipe_ref_apos_a_dele"] = len(da_equipe)
            return info

        ultima_dele = dele_na_equipe[-1]
        info["ultima_partida_dele_na_equipe_ref"] = ultima_dele[:10]
        info["partidas_da_equipe_ref_apos_a_dele"] = sum(
            1 for d in datas_equipe if d > ultima_dele
        )
        return info

    # CSV resumo
    summary_rows: List[Dict[str, Any]] = []
    for chave, item in summary.items():
        teams_summary_parts = []
        for team, match_ids in sorted(
            item.teams_to_matches.items(),
            key=lambda kv: (-len(kv[1]), kv[0]),
        ):
            n = len(match_ids)
            palavra = "partida" if n == 1 else "partidas"
            teams_summary_parts.append(f"{n} {palavra} com {team}")

        files_sorted = sorted(item.files)
        riot_ids_sorted = sorted(item.sample_riot_ids)

        summary_rows.append({
            "severidade": severidade_do_conjunto(item.issue_types),
            "riot_id": item.display_riot_id,
            "puuid": item.puuid or "",
            "total_partidas": len(item.total_matches),
            "aparicoes_por_time": "; ".join(teams_summary_parts),
            "issues": "; ".join(sorted(item.issue_types)),
            "registered_player": item.registered_player or "",
            "registered_current_team": item.registered_current_team or "",
            "historico_de_equipes": "; ".join(item.registered_team_history),
            "completes": "; ".join(item.registered_completes),
            **metricas_de_atividade(chave, item),
            "riot_ids_vistos": "; ".join(riot_ids_sorted),
            "arquivos_qtd": len(files_sorted),
            "arquivos_exemplo": "; ".join(files_sorted[:10]),
        })

    summary_rows.sort(key=lambda r: (
        ORDEM_SEVERIDADE.get(r["severidade"], 9),
        -int(r["total_partidas"]),
        r["riot_id"].casefold(),
    ))

    resumo_csv = output_dir / "nicks_fora_do_cadastro_resumo.csv"
    detalhes_csv = output_dir / "nicks_fora_do_cadastro_detalhado.csv"
    resumo_json = output_dir / "nicks_fora_do_cadastro_resumo.json"
    debug_json = output_dir / "team_detection_debug.json"

    with resumo_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "severidade",
                "riot_id",
                "puuid",
                "total_partidas",
                "aparicoes_por_time",
                "issues",
                "registered_player",
                "registered_current_team",
                "historico_de_equipes",
                "completes",
                "primeira_partida",
                "ultima_partida",
                "equipe_de_referencia",
                "ultima_partida_dele_na_equipe_ref",
                "ultima_partida_da_equipe_ref",
                "partidas_da_equipe_ref_apos_a_dele",
                "partidas_da_equipe_ref_sem_ele",
                "total_partidas_da_equipe_ref",
                "partidas_sem_equipe_apurada",
                "riot_ids_vistos",
                "arquivos_qtd",
                "arquivos_exemplo",
            ],
        )
        writer.writeheader()
        writer.writerows(linhas_seguras(summary_rows))

    with detalhes_csv.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "riot_id",
                "puuid",
                "issues",
                "associated_team",
                "side",
                "file_name",
                "match_id",
                "match_date_brt",
                "registered_player",
                "registered_current_team",
                "registered_team_history",
                "source_for_associated_team",
                "team_by_players_3plus",
                "team_by_filename_score",
                "filename_teamA",
                "filename_scoreA",
                "filename_scoreB",
                "filename_teamB",
            ],
        )
        writer.writeheader()
        writer.writerows(linhas_seguras(detail_rows))

    resumo_json.write_text(
        json.dumps(summary_rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    debug_json.write_text(
        json.dumps(debug_rows, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    result = {
        "players_path": str(players_path),
        "teams_path": str(teams_path),
        "matches_dir": str(matches_dir),
        "match_source": fonte,
        "source_dirs": [str(d) for d in pastas],
        "output_dir": str(output_dir),
        "processed_files": processed_files,
        "invalid_files": invalid_files,
        "arquivos_sem_jogadores": arquivos_sem_jogadores,
        "duplicatas": duplicatas,
        "total_active_players_seen": total_active_players_seen,
        "reported_unique_accounts": len(summary_rows),
        "contagem_por_issue": {
            issue: sum(1 for row in summary_rows if issue in row["issues"].split("; "))
            for issue in SEVERIDADE_POR_ISSUE
        },
        "contagem_por_severidade": {
            sev: sum(1 for row in summary_rows if row["severidade"] == sev)
            for sev in ORDEM_SEVERIDADE
        },
        "reported_occurrences": reported_occurrences,
        "summary_csv": str(resumo_csv),
        "details_csv": str(detalhes_csv),
        "summary_json": str(resumo_json),
        "debug_json": str(debug_json),
        "top_rows": summary_rows[:40],
    }

    return result


def print_console_summary(result: Dict[str, Any]) -> None:
    print("Fonte de partidas:", result.get("match_source", "-"))
    print("Pasta de partidas:")
    for d in result.get("source_dirs", []):
        print(" -", d)
    print("Arquivos JSON processados:", result["processed_files"])
    print("Arquivos JSON inválidos/ignorados:", result["invalid_files"])
    print("Arquivos sem jogadores (não são partida):", result.get("arquivos_sem_jogadores", 0))
    duplicatas = result.get("duplicatas") or []
    if duplicatas:
        print(f"Partidas gravadas em duplicata: {len(duplicatas)}")
        for d in duplicatas:
            print(f"   mantido    {d['mantido']}")
            print(f"   descartado {d['descartado']}   ({d['criterio']})")
    print("Jogadores ativos vistos nos JSONs:", result["total_active_players_seen"])
    print("Ocorrências reportadas:", result["reported_occurrences"])
    print("Contas únicas reportadas:", result["reported_unique_accounts"])
    print()
    print("Saídas:")
    print(" -", result["summary_csv"])
    print(" -", result["details_csv"])
    print(" -", result["summary_json"])
    print(" -", result["debug_json"])
    print()

    contagem = result.get("contagem_por_issue") or {}
    if contagem:
        print("Achados por tipo:")
        for issue, sev in SEVERIDADE_POR_ISSUE.items():
            n = contagem.get(issue, 0)
            marca = " " if n == 0 else ">"
            print(f" {marca} [{sev:5}] {issue:30} {n:4}   {DESCRICAO_POR_ISSUE.get(issue, '')}")
        print()

    top_rows = result.get("top_rows") or []
    if not top_rows:
        print("Nada a corrigir: toda conta vista bate com o cadastro.")
        return

    for sev in ORDEM_SEVERIDADE:
        linhas = [row for row in top_rows if row["severidade"] == sev]
        if not linhas:
            continue
        print(f"--- {sev} ---")
        for row in linhas:
            print(f"{row['riot_id']}   {row['aparicoes_por_time']}   [{row['issues']}]")
            cadastro = row.get("registered_player") or "(sem linha no players.xlsx)"
            atual = row.get("registered_current_team") or "-"
            historico = row.get("historico_de_equipes") or "-"
            print(f"    cadastro: {cadastro} | current_team: {atual} | histórico: {historico}")
            print(f"    jogou de {row.get('primeira_partida') or '?'} a {row.get('ultima_partida') or '?'}")
            ref = row.get("equipe_de_referencia")
            if ref:
                depois = row.get("partidas_da_equipe_ref_apos_a_dele")
                print(
                    f"    {ref}: última dele {row.get('ultima_partida_dele_na_equipe_ref') or '?'}"
                    f" | última da equipe {row.get('ultima_partida_da_equipe_ref') or '?'}"
                    f" | {depois} partida(s) da equipe depois disso"
                    f" (de {row.get('total_partidas_da_equipe_ref')} no total)"
                )
        print()


# =========================
# Execução direta no VS Code
# =========================
# A pasta de trabalho fica fora do repo: e la que estao os JSONs brutos e as
# saidas de analise. Ver config_hub.
BASE_DIR = config_hub.TRABALHO_DIR
DEFAULT_PLAYERS_XLSX = config_hub.PLAYERS_XLSX
DEFAULT_TEAMS_XLSX = config_hub.TEAMS_XLSX
DEFAULT_MATCHES_DIR = BASE_DIR / "Finalizados"
DEFAULT_OUTPUT_DIR = BASE_DIR / "out_auditoria_nicks"
DEFAULT_RECURSIVE = True


def main() -> None:
    result = audit_matches(
        players_path=DEFAULT_PLAYERS_XLSX,
        teams_path=DEFAULT_TEAMS_XLSX,
        matches_dir=DEFAULT_MATCHES_DIR,
        output_dir=DEFAULT_OUTPUT_DIR,
        min_team_players=MIN_TEAM_PLAYERS_TO_IDENTIFY,
        recursive=DEFAULT_RECURSIVE,
        match_source=choose_match_source_runtime(),
    )
    print_console_summary(result)


if __name__ == "__main__":
    main()
