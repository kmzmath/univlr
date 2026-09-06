import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ---- CONFIG PARA RODAR DIRETO NO VS CODE ----
import config_hub

# A pasta de trabalho fica fora do repo: e la que estao os JSONs brutos e as
# saidas de analise. Ver config_hub.
BASE_DIR = config_hub.TRABALHO_DIR

# Edite esta pasta para o campeonato que você quer analisar.
# Exemplos:
# DEFAULT_INPUT_DIR = str(BASE_DIR / "Finalizados" / "Válidos" / "CIA 2026")
# DEFAULT_INPUT_DIR = str(BASE_DIR / "Finalizados" / "Válidos" / "UNI Ascension")
DEFAULT_INPUT_DIR = str(BASE_DIR / "Finalizados" / "Válidos" / "Univavá" / "Classificatórias 1")

# Se vazio, gera automaticamente: analiseAgentes_<nome_da_pasta>.xlsx
DEFAULT_OUTPUT_XLSX = ""
DEFAULT_RECURSIVE = True
DEFAULT_TRADE_WINDOW_MS = 2000
DEFAULT_STATE_WINRATES_XLSX = str(config_hub.ROUND_STATE_WINRATES_XLSX)
OPEN_FILE_AFTER_SAVE = True


# ---- MAPAS: internalName/uuid -> displayName ----
MAP_INTERNAL_TO_DISPLAY = {
    "Infinity": "Abyss",
    "Ascent": "Ascent",
    "Duality": "Bind",
    "Foxtrot": "Breeze",
    "Rook": "Corrode",
    "Canyon": "Fracture",
    "Triad": "Haven",
    "Port": "Icebox",
    "Jam": "Lotus",
    "Pitt": "Pearl",
    "Bonsai": "Split",
    "Juliett": "Sunset",
}

MAP_UUID_TO_DISPLAY = {
    "224b0a95-48b9-f703-1bd8-67aca101a61f": "Abyss",
    "7eaecc1b-4337-bbf6-6ab9-04b8f06b3319": "Ascent",
    "2c9d57ec-4431-9c5e-2939-8f9ef6dd5cba": "Bind",
    "2fb9a4fd-47b8-4e7d-a969-74b4046ebd53": "Breeze",
    "1c18ab1f-420d-0d8b-71d0-77ad3c439115": "Corrode",
    "b529448b-4d60-346e-e89e-00a4c527a405": "Fracture",
    "2bee0dc9-4ffe-519b-1cbd-7fbe763a6047": "Haven",
    "e2ad5c54-4114-a870-9641-8ea21279579a": "Icebox",
    "2fe4ed3a-450a-948b-6d6b-e89a78e680a9": "Lotus",
    "fd267378-4d1d-484f-ff52-77821ed10dc2": "Pearl",
    "d960549e-485c-e861-8d71-aa9d1aed12a2": "Split",
    "92584fbe-486a-b1b2-9faa-39b0f486b498": "Sunset",
}


# ---- AGENTES: uuid -> displayName ----
AGENT_ID_TO_NAME: Dict[str, str] = {
    "41fb69c1-4189-7b37-f117-bcaf1e96f1bf": "Astra",
    "5f8d3a7f-467b-97f3-062c-13acf203c006": "Breach",
    "9f0d8ba9-4140-b941-57d3-a7ad57c6b417": "Brimstone",
    "22697a3d-45bf-8dd7-4fec-84a9e28c69d7": "Chamber",
    "1dbf2edd-4729-0984-3115-daa5eed44993": "Clove",
    "117ed9e3-49f3-6512-3ccf-0cada7e3823b": "Cypher",
    "cc8b64c8-4b25-4ff9-6e7f-37b4da43d235": "Deadlock",
    "dade69b4-4f5a-8528-247b-219e5a1facd6": "Fade",
    "e370fa57-4757-3604-3648-499e1f642d3f": "Gekko",
    "95b78ed7-4637-86d9-7e41-71ba8c293152": "Harbor",
    "0e38b510-41a8-5780-5e8f-568b2a4f2d6c": "Iso",
    "add6443a-41bd-e414-f6ad-e58d267f4e95": "Jett",
    "601dbbe7-43ce-be57-2a40-4abd24953621": "KAY/O",
    "1e58de9c-4950-5125-93e9-a0aee9f98746": "Killjoy",
    "7c8a4701-4de6-9355-b254-e09bc2a34b72": "Miks",
    "bb2a4828-46eb-8cd1-e765-15848195d751": "Neon",
    "8e253930-4c05-31dd-1b6c-968525494517": "Omen",
    "eb93336a-449b-9c1b-0a54-a891f7921d69": "Phoenix",
    "f94c3b30-42be-e959-889c-5aa313dba261": "Raze",
    "a3bfb853-43b2-7238-a4f1-ad90e9e46bcc": "Reyna",
    "569fdd95-4d10-43ab-ca70-79becc718b46": "Sage",
    "6f2a04ca-43e0-be17-7f36-b3908627744d": "Skye",
    "320b2a48-4d9b-a075-30f1-1f93a9b638fa": "Sova",
    "b444168c-4e35-8076-db47-ef9bf368f384": "Tejo",
    "92eeef5d-43b5-1d4a-8d03-b3927a09034b": "Veto",
    "707eab51-4836-f488-046a-cda6bf494859": "Viper",
    "efba5359-4016-a1e5-7626-b1ae76895940": "Vyse",
    "df1cb487-4902-002e-5c17-d28e83e78588": "Waylay",
    "7f94d92c-4234-0a36-9646-3a87eb8b5c89": "Yoru",
}

AGENT_TO_ROLE = {
    "Brimstone": "Controlador",
    "Viper": "Controlador",
    "Omen": "Controlador",
    "Astra": "Controlador",
    "Harbor": "Controlador",
    "Clove": "Controlador",
    "Miks": "Controlador",
    "Phoenix": "Duelista",
    "Jett": "Duelista",
    "Reyna": "Duelista",
    "Raze": "Duelista",
    "Yoru": "Duelista",
    "Neon": "Duelista",
    "Iso": "Duelista",
    "Waylay": "Duelista",
    "Sova": "Iniciador",
    "Breach": "Iniciador",
    "Skye": "Iniciador",
    "KAY/O": "Iniciador",
    "Fade": "Iniciador",
    "Gekko": "Iniciador",
    "Tejo": "Iniciador",
    "Killjoy": "Sentinela",
    "Cypher": "Sentinela",
    "Sage": "Sentinela",
    "Chamber": "Sentinela",
    "Deadlock": "Sentinela",
    "Vyse": "Sentinela",
    "Veto": "Sentinela",
}

ARMOR_ID_TO_EXTRA_HP = {
    "": 0,
    "4DEC83D5-4902-9AB3-BED6-A7A390761157": 25,  # Light
    "822BCAB2-40A2-324E-C137-E09195AD7692": 50,  # Heavy
}

SUMMARY_HEADERS = [
    "Agente",
    "Função",
    "Picks",
    "Pickrate",
    "Map Win%",
    "Round Win%",
    "KAST",
    "Rounds",
    "R.Ganhos",
    "R.Perdidos",
    "R.KAST",
    "ACS",
    "Kills",
    "Deaths",
    "Assists",
    "KPR",
    "DPR",
    "APR",
    "ADR",
    "FK",
    "FD",
    "FK-FD",
    "1K",
    "2K",
    "3K",
    "4K",
    "5K",
    "Clutches",
    "Imp.Total (pp)",
    "Imp/Round (pp)",
    "Rating",
]

MAP_HEADERS = ["Mapa"] + SUMMARY_HEADERS

DETAIL_HEADERS = [
    "Arquivo",
    "Player",
    "Riot ID",
    "Side",
    "Agente",
    "Função",
    "Mapa",
    "Map Win",
    "KAST",
    "Rounds",
    "R.Ganhos",
    "R.Perdidos",
    "R.KAST",
    "ACS",
    "Kills",
    "Deaths",
    "Assists",
    "KPR",
    "DPR",
    "APR",
    "ADR",
    "FK",
    "FD",
    "1K",
    "2K",
    "3K",
    "4K",
    "5K",
    "Clutches",
    "Imp.Total (pp)",
    "Imp/Round (pp)",
    "Rating",
]


def safe_div(n: float, d: float) -> float:
    return (n / d) if d else 0.0


def resolve_user_path(raw: str, base_dir: Path = BASE_DIR) -> Path:
    s = str(raw or "").strip()
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        s = s[1:-1].strip()
    p = Path(s).expanduser()
    if not p.is_absolute():
        p = base_dir / p
    return p.resolve()


def build_default_output_path(input_dir: Path) -> Path:
    folder_name = input_dir.name.strip() or "campeonato"
    safe_name = re.sub(r"[^A-Za-z0-9À-ÿ._ -]+", "_", folder_name).strip(" ._-") or "campeonato"
    safe_name = re.sub(r"\s+", "_", safe_name)
    return BASE_DIR / f"analiseAgentes_{safe_name}.xlsx"


def iter_json_files(input_dir: Path, recursive: bool) -> Iterable[Path]:
    pattern = "**/*.json" if recursive else "*.json"
    yield from sorted(input_dir.glob(pattern))


def load_match(path: Path) -> Optional[Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return None
        if "matchInfo" not in data or "players" not in data or "roundResults" not in data:
            return None
        return data
    except Exception:
        return None


def riot_id(game_name: Optional[str], tag_line: Optional[str]) -> str:
    gn = (game_name or "").strip()
    tl = (tag_line or "").strip()
    if gn and tl:
        return f"{gn}#{tl}"
    return gn or ""


def map_name_from_map_id(map_id: Optional[str]) -> str:
    if not map_id:
        return ""
    s = str(map_id).strip()
    if s in MAP_UUID_TO_DISPLAY:
        return MAP_UUID_TO_DISPLAY[s]
    parts = [p for p in s.split("/") if p]
    internal = parts[-1] if parts else s
    return MAP_INTERNAL_TO_DISPLAY.get(internal, internal)


def agent_name_from_character_id(character_id: Optional[str]) -> str:
    if not character_id:
        return ""
    cid = str(character_id).strip()
    return AGENT_ID_TO_NAME.get(cid, cid)


def armor_extra_hp(armor_id: Any) -> int:
    a = str(armor_id or "").strip().upper()
    return ARMOR_ID_TO_EXTRA_HP.get(a, 50)


def rating_value(kast_frac: float, kpr: float, dpr: float, apr: float, adr: float) -> float:
    kast_pct = kast_frac * 100.0
    return (
        0.0073 * kast_pct
        + 0.3591 * kpr
        - 0.5329 * dpr
        + (0.2372 * (2.13 * kpr + 0.42 * apr - 0.41))
        + 0.0032 / 1.5 * adr
        + 0.1587
    )


@dataclass(frozen=True)
class KillEvent:
    t_ms: int
    killer: str
    victim: str
    assistants: Tuple[str, ...]


@dataclass
class AgentPickRecord:
    file_name: str
    player: str
    riot_id: str
    side: str
    agent: str
    role: str
    map_name: str
    map_win: int
    rounds: int
    rounds_won: int
    rounds_lost: int
    kast_rounds: int
    acs: float
    kills: int
    deaths: int
    assists: int
    kpr: float
    dpr: float
    apr: float
    adr: float
    fk: int
    fd: int
    k1: int
    k2: int
    k3: int
    k4: int
    k5: int
    clutches: int
    impact_total: float
    impact_per_round: float
    rating: float

    @property
    def kast(self) -> float:
        return safe_div(self.kast_rounds, self.rounds)

    def detail_row(self) -> List[Any]:
        return [
            self.file_name,
            self.player,
            self.riot_id,
            self.side,
            self.agent,
            self.role,
            self.map_name,
            self.map_win,
            self.kast,
            self.rounds,
            self.rounds_won,
            self.rounds_lost,
            self.kast_rounds,
            self.acs,
            self.kills,
            self.deaths,
            self.assists,
            self.kpr,
            self.dpr,
            self.apr,
            self.adr,
            self.fk,
            self.fd,
            self.k1,
            self.k2,
            self.k3,
            self.k4,
            self.k5,
            self.clutches,
            self.impact_total,
            self.impact_per_round,
            self.rating,
        ]


# ---------------- Winrates por estado XvY ----------------

def _parse_win_percent(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        x = float(v)
        if x <= 1.0:
            return max(0.0, min(1.0, x))
        return max(0.0, min(1.0, x / 100.0))
    s = str(v).strip().replace(",", ".")
    if not s:
        return 0.0
    if s.endswith("%"):
        try:
            return max(0.0, min(1.0, float(s[:-1].strip()) / 100.0))
        except Exception:
            return 0.0
    try:
        x = float(s)
        if x <= 1.0:
            return max(0.0, min(1.0, x))
        return max(0.0, min(1.0, x / 100.0))
    except Exception:
        return 0.0


def _parse_situacao(s: Any) -> Optional[Tuple[int, int]]:
    if s is None:
        return None
    txt = str(s).strip().lower()
    m = re.match(r"^\s*(\d)\s*v\s*(\d)\s*$", txt)
    if not m:
        return None
    x = int(m.group(1))
    y = int(m.group(2))
    if not (1 <= x <= 5 and 1 <= y <= 5):
        return None
    return (x, y)


def load_state_winrates_xlsx(path: Path) -> Dict[Tuple[int, int], float]:
    """
    Carrega round_state_winrates.xlsx. Se não existir, usa fallback neutro.
    Com fallback neutro, impacto fica praticamente 0, mas o script continua rodando.
    """
    if not path.exists():
        print(f"[WARNING] Arquivo de winrates não encontrado: {path}")
        print("[WARNING] Impacto será calculado com fallback neutro.")
        return {}

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print(f"[WARNING] Arquivo de winrates vazio: {path}")
        return {}

    headers = [str(h or "").strip() for h in rows[0]]
    hmap = {h.casefold(): i for i, h in enumerate(headers)}

    idx_sit = None
    idx_winp = None
    for k, i in hmap.items():
        if k in ("situação", "situacao", "situacão"):
            idx_sit = i
        if k in ("win%", "win %", "winrate", "win rate", "win"):
            idx_winp = i

    if idx_sit is None:
        print(f"[WARNING] Coluna 'Situação' não encontrada em {path}. Impacto usará fallback neutro.")
        return {}

    if idx_winp is None:
        for i, h in enumerate(headers):
            if "win" in h.casefold() and "%" in h:
                idx_winp = i
                break

    if idx_winp is None:
        print(f"[WARNING] Coluna 'Win%' não encontrada em {path}. Impacto usará fallback neutro.")
        return {}

    out: Dict[Tuple[int, int], float] = {}
    for r in rows[1:]:
        if not r or idx_sit >= len(r):
            continue
        key = _parse_situacao(r[idx_sit])
        if not key:
            continue
        v = r[idx_winp] if idx_winp < len(r) else None
        out[key] = _parse_win_percent(v)

    return out


def get_state_wr(state_wr: Dict[Tuple[int, int], float], x: int, y: int) -> float:
    if x <= 0:
        return 0.0
    if y <= 0:
        return 1.0
    return state_wr.get((x, y), 0.5)


def delta_kill_pp(state_wr: Dict[Tuple[int, int], float], x: int, y: int) -> float:
    before = get_state_wr(state_wr, x, y)
    after = get_state_wr(state_wr, x, y - 1)
    return (after - before) * 100.0


def delta_death_pp(state_wr: Dict[Tuple[int, int], float], x: int, y: int) -> float:
    before = get_state_wr(state_wr, x, y)
    after = get_state_wr(state_wr, x - 1, y)
    return (after - before) * 100.0


# ---------------- Core ----------------

def build_agent_records_for_match(
    data: Dict[str, Any],
    filename: str,
    trade_window_ms: int,
    state_wr: Dict[Tuple[int, int], float],
) -> List[AgentPickRecord]:
    match_info = data.get("matchInfo") or {}
    map_name = map_name_from_map_id(match_info.get("mapId")) or "DESCONHECIDO"

    players = data.get("players") or []
    round_results = data.get("roundResults") or []
    rounds_total_match = len(round_results)

    side_by_puuid: Dict[str, str] = {}
    meta_by_puuid: Dict[str, Dict[str, Any]] = {}

    for p in players:
        if not isinstance(p, dict):
            continue
        if p.get("isObserver"):
            continue
        puuid = p.get("puuid")
        side = p.get("teamId")
        if puuid and side in ("Blue", "Red"):
            puuid = str(puuid)
            side_by_puuid[puuid] = str(side)
            meta_by_puuid[puuid] = p

    roster_blue = {pu for pu, side in side_by_puuid.items() if side == "Blue"}
    roster_red = {pu for pu, side in side_by_puuid.items() if side == "Red"}

    rounds_seen = defaultdict(int)
    rounds_won = defaultdict(int)
    rounds_lost = defaultdict(int)
    kast_rounds = defaultdict(int)
    dmg_total = defaultdict(float)
    fk = defaultdict(int)
    fd = defaultdict(int)
    k1 = defaultdict(int)
    k2 = defaultdict(int)
    k3 = defaultdict(int)
    k4 = defaultdict(int)
    k5 = defaultdict(int)
    clutches = defaultdict(int)
    impact_total_pp = defaultdict(float)

    for rr in round_results:
        winning_side = rr.get("winningTeam")
        player_stats = rr.get("playerStats") or []

        armor_extra_by_puuid: Dict[str, int] = {}
        for ps in player_stats:
            pu = ps.get("puuid")
            if not pu:
                continue
            pu = str(pu)
            if pu not in side_by_puuid:
                continue
            armor_id = (ps.get("economy") or {}).get("armor")
            armor_extra_by_puuid[pu] = armor_extra_hp(armor_id)

        damage_by_victim: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))
        damage_by_receiver: Dict[str, Set[str]] = defaultdict(set)
        participants: Set[str] = set()
        kills_count_round: Dict[str, int] = defaultdict(int)
        events: List[KillEvent] = []
        events_seen: Set[Tuple[int, str, str]] = set()
        assistant_set: Set[str] = set()
        victims_set: Set[str] = set()

        for ps in player_stats:
            puuid = ps.get("puuid")
            if not puuid:
                continue
            puuid = str(puuid)
            if puuid not in side_by_puuid:
                continue

            participants.add(puuid)

            for d in (ps.get("damage") or []):
                receiver = d.get("receiver")
                dmg = d.get("damage", 0) or 0
                if not receiver or dmg <= 0:
                    continue
                receiver = str(receiver)
                if receiver not in side_by_puuid:
                    continue
                if receiver == puuid:
                    continue
                if side_by_puuid.get(receiver) == side_by_puuid.get(puuid):
                    continue
                damage_by_victim[receiver][puuid] += float(dmg)
                damage_by_receiver[receiver].add(puuid)

            kills_list = ps.get("kills") or []
            kills_count_round[puuid] += len(kills_list)
            for k in kills_list:
                t = k.get("timeSinceRoundStartMillis")
                if t is None:
                    continue
                killer = str(k.get("killer") or "")
                victim = str(k.get("victim") or "")
                assistants = tuple(str(a) for a in (k.get("assistants") or []))

                if victim and victim not in side_by_puuid:
                    continue
                if killer and killer not in side_by_puuid:
                    killer = ""

                key = (int(t), killer, victim)
                if key in events_seen:
                    continue
                events_seen.add(key)

                events.append(KillEvent(int(t), killer, victim, assistants))
                if victim:
                    victims_set.add(victim)
                for a in assistants:
                    if a in side_by_puuid:
                        assistant_set.add(a)

        for pu in participants:
            rounds_seen[pu] += 1
            if winning_side in ("Blue", "Red"):
                if side_by_puuid[pu] == winning_side:
                    rounds_won[pu] += 1
                else:
                    rounds_lost[pu] += 1

        events.sort(key=lambda e: e.t_ms)

        if events:
            min_t = events[0].t_ms
            for e in events:
                if e.t_ms != min_t:
                    break
                if e.killer and e.killer in side_by_puuid:
                    fk[e.killer] += 1
                if e.victim in side_by_puuid:
                    fd[e.victim] += 1

        killer_of_victim: Dict[str, str] = {}
        killer_time: Dict[str, int] = {}
        for e in events:
            if not e.victim or e.victim not in side_by_puuid:
                continue
            if not e.killer or e.killer not in side_by_puuid:
                continue
            prev_t = killer_time.get(e.victim)
            if prev_t is None or e.t_ms > prev_t:
                killer_time[e.victim] = e.t_ms
                killer_of_victim[e.victim] = e.killer

        # Dano efetivo com cap por vida+escudo, igual à lógica do analiseMatches.
        for victim, by_attacker in damage_by_victim.items():
            raw_total = sum(by_attacker.values())
            if raw_total <= 0:
                continue

            cap = 100 + armor_extra_by_puuid.get(victim, 50)
            over = raw_total - float(cap)
            effective = dict(by_attacker)

            if over > 0:
                last = killer_of_victim.get(victim)
                if not last or last not in effective:
                    last = max(effective.items(), key=lambda kv: kv[1])[0]

                take = min(over, effective.get(last, 0.0))
                effective[last] = max(0.0, effective.get(last, 0.0) - take)
                over -= take

                if over > 0:
                    for attacker, _ in sorted(effective.items(), key=lambda kv: kv[1], reverse=True):
                        if attacker == last:
                            continue
                        if over <= 0:
                            break
                        take = min(over, effective[attacker])
                        effective[attacker] = max(0.0, effective[attacker] - take)
                        over -= take

            for attacker, dmg in effective.items():
                if dmg > 0:
                    dmg_total[attacker] += dmg

        # Impacto por kill/death usando estados XvY.
        alive_blue_i = set(roster_blue)
        alive_red_i = set(roster_red)
        for e in events:
            victim = e.victim
            killer = e.killer
            if not victim or victim not in side_by_puuid:
                continue

            if victim in alive_blue_i:
                v_side = "Blue"
            elif victim in alive_red_i:
                v_side = "Red"
            else:
                continue

            if v_side == "Blue":
                x_v = len(alive_blue_i)
                y_v = len(alive_red_i)
            else:
                x_v = len(alive_red_i)
                y_v = len(alive_blue_i)
            impact_total_pp[victim] += delta_death_pp(state_wr, x_v, y_v)

            if killer and killer in side_by_puuid:
                k_side = side_by_puuid[killer]
                if k_side == "Blue":
                    x_k = len(alive_blue_i)
                    y_k = len(alive_red_i)
                else:
                    x_k = len(alive_red_i)
                    y_k = len(alive_blue_i)
                impact_total_pp[killer] += delta_kill_pp(state_wr, x_k, y_k)

            if v_side == "Blue":
                alive_blue_i.remove(victim)
            else:
                alive_red_i.remove(victim)

        for pu in participants:
            kc = kills_count_round.get(pu, 0)
            if kc == 1:
                k1[pu] += 1
            elif kc == 2:
                k2[pu] += 1
            elif kc == 3:
                k3[pu] += 1
            elif kc == 4:
                k4[pu] += 1
            elif kc == 5:
                k5[pu] += 1

        deaths_by_victim: Dict[str, List[KillEvent]] = defaultdict(list)
        for e in events:
            if e.victim:
                deaths_by_victim[e.victim].append(e)

        for pu in participants:
            kill_cond = kills_count_round.get(pu, 0) > 0
            assist_cond = pu in assistant_set
            survived_cond = pu not in victims_set

            traded_cond = False
            if not survived_cond:
                victim_side = side_by_puuid.get(pu)
                for death in deaths_by_victim.get(pu, []):
                    t0 = death.t_ms
                    t1 = t0 + trade_window_ms
                    involved: Set[str] = set()
                    if death.killer:
                        involved.add(death.killer)
                    involved.update(death.assistants)
                    involved.update(damage_by_receiver.get(pu, set()))
                    involved = {x for x in involved if x in side_by_puuid and side_by_puuid.get(x) != victim_side}
                    if not involved:
                        continue
                    for ev in events:
                        if ev.t_ms <= t0:
                            continue
                        if ev.t_ms > t1:
                            break
                        if side_by_puuid.get(ev.killer) == victim_side and ev.victim in involved:
                            traded_cond = True
                            break
                    if traded_cond:
                        break

            if kill_cond or assist_cond or survived_cond or traded_cond:
                kast_rounds[pu] += 1

        alive_blue = set(roster_blue)
        alive_red = set(roster_red)
        clutch_candidate_blue: Optional[str] = None
        clutch_candidate_red: Optional[str] = None

        for e in events:
            if not e.victim or e.victim not in side_by_puuid:
                continue
            vside = side_by_puuid[e.victim]
            if vside == "Blue":
                alive_blue.discard(e.victim)
                if clutch_candidate_blue is None and len(alive_blue) == 1:
                    last = next(iter(alive_blue))
                    if len(alive_red) > 1:
                        clutch_candidate_blue = last
            else:
                alive_red.discard(e.victim)
                if clutch_candidate_red is None and len(alive_red) == 1:
                    last = next(iter(alive_red))
                    if len(alive_blue) > 1:
                        clutch_candidate_red = last

        if winning_side == "Blue" and clutch_candidate_blue and clutch_candidate_blue not in victims_set:
            clutches[clutch_candidate_blue] += 1
        elif winning_side == "Red" and clutch_candidate_red and clutch_candidate_red not in victims_set:
            clutches[clutch_candidate_red] += 1

    out: List[AgentPickRecord] = []
    for puuid, p in meta_by_puuid.items():
        stats = p.get("stats") or {}
        rounds_played = int(stats.get("roundsPlayed") or rounds_seen.get(puuid, 0))
        if rounds_played <= 0:
            continue

        agent = agent_name_from_character_id(p.get("characterId")) or "DESCONHECIDO"
        role = AGENT_TO_ROLE.get(agent, "DESCONHECIDO")
        side = side_by_puuid.get(puuid, "")
        r_won = int(rounds_won.get(puuid, 0))
        r_lost = int(rounds_lost.get(puuid, 0))
        map_win = 1 if r_won > r_lost else 0

        score = float(stats.get("score") or 0)
        kills = int(stats.get("kills") or 0)
        deaths = int(stats.get("deaths") or 0)
        assists = int(stats.get("assists") or 0)
        kast_cnt = int(kast_rounds.get(puuid, 0))
        kast_frac = safe_div(kast_cnt, rounds_played)
        acs = safe_div(score, rounds_played)
        kpr = safe_div(kills, rounds_played)
        dpr = safe_div(deaths, rounds_played)
        apr = safe_div(assists, rounds_played)
        adr = safe_div(float(dmg_total.get(puuid, 0.0)), rounds_played)
        imp_total = float(impact_total_pp.get(puuid, 0.0))
        imp_pr = safe_div(imp_total, rounds_played)
        rtg = rating_value(kast_frac, kpr, dpr, apr, adr)

        rid = riot_id(p.get("gameName"), p.get("tagLine"))
        player_display = rid or puuid
        player_name = str(p.get("gameName") or "").strip() or player_display

        out.append(
            AgentPickRecord(
                file_name=filename,
                player=player_name,
                riot_id=rid,
                side=side,
                agent=agent,
                role=role,
                map_name=map_name,
                map_win=map_win,
                rounds=rounds_total_match or rounds_played,
                rounds_won=r_won,
                rounds_lost=r_lost,
                kast_rounds=kast_cnt,
                acs=acs,
                kills=kills,
                deaths=deaths,
                assists=assists,
                kpr=kpr,
                dpr=dpr,
                apr=apr,
                adr=adr,
                fk=int(fk.get(puuid, 0)),
                fd=int(fd.get(puuid, 0)),
                k1=int(k1.get(puuid, 0)),
                k2=int(k2.get(puuid, 0)),
                k3=int(k3.get(puuid, 0)),
                k4=int(k4.get(puuid, 0)),
                k5=int(k5.get(puuid, 0)),
                clutches=int(clutches.get(puuid, 0)),
                impact_total=imp_total,
                impact_per_round=imp_pr,
                rating=rtg,
            )
        )

    return out


# ---------------- Agregação ----------------

def aggregate_records(records: List[AgentPickRecord], pickrate_denominator: int, agent_name: str, role: str) -> List[Any]:
    picks = len(records)
    total_rounds = sum(r.rounds for r in records)
    total_wins = sum(r.rounds_won for r in records)
    total_losses = sum(r.rounds_lost for r in records)
    total_kast = sum(r.kast_rounds for r in records)
    total_kills = sum(r.kills for r in records)
    total_deaths = sum(r.deaths for r in records)
    total_assists = sum(r.assists for r in records)
    total_fk = sum(r.fk for r in records)
    total_fd = sum(r.fd for r in records)
    total_1k = sum(r.k1 for r in records)
    total_2k = sum(r.k2 for r in records)
    total_3k = sum(r.k3 for r in records)
    total_4k = sum(r.k4 for r in records)
    total_5k = sum(r.k5 for r in records)
    total_clutches = sum(r.clutches for r in records)
    total_impact = sum(r.impact_total for r in records)
    map_wins = sum(r.map_win for r in records)

    acs = safe_div(sum(r.acs * r.rounds for r in records), total_rounds)
    adr = safe_div(sum(r.adr * r.rounds for r in records), total_rounds)
    kast = safe_div(total_kast, total_rounds)
    kpr = safe_div(total_kills, total_rounds)
    dpr = safe_div(total_deaths, total_rounds)
    apr = safe_div(total_assists, total_rounds)
    imp_pr = safe_div(total_impact, total_rounds)
    rtg = rating_value(kast, kpr, dpr, apr, adr)

    return [
        agent_name,
        role,
        picks,
        safe_div(picks, pickrate_denominator),
        safe_div(map_wins, picks),
        safe_div(total_wins, total_rounds),
        kast,
        total_rounds,
        total_wins,
        total_losses,
        total_kast,
        acs,
        total_kills,
        total_deaths,
        total_assists,
        kpr,
        dpr,
        apr,
        adr,
        total_fk,
        total_fd,
        total_fk - total_fd,
        total_1k,
        total_2k,
        total_3k,
        total_4k,
        total_5k,
        total_clutches,
        total_impact,
        imp_pr,
        rtg,
    ]


def aggregate_role_records(records: List[AgentPickRecord], pickrate_denominator: int, role_name: str) -> List[Any]:
    # Reutiliza formato de summary; coluna Agente recebe nome da função.
    return aggregate_records(records, pickrate_denominator, role_name, role_name)


# ---------------- Excel ----------------

def write_headers(ws, headers: List[str]) -> None:
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")


def apply_summary_formats(ws, start_row: int, end_row: int, offset: int = 0) -> None:
    if end_row < start_row:
        return
    # SUMMARY_HEADERS columns, with optional offset for Map in agent_by_map.
    pct_cols = [4, 5, 6, 7]
    decimal_2_cols = [12, 16, 17, 18, 19, 29, 30]
    rating_col = 31

    for r in range(start_row, end_row + 1):
        for c in pct_cols:
            ws.cell(row=r, column=c + offset).number_format = "0.0%"
        for c in decimal_2_cols:
            ws.cell(row=r, column=c + offset).number_format = "0.00"
        ws.cell(row=r, column=rating_col + offset).number_format = "0.000"


def apply_detail_formats(ws, start_row: int, end_row: int) -> None:
    if end_row < start_row:
        return
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=9).number_format = "0.0%"   # KAST
        ws.cell(row=r, column=14).number_format = "0.00"  # ACS
        for c in (18, 19, 20, 21, 30, 31):
            ws.cell(row=r, column=c).number_format = "0.00"
        ws.cell(row=r, column=32).number_format = "0.000"


def auto_width(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 45)


def write_workbook(records: List[AgentPickRecord], output_path: Path) -> None:
    wb = Workbook()
    total_picks = len(records)

    # 1) agent_summary
    ws = wb.active
    ws.title = "agent_summary"
    write_headers(ws, SUMMARY_HEADERS)

    by_agent: Dict[str, List[AgentPickRecord]] = defaultdict(list)
    for r in records:
        by_agent[r.agent].append(r)

    rows = []
    for agent, recs in by_agent.items():
        role = AGENT_TO_ROLE.get(agent, recs[0].role if recs else "DESCONHECIDO")
        rows.append(aggregate_records(recs, total_picks, agent, role))
    rows.sort(key=lambda row: (-row[2], row[0].casefold()))
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{ws.max_row}"
    apply_summary_formats(ws, 2, ws.max_row)
    auto_width(ws)

    # 2) agent_by_map
    ws_map = wb.create_sheet("agent_by_map")
    write_headers(ws_map, MAP_HEADERS)
    by_agent_map: Dict[Tuple[str, str], List[AgentPickRecord]] = defaultdict(list)
    for r in records:
        by_agent_map[(r.map_name, r.agent)].append(r)
    map_rows = []
    for (map_name, agent), recs in by_agent_map.items():
        role = AGENT_TO_ROLE.get(agent, recs[0].role if recs else "DESCONHECIDO")
        map_rows.append([map_name] + aggregate_records(recs, total_picks, agent, role))
    map_rows.sort(key=lambda row: (str(row[0]).casefold(), -row[3], str(row[1]).casefold()))
    for row in map_rows:
        ws_map.append(row)
    ws_map.freeze_panes = "A2"
    ws_map.auto_filter.ref = f"A1:{get_column_letter(len(MAP_HEADERS))}{ws_map.max_row}"
    apply_summary_formats(ws_map, 2, ws_map.max_row, offset=1)
    auto_width(ws_map)

    # 3) role_summary
    ws_role = wb.create_sheet("role_summary")
    write_headers(ws_role, SUMMARY_HEADERS)
    by_role: Dict[str, List[AgentPickRecord]] = defaultdict(list)
    for r in records:
        by_role[r.role].append(r)
    role_rows = []
    for role, recs in by_role.items():
        role_rows.append(aggregate_role_records(recs, total_picks, role))
    role_rows.sort(key=lambda row: (-row[2], row[0].casefold()))
    for row in role_rows:
        ws_role.append(row)
    ws_role.freeze_panes = "A2"
    ws_role.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}{ws_role.max_row}"
    apply_summary_formats(ws_role, 2, ws_role.max_row)
    auto_width(ws_role)

    # 4) raw_picks
    ws_raw = wb.create_sheet("raw_picks")
    write_headers(ws_raw, DETAIL_HEADERS)
    for r in sorted(records, key=lambda x: (x.file_name.casefold(), x.agent.casefold(), x.player.casefold())):
        ws_raw.append(r.detail_row())
    ws_raw.freeze_panes = "A2"
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_HEADERS))}{ws_raw.max_row}"
    apply_detail_formats(ws_raw, 2, ws_raw.max_row)
    auto_width(ws_raw)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def open_file_after_save(path: Path) -> None:
    if not OPEN_FILE_AFTER_SAVE:
        return
    p = str(path)
    try:
        if sys.platform.startswith("win"):
            os.startfile(p)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", p], check=False)
        else:
            subprocess.run(["xdg-open", p], check=False)
    except Exception as e:
        print(f"Não foi possível abrir o arquivo automaticamente: {e}")


def main() -> int:
    input_dir = resolve_user_path(DEFAULT_INPUT_DIR)
    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"Pasta inválida: {input_dir}")

    output_xlsx = resolve_user_path(DEFAULT_OUTPUT_XLSX) if DEFAULT_OUTPUT_XLSX else build_default_output_path(input_dir)
    state_wr = load_state_winrates_xlsx(resolve_user_path(DEFAULT_STATE_WINRATES_XLSX))

    all_records: List[AgentPickRecord] = []
    processed = 0
    ignored = 0

    for json_path in iter_json_files(input_dir, recursive=DEFAULT_RECURSIVE):
        data = load_match(json_path)
        if not data:
            ignored += 1
            continue
        processed += 1
        all_records.extend(
            build_agent_records_for_match(
                data=data,
                filename=json_path.name,
                trade_window_ms=DEFAULT_TRADE_WINDOW_MS,
                state_wr=state_wr,
            )
        )

    if not all_records:
        raise SystemExit(f"Nenhuma linha de agente foi gerada. Verifique a pasta: {input_dir}")

    write_workbook(all_records, output_xlsx)

    unique_agents = len({r.agent for r in all_records})
    unique_maps = len({r.map_name for r in all_records})
    print("Input:", input_dir)
    print("JSONs processados:", processed)
    print("JSONs ignorados:", ignored)
    print("Picks analisados:", len(all_records))
    print("Agentes encontrados:", unique_agents)
    print("Mapas encontrados:", unique_maps)
    print("Output:", output_xlsx)

    open_file_after_save(output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
