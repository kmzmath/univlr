import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ---- CONFIG PARA RODAR DIRETO NO VS CODE (sem argumentos) ----
import config_hub

# A pasta de trabalho fica fora do repo: e la que estao os JSONs brutos e as
# saidas de analise. Ver config_hub.
BASE_DIR = config_hub.TRABALHO_DIR
DEFAULT_INPUT_DIR = str(BASE_DIR / "Finalizados" / "Válidos" / "Univavá" / "Classificatórias 3")
DEFAULT_OUTPUT_XLSX = ""  # vazio = gera analiseMatches_<nome_da_pasta>.xlsx automaticamente
DEFAULT_RECURSIVE = True


DEFAULT_TRADE_WINDOW_MS = 2000


def _strip_quotes(value: str) -> str:
    s = str(value or "").strip()
    if (s.startswith('"') and s.endswith('"')) or (s.startswith("'") and s.endswith("'")):
        return s[1:-1].strip()
    return s


def resolve_user_path(raw: str, base_dir: Path = BASE_DIR) -> Path:
    """Aceita caminho absoluto ou relativo à pasta do script."""
    raw = _strip_quotes(raw).strip()
    raw = raw.replace("/", "\\") if sys.platform.startswith("win") else raw
    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = base_dir / p
    return p.resolve()


def build_default_output_path(input_dir: Path) -> Path:
    folder_name = input_dir.name.strip() or "campeonato"
    safe_name = re.sub(r'[^A-Za-z0-9À-ÿ._ -]+', '_', folder_name).strip(' ._-') or "campeonato"
    safe_name = re.sub(r'\s+', '_', safe_name)
    return BASE_DIR / f"analiseMatches_{safe_name}.xlsx"



DEFAULT_TEAMS_XLSX = str(config_hub.TEAMS_XLSX)
DEFAULT_PLAYERS_XLSX = str(config_hub.PLAYERS_XLSX)
DEFAULT_STATE_WINRATES_XLSX = str(config_hub.ROUND_STATE_WINRATES_XLSX)




# ---- MAPAS: internalName/uuid -> displayName ----
MAP_INTERNAL_TO_DISPLAY = {
    "Infinity": "Abyss",
    "Ascent": "Ascent",
    "Duality": "Bind",
    "Foxtrot": "Breeze",
    "Rook": "Corrode",
    "Canyon": "Fracture",
    "Triad": "Haven",
    "Port": "Icebox",
    "Jam": "Lotus",
    "Pitt": "Pearl",
    "Bonsai": "Split",
    "Juliett": "Sunset",
}

MAP_UUID_TO_DISPLAY = {
    "224b0a95-48b9-f703-1bd8-67aca101a61f": "Abyss",
    "7eaecc1b-4337-bbf6-6ab9-04b8f06b3319": "Ascent",
    "2c9d57ec-4431-9c5e-2939-8f9ef6dd5cba": "Bind",
    "2fb9a4fd-47b8-4e7d-a969-74b4046ebd53": "Breeze",
    "1c18ab1f-420d-0d8b-71d0-77ad3c439115": "Corrode",
    "b529448b-4d60-346e-e89e-00a4c527a405": "Fracture",
    "2bee0dc9-4ffe-519b-1cbd-7fbe763a6047": "Haven",
    "e2ad5c54-4114-a870-9641-8ea21279579a": "Icebox",
    "2fe4ed3a-450a-948b-6d6b-e89a78e680a9": "Lotus",
    "fd267378-4d1d-484f-ff52-77821ed10dc2": "Pearl",
    "d960549e-485c-e861-8d71-aa9d1aed12a2": "Split",
    "92584fbe-486a-b1b2-9faa-39b0f486b498": "Sunset",
}


# ---- AGENTES: uuid -> displayName ----
AGENT_ID_TO_NAME: Dict[str, str] = {
    "41fb69c1-4189-7b37-f117-bcaf1e96f1bf": "Astra",
    "5f8d3a7f-467b-97f3-062c-13acf203c006": "Breach",
    "9f0d8ba9-4140-b941-57d3-a7ad57c6b417": "Brimstone",
    "22697a3d-45bf-8dd7-4fec-84a9e28c69d7": "Chamber",
    "1dbf2edd-4729-0984-3115-daa5eed44993": "Clove",
    "117ed9e3-49f3-6512-3ccf-0cada7e3823b": "Cypher",
    "cc8b64c8-4b25-4ff9-6e7f-37b4da43d235": "Deadlock",
    "dade69b4-4f5a-8528-247b-219e5a1facd6": "Fade",
    "e370fa57-4757-3604-3648-499e1f642d3f": "Gekko",
    "95b78ed7-4637-86d9-7e41-71ba8c293152": "Harbor",
    "0e38b510-41a8-5780-5e8f-568b2a4f2d6c": "Iso",
    "add6443a-41bd-e414-f6ad-e58d267f4e95": "Jett",
    "601dbbe7-43ce-be57-2a40-4abd24953621": "KAY/O",
    "1e58de9c-4950-5125-93e9-a0aee9f98746": "Killjoy",
    "7c8a4701-4de6-9355-b254-e09bc2a34b72": "Miks",
    "bb2a4828-46eb-8cd1-e765-15848195d751": "Neon",
    "8e253930-4c05-31dd-1b6c-968525494517": "Omen",
    "eb93336a-449b-9c1b-0a54-a891f7921d69": "Phoenix",
    "f94c3b30-42be-e959-889c-5aa313dba261": "Raze",
    "a3bfb853-43b2-7238-a4f1-ad90e9e46bcc": "Reyna",
    "569fdd95-4d10-43ab-ca70-79becc718b46": "Sage",
    "6f2a04ca-43e0-be17-7f36-b3908627744d": "Skye",
    "320b2a48-4d9b-a075-30f1-1f93a9b638fa": "Sova",
    "b444168c-4e35-8076-db47-ef9bf368f384": "Tejo",
    "92eeef5d-43b5-1d4a-8d03-b3927a09034b": "Veto",
    "707eab51-4836-f488-046a-cda6bf494859": "Viper",
    "efba5359-4016-a1e5-7626-b1ae76895940": "Vyse",
    "df1cb487-4902-002e-5c17-d28e83e78588": "Waylay",
    "7f94d92c-4234-0a36-9646-3a87eb8b5c89": "Yoru",
}

# ---- FUNÇÃO (ROLE) POR AGENTE ----
AGENT_TO_ROLE = {
    "Brimstone": "Controlador",
    "Viper": "Controlador",
    "Omen": "Controlador",
    "Astra": "Controlador",
    "Harbor": "Controlador",
    "Clove": "Controlador",
    "Miks": "Controlador",
    "Phoenix": "Duelista",
    "Jett": "Duelista",
    "Reyna": "Duelista",
    "Raze": "Duelista",
    "Yoru": "Duelista",
    "Neon": "Duelista",
    "Iso": "Duelista",
    "Waylay": "Duelista",
    "Sova": "Iniciador",
    "Breach": "Iniciador",
    "Skye": "Iniciador",
    "KAY/O": "Iniciador",
    "Fade": "Iniciador",
    "Gekko": "Iniciador",
    "Tejo": "Iniciador",
    "Killjoy": "Sentinela",
    "Cypher": "Sentinela",
    "Sage": "Sentinela",
    "Chamber": "Sentinela",
    "Deadlock": "Sentinela",
    "Vyse": "Sentinela",
    "Veto": "Sentinela",
}


HEADERS = [
    "Player",
    "Agente",
    "Mapa",
    "KAST",
    "Rounds",
    "R.Ganhos",
    "R.Perdidos",
    "R.KAST",
    "ACS",
    "Kills",
    "Deaths",
    "Assists",
    "KPR",
    "DPR",
    "APR",
    "ADR",
    "FK",
    "FD",
    "1K",
    "2K",
    "3K",
    "4K",
    "5K",
    "Clutches",
    "Imp.Total (pp)",
    "Imp/Round (pp)",
    "Rating",
    "rAAting 1.0",
    "Partidas",
]

ARMOR_ID_TO_EXTRA_HP = {
    "": 0,
    "4DEC83D5-4902-9AB3-BED6-A7A390761157": 25,  # Light
    "822BCAB2-40A2-324E-C137-E09195AD7692": 50,  # Heavy
}


def armor_extra_hp(armor_id: Any) -> int:
    a = str(armor_id or "").strip().upper()
    return ARMOR_ID_TO_EXTRA_HP.get(a, 50)  # desconhecido -> assume 50


TOTAL_HEADERS = HEADERS.copy()
TOTAL_HEADERS[2] = "Equipe"


def safe_div(n: float, d: float) -> float:
    return (n / d) if d else 0.0


def riot_id(game_name: Optional[str], tag_line: Optional[str]) -> str:
    gn = (game_name or "").strip()
    tl = (tag_line or "").strip()
    if gn and tl:
        return f"{gn}#{tl}"
    return gn or ""


def map_name_from_map_id(map_id: Optional[str]) -> str:
    if not map_id:
        return ""
    s = str(map_id).strip()

    if s in MAP_UUID_TO_DISPLAY:
        return MAP_UUID_TO_DISPLAY[s]

    parts = [p for p in s.split("/") if p]
    internal = parts[-1] if parts else s
    return MAP_INTERNAL_TO_DISPLAY.get(internal, internal)


def agent_name_from_character_id(character_id: Optional[str]) -> str:
    if not character_id:
        return ""
    cid = str(character_id).strip()
    return AGENT_ID_TO_NAME.get(cid, cid)


@dataclass(frozen=True)
class KillEvent:
    t_ms: int
    killer: str
    victim: str
    assistants: Tuple[str, ...]


@dataclass(frozen=True)
class RowRecord:
    team_key: str
    cells: List[Any]


def iter_json_files(input_dir: Path, recursive: bool) -> Iterable[Path]:
    pattern = "**/*.json" if recursive else "*.json"
    yield from sorted(input_dir.glob(pattern))

def load_match(path: Path) -> Optional[Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return None
        if "matchInfo" not in data or "players" not in data or "roundResults" not in data:
            return None
        return data
    except Exception:
        return None


def rating_value(kast_frac: float, kpr: float, dpr: float, apr: float, adr: float) -> float:
    kast_pct = kast_frac * 100.0
    return (
        0.0073 * kast_pct
        + 0.3591 * kpr
        - 0.5329 * dpr
        + (0.2372 * (2.13 * kpr + 0.42 * apr - 0.41))
        + 0.0032 / 1.5 * adr
        + 0.1587
    )


def clamp(value: float, lo: float, hi: float) -> float:
    return max(lo, min(hi, value))


def rAAting_1_0_value(
    kast_frac: float,
    kpr: float,
    dpr: float,
    apr: float,
    adr: float,
    imp_round_pp: float,
) -> float:
    """
    rAAting 1.0 fixo calibrado na base analiseMatches_Todas.

    Escala:
      - mediana da base de calibração = 1.00
      - usa KAST como fração, ex.: 0.6923 para 69,23%
      - usa Imp/Round em pontos percentuais, ex.: 2.50 para +2,50 pp
    """
    z_kpr = clamp((kpr - 0.6941176471) / 0.1941309654, -3.0, 3.0)
    z_adr = clamp((adr - 104.4644808743) / 28.3082643335, -3.0, 3.0)
    z_kast = clamp((kast_frac - 0.6923076923) / 0.0802383950, -3.0, 3.0)
    z_imp = clamp((imp_round_pp + 0.4159106155) / 4.4510585893, -3.0, 3.0)
    z_dpr_inv = -clamp((dpr - 0.7136563877) / 0.0969349482, -3.0, 3.0)
    z_apr = clamp((apr - 0.2588235294) / 0.1203522890, -3.0, 3.0)

    score_wr = (
        0.22 * z_kpr
        + 0.18 * z_adr
        + 0.18 * z_kast
        + 0.17 * z_imp
        + 0.15 * z_dpr_inv
        + 0.10 * z_apr
    )

    rating = 1 + 0.2883448104 * (score_wr - 0.0853107262)
    return clamp(rating, 0.30, 1.80)


def sheet_name_safe(name: str) -> str:
    cleaned = re.sub(r"[\[\]\:\*\?\/\\]+", "_", (name or "").strip())
    cleaned = cleaned[:31] if cleaned else "sheet"
    return cleaned


def _unique_preserve(items: List[str]) -> List[str]:
    seen: Set[str] = set()
    out: List[str] = []
    for x in items:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def _read_first_sheet_as_dict_rows(xlsx_path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        d: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = r[i] if i < len(r) else None
        if any(v is not None and str(v).strip() != "" for v in d.values()):
            out.append(d)
    return out

def normalize_team(value: Any) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == "-":
        return None
    return s


def team_index_from_header(header: Any) -> Optional[int]:
    """
    Identifica colunas cronológicas de equipe no players.xlsx.

    Aceita cabeçalhos como:
      - team1, team 1, team_1, team-1
      - time1, time 1
      - equipe1, equipe 1

    A coluna sem número, ex. "team", é tratada como team 1.
    """
    if header is None:
        return None

    s = str(header).strip().lower()
    if not s:
        return None

    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    compact = re.sub(r"\s+", "", s)

    m = re.fullmatch(r"(?:team|time|equipe)(\d*)", compact)
    if not m:
        return None

    raw_idx = m.group(1)
    idx = int(raw_idx) if raw_idx else 1
    return idx if idx > 0 else None


def get_team_values_from_row(row: Dict[str, Any]) -> List[Any]:
    """
    Retorna os valores das colunas team/time/equipe em ordem cronológica.
    Ex.: team 1, team 2, team 3, ...
    """
    team_items: List[Tuple[int, Any]] = []

    for header, value in row.items():
        idx = team_index_from_header(header)
        if idx is not None:
            team_items.append((idx, value))

    team_items.sort(key=lambda item: item[0])
    return [value for _idx, value in team_items]


def get_current_team(*team_values: Any) -> Optional[str]:
    """
    Retorna a equipe atual usando a última coluna team N preenchida.

    Compatível com:
      get_current_team(team1, team2)
      get_current_team([team1, team2, team3, ...])
    """
    if len(team_values) == 1 and isinstance(team_values[0], (list, tuple)):
        values = list(team_values[0])
    else:
        values = list(team_values)

    for value in reversed(values):
        team = normalize_team(value)
        if team is not None:
            return team

    return None

def normalize_no_team_value(value: Any) -> Optional[str]:
    s = str(value or "").strip()
    if not s:
        return None
    folded = s.casefold().replace("-", "_").replace(" ", "_")
    if folded in {"sem_time", "sem_equipe", "free_agent", "none", "null", "n/a", "na", "-"}:
        return None
    return s


def get_row_value_by_normalized_header(row: Dict[str, Any], accepted_headers: Set[str]) -> Any:
    for k, v in row.items():
        normalized = re.sub(r"[\s_\-]+", "", str(k or "").strip().casefold())
        if normalized in accepted_headers:
            return v
    return None


def get_current_team_from_row(row: Dict[str, Any]) -> Optional[str]:
    """Usa current_team como fonte de verdade; fallback legado: último team N."""
    value = get_row_value_by_normalized_header(
        row,
        {"currentteam", "teamatual", "timeatual", "equipeatual"},
    )
    if value is not None:
        return normalize_no_team_value(value)

    # Fallback para planilhas antigas sem current_team.
    return get_current_team(get_team_values_from_row(row))



def load_rosters(teams_xlsx: Path, players_xlsx: Path) -> Tuple[List[str], Dict[str, Tuple[str, str]]]:
    team_order: List[str] = []
    nick_to_info: Dict[str, Tuple[str, str]] = {}
    current_teams_from_players: List[str] = []

    # players.xlsx (Jogador, team 1, team 2, team 3, ..., nick 1, nick 2, ...)
    if players_xlsx.exists():
        rows = _read_first_sheet_as_dict_rows(players_xlsx)
        for r in rows:
            jogador = str(r.get("Jogador") or "").strip()
            if not jogador:
                continue

            # Regra atual: current_team define o roster ativo.
            # team 1, team 2, team 3... ficam apenas como histórico.
            current_team = get_current_team_from_row(r) or ""
            if current_team:
                current_teams_from_players.append(current_team)

            for k, v in r.items():
                if str(k).lower().startswith("nick"):
                    nick = str(v or "").strip()
                    if nick:
                        nick_to_info[nick.lower()] = (jogador, current_team)

    # teams.xlsx (team, player1..)
    if teams_xlsx.exists():
        rows = _read_first_sheet_as_dict_rows(teams_xlsx)
        for r in rows:
            team = str(r.get("team") or r.get("Team") or "").strip()
            if not team:
                continue

            team_order.append(team)

            for k, v in r.items():
                if str(k).lower().startswith("player"):
                    nick = str(v or "").strip()
                    if not nick:
                        continue

                    key = nick.lower()

                    # se não veio do players.xlsx, usa teams.xlsx como fallback
                    if key not in nick_to_info:
                        nick_to_info[key] = (nick, team)
                    else:
                        jogador, t = nick_to_info[key]
                        if not t:
                            nick_to_info[key] = (jogador, team)

    # Garante que equipes atuais presentes só no players.xlsx também ganhem aba no relatório.
    # Isso é importante para lines temporárias, como uma line única do JUBS.
    if current_teams_from_players:
        team_order.extend(sorted(set(current_teams_from_players), key=lambda s: s.casefold()))

    team_order = _unique_preserve(team_order)
    return team_order, nick_to_info

def resolve_player_and_team(
    game_name: Optional[str],
    tag_line: Optional[str],
    puuid_fallback: str,
    nick_to_info: Dict[str, Tuple[str, str]],
) -> Tuple[str, str]:
    rid = riot_id(game_name, tag_line) or puuid_fallback
    key = rid.strip().lower()
    if key in nick_to_info:
        jogador, team = nick_to_info[key]
        jogador = (jogador or "").strip() or rid
        team = (team or "").strip()
        return jogador, team
    return rid, ""


def infer_team_for_player(recs: List[RowRecord]) -> str:
    counts: Dict[str, int] = defaultdict(int)
    for r in recs:
        t = (r.team_key or "").strip()
        if t:
            counts[t] += 1
    if not counts:
        return "-"
    return max(counts.items(), key=lambda kv: kv[1])[0]


# ---------------- Winrates por estado (carrega do XLSX) ----------------

def _parse_win_percent(v: Any) -> float:
    """
    Retorna fração [0..1].
    Aceita:
      - Excel % (0.5) -> 0.5
      - "50.0%" -> 0.5
      - 50 -> 0.5 (assume percent)
    """
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        x = float(v)
        if x <= 1.0:
            return max(0.0, min(1.0, x))
        return max(0.0, min(1.0, x / 100.0))
    s = str(v).strip().replace(",", ".")
    if not s:
        return 0.0
    if s.endswith("%"):
        try:
            return max(0.0, min(1.0, float(s[:-1].strip()) / 100.0))
        except Exception:
            return 0.0
    try:
        x = float(s)
        if x <= 1.0:
            return max(0.0, min(1.0, x))
        return max(0.0, min(1.0, x / 100.0))
    except Exception:
        return 0.0


def _parse_situacao(s: Any) -> Optional[Tuple[int, int]]:
    if s is None:
        return None
    txt = str(s).strip().lower()
    m = re.match(r"^\s*(\d)\s*v\s*(\d)\s*$", txt)
    if not m:
        return None
    x = int(m.group(1))
    y = int(m.group(2))
    if not (1 <= x <= 5 and 1 <= y <= 5):
        return None
    return (x, y)


def load_state_winrates_xlsx(path: Path) -> Dict[Tuple[int, int], float]:
    if not path.exists():
        raise SystemExit(f"Arquivo de winrates não encontrado: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Arquivo de winrates vazio: {path}")

    headers = [str(h or "").strip() for h in rows[0]]
    hmap = {h.casefold(): i for i, h in enumerate(headers)}

    # tenta achar colunas
    idx_sit = None
    idx_winp = None

    for k, i in hmap.items():
        if k in ("situação", "situacao", "situacão"):
            idx_sit = i
        if k in ("win%", "win %", "winrate", "win rate", "win"):
            idx_winp = i

    if idx_sit is None:
        raise SystemExit(f"Coluna 'Situação' não encontrada em: {path}")

    # se não achar Win%, tenta usar "Win%" aproximado pelo nome
    if idx_winp is None:
        for i, h in enumerate(headers):
            if "win" in h.casefold() and "%" in h:
                idx_winp = i
                break

    if idx_winp is None:
        raise SystemExit(f"Coluna 'Win%' não encontrada em: {path}")

    out: Dict[Tuple[int, int], float] = {}
    for r in rows[1:]:
        if not r or idx_sit >= len(r):
            continue
        key = _parse_situacao(r[idx_sit])
        if not key:
            continue
        v = r[idx_winp] if idx_winp < len(r) else None
        out[key] = _parse_win_percent(v)

    if not out:
        raise SystemExit(f"Nenhuma linha válida de Situação/Win% encontrada em: {path}")

    return out


# ---------------- Impacto usando tabela XvY ----------------

def get_state_wr(state_wr: Dict[Tuple[int, int], float], x: int, y: int) -> float:
    if x <= 0:
        return 0.0
    if y <= 0:
        return 1.0
    return state_wr.get((x, y), 0.5)  # fallback neutro


def delta_kill_pp(state_wr: Dict[Tuple[int, int], float], x: int, y: int) -> float:
    before = get_state_wr(state_wr, x, y)
    after = get_state_wr(state_wr, x, y - 1)
    return (after - before) * 100.0


def delta_death_pp(state_wr: Dict[Tuple[int, int], float], x: int, y: int) -> float:
    before = get_state_wr(state_wr, x, y)
    after = get_state_wr(state_wr, x - 1, y)
    return (after - before) * 100.0


# ---------------- Core: build_records_for_match ----------------

def build_records_for_match(
    data: Dict[str, Any],
    filename: str,
    trade_window_ms: int,
    nick_to_info: Dict[str, Tuple[str, str]],
    state_wr: Dict[Tuple[int, int], float],
) -> List[RowRecord]:
    match_info = data.get("matchInfo") or {}
    map_name = map_name_from_map_id(match_info.get("mapId"))

    players = data.get("players") or []
    round_results = data.get("roundResults") or []
    rounds_total_match = len(round_results)

    side_by_puuid: Dict[str, str] = {}
    meta_by_puuid: Dict[str, Dict[str, Any]] = {}

    external_name_by_puuid: Dict[str, str] = {}
    external_team_by_puuid: Dict[str, str] = {}

    for p in players:
        if p.get("isObserver"):
            continue
        puuid = p.get("puuid")
        side = p.get("teamId")
        if puuid and side in ("Blue", "Red"):
            puuid = str(puuid)
            side_by_puuid[puuid] = str(side)
            meta_by_puuid[puuid] = p

            disp, team_key = resolve_player_and_team(
                p.get("gameName"), p.get("tagLine"), puuid, nick_to_info
            )
            external_name_by_puuid[puuid] = disp
            external_team_by_puuid[puuid] = team_key

    rounds_seen = defaultdict(int)
    rounds_won = defaultdict(int)
    rounds_lost = defaultdict(int)
    kast_rounds = defaultdict(int)

    dmg_total = defaultdict(float)

    fk = defaultdict(int)
    fd = defaultdict(int)

    k1 = defaultdict(int)
    k2 = defaultdict(int)
    k3 = defaultdict(int)
    k4 = defaultdict(int)
    k5 = defaultdict(int)

    clutches = defaultdict(int)

    impact_total_pp = defaultdict(float)  # soma (pontos percentuais) por player na partida

    roster_blue = {pu for pu, s in side_by_puuid.items() if s == "Blue"}
    roster_red = {pu for pu, s in side_by_puuid.items() if s == "Red"}

    for rr in round_results:
        winning_side = rr.get("winningTeam")
        player_stats = rr.get("playerStats") or []

        # armor (para cap de dano efetivo por vítima no round)
        armor_extra_by_puuid: Dict[str, int] = {}
        for ps in player_stats:
            pu = ps.get("puuid")
            if not pu:
                continue
            pu = str(pu)
            if pu not in side_by_puuid:
                continue
            armor_id = (ps.get("economy") or {}).get("armor")
            armor_extra_by_puuid[pu] = armor_extra_hp(armor_id)

        # dano do round organizado por vítima -> atacante -> dano
        damage_by_victim: Dict[str, Dict[str, float]] = defaultdict(lambda: defaultdict(float))

        participants: Set[str] = set()
        kills_count_round: Dict[str, int] = defaultdict(int)

        # usado na lógica de trade
        damage_by_receiver: Dict[str, Set[str]] = defaultdict(set)
        events: List[KillEvent] = []
        events_seen: Set[Tuple[int, str, str]] = set()

        assistant_set: Set[str] = set()
        victims_set: Set[str] = set()

        for ps in player_stats:
            puuid = ps.get("puuid")
            if not puuid:
                continue
            puuid = str(puuid)
            if puuid not in side_by_puuid:
                continue

            participants.add(puuid)

            # coleta dano por vítima/atacante (sem somar direto no dmg_total)
            for d in (ps.get("damage") or []):
                receiver = d.get("receiver")
                dmg = d.get("damage", 0) or 0
                if not receiver or dmg <= 0:
                    continue

                receiver = str(receiver)
                if receiver not in side_by_puuid:
                    continue

                # ignora self-damage e friendly damage
                if receiver == puuid:
                    continue
                if side_by_puuid.get(receiver) == side_by_puuid.get(puuid):
                    continue

                damage_by_victim[receiver][puuid] += float(dmg)
                damage_by_receiver[receiver].add(puuid)  # trade (só para inimigos)

            # kills
            kills_list = ps.get("kills") or []
            kills_count_round[puuid] += len(kills_list)
            for k in kills_list:
                t = k.get("timeSinceRoundStartMillis")
                if t is None:
                    continue
                killer = str(k.get("killer") or "")
                victim = str(k.get("victim") or "")
                assistants = tuple(str(a) for a in (k.get("assistants") or []))

                if victim and victim not in side_by_puuid:
                    continue
                if killer and killer not in side_by_puuid:
                    killer = ""  # ambiente/unknown

                key = (int(t), killer, victim)
                if key in events_seen:
                    continue
                events_seen.add(key)

                events.append(KillEvent(int(t), killer, victim, assistants))
                if victim:
                    victims_set.add(victim)
                for a in assistants:
                    if a in side_by_puuid:
                        assistant_set.add(a)

        # rounds win/loss
        for pu in participants:
            rounds_seen[pu] += 1
            if winning_side in ("Blue", "Red"):
                if side_by_puuid[pu] == winning_side:
                    rounds_won[pu] += 1
                else:
                    rounds_lost[pu] += 1

        # ordena eventos para FK/FD e também para "último killer"
        events.sort(key=lambda e: e.t_ms)

        # FK/FD
        if events:
            min_t = events[0].t_ms
            for e in events:
                if e.t_ms != min_t:
                    break
                if e.killer and e.killer in side_by_puuid:
                    fk[e.killer] += 1
                if e.victim in side_by_puuid:
                    fd[e.victim] += 1

        # quem matou cada vítima (último hit) neste round
        killer_of_victim: Dict[str, str] = {}
        killer_time: Dict[str, int] = {}
        for e in events:
            if not e.victim:
                continue
            if e.victim not in side_by_puuid:
                continue
            if not e.killer or e.killer not in side_by_puuid:
                continue
            prev_t = killer_time.get(e.victim)
            if prev_t is None or e.t_ms > prev_t:
                killer_time[e.victim] = e.t_ms
                killer_of_victim[e.victim] = e.killer

        # aplica "cap" por vítima no round e desconta o overkill do ÚLTIMO (killer)
        for victim, by_attacker in damage_by_victim.items():
            raw_total = sum(by_attacker.values())
            if raw_total <= 0:
                continue

            cap = 100 + armor_extra_by_puuid.get(victim, 50)
            over = raw_total - float(cap)

            effective = dict(by_attacker)

            if over > 0:
                last = killer_of_victim.get(victim)

                # fallback: se não achar killer (ou killer não tem dano registrado), usa quem mais deu dano
                if not last or last not in effective:
                    last = max(effective.items(), key=lambda kv: kv[1])[0]

                take = min(over, effective.get(last, 0.0))
                effective[last] = max(0.0, effective.get(last, 0.0) - take)
                over -= take

                # se ainda sobrar, tira dos maiores contribuidores para garantir total == cap
                if over > 0:
                    for attacker, _ in sorted(effective.items(), key=lambda kv: kv[1], reverse=True):
                        if attacker == last:
                            continue
                        if over <= 0:
                            break
                        take = min(over, effective[attacker])
                        effective[attacker] = max(0.0, effective[attacker] - take)
                        over -= take

            for attacker, dmg in effective.items():
                if dmg > 0:
                    dmg_total[attacker] += dmg

        # -------- Impacto (pp) por kill/death usando estados XvY (do XLSX) --------
        alive_blue_i = set(roster_blue)
        alive_red_i = set(roster_red)

        for e in events:
            victim = e.victim
            killer = e.killer

            if not victim or victim not in side_by_puuid:
                continue

            if victim in alive_blue_i:
                v_side = "Blue"
            elif victim in alive_red_i:
                v_side = "Red"
            else:
                continue  # já morto / repetido

            # estado antes da morte, sempre do ponto de vista do time do player em questão
            if v_side == "Blue":
                x_v = len(alive_blue_i)
                y_v = len(alive_red_i)
            else:
                x_v = len(alive_red_i)
                y_v = len(alive_blue_i)

            # impacto negativo para quem morreu
            impact_total_pp[victim] += delta_death_pp(state_wr, x_v, y_v)

            # impacto positivo para o killer (se conhecido e válido)
            if killer and killer in side_by_puuid:
                k_side = side_by_puuid[killer]
                if k_side == "Blue":
                    x_k = len(alive_blue_i)
                    y_k = len(alive_red_i)
                else:
                    x_k = len(alive_red_i)
                    y_k = len(alive_blue_i)
                impact_total_pp[killer] += delta_kill_pp(state_wr, x_k, y_k)

            # aplica a morte no estado
            if v_side == "Blue":
                alive_blue_i.remove(victim)
            else:
                alive_red_i.remove(victim)

        # multi-kills por round
        for pu in participants:
            kc = kills_count_round.get(pu, 0)
            if kc == 1:
                k1[pu] += 1
            elif kc == 2:
                k2[pu] += 1
            elif kc == 3:
                k3[pu] += 1
            elif kc == 4:
                k4[pu] += 1
            elif kc == 5:
                k5[pu] += 1

        # KAST
        deaths_by_victim: Dict[str, List[KillEvent]] = defaultdict(list)
        for e in events:
            if e.victim:
                deaths_by_victim[e.victim].append(e)

        for pu in participants:
            kill_cond = kills_count_round.get(pu, 0) > 0
            assist_cond = pu in assistant_set
            survived_cond = pu not in victims_set

            traded_cond = False
            if not survived_cond:
                victim_side = side_by_puuid.get(pu)
                for death in deaths_by_victim.get(pu, []):
                    t0 = death.t_ms
                    t1 = t0 + trade_window_ms

                    involved: Set[str] = set()
                    if death.killer:
                        involved.add(death.killer)
                    involved.update(death.assistants)
                    involved.update(damage_by_receiver.get(pu, set()))
                    involved = {
                        x for x in involved
                        if x in side_by_puuid and side_by_puuid.get(x) != victim_side
                    }
                    if not involved:
                        continue

                    for ev in events:
                        if ev.t_ms <= t0:
                            continue
                        if ev.t_ms > t1:
                            break
                        if side_by_puuid.get(ev.killer) == victim_side and ev.victim in involved:
                            traded_cond = True
                            break
                    if traded_cond:
                        break

            if kill_cond or assist_cond or survived_cond or traded_cond:
                kast_rounds[pu] += 1

        # clutch
        alive_blue = set(roster_blue)
        alive_red = set(roster_red)

        clutch_candidate_blue: Optional[str] = None
        clutch_candidate_red: Optional[str] = None

        for e in events:
            if not e.victim or e.victim not in side_by_puuid:
                continue
            vside = side_by_puuid[e.victim]

            if vside == "Blue":
                alive_blue.discard(e.victim)
                if clutch_candidate_blue is None and len(alive_blue) == 1:
                    last = next(iter(alive_blue))
                    if len(alive_red) > 1:
                        clutch_candidate_blue = last
            else:
                alive_red.discard(e.victim)
                if clutch_candidate_red is None and len(alive_red) == 1:
                    last = next(iter(alive_red))
                    if len(alive_blue) > 1:
                        clutch_candidate_red = last

        if winning_side == "Blue" and clutch_candidate_blue and clutch_candidate_blue not in victims_set:
            clutches[clutch_candidate_blue] += 1
        elif winning_side == "Red" and clutch_candidate_red and clutch_candidate_red not in victims_set:
            clutches[clutch_candidate_red] += 1

    out: List[RowRecord] = []
    for puuid, p in meta_by_puuid.items():
        stats = p.get("stats") or {}
        rounds_played = int(stats.get("roundsPlayed") or rounds_seen.get(puuid, 0))
        if rounds_played <= 0:
            continue

        score = float(stats.get("score") or 0)
        kills = int(stats.get("kills") or 0)
        deaths = int(stats.get("deaths") or 0)
        assists = int(stats.get("assists") or 0)

        kast_cnt = int(kast_rounds.get(puuid, 0))
        seen = int(rounds_seen.get(puuid, 0)) or rounds_played
        kast_frac = safe_div(kast_cnt, seen)

        acs = safe_div(score, rounds_played)
        kpr = safe_div(kills, rounds_played)
        dpr = safe_div(deaths, rounds_played)
        apr = safe_div(assists, rounds_played)
        adr = safe_div(dmg_total.get(puuid, 0.0), rounds_played)

        imp_total = float(impact_total_pp.get(puuid, 0.0))
        imp_pr = safe_div(imp_total, float(rounds_played))

        rtg = rating_value(kast_frac, kpr, dpr, apr, adr)
        raa = rAAting_1_0_value(kast_frac, kpr, dpr, apr, adr, imp_pr)

        player_display = external_name_by_puuid.get(puuid) or resolve_player_and_team(
            p.get("gameName"), p.get("tagLine"), puuid, nick_to_info
        )[0]
        team_key = external_team_by_puuid.get(puuid, "")

        out.append(
            RowRecord(
                team_key=team_key,
                cells=[
                    player_display,
                    agent_name_from_character_id(p.get("characterId")),
                    map_name,
                    kast_frac,
                    rounds_total_match,
                    int(rounds_won.get(puuid, 0)),
                    int(rounds_lost.get(puuid, 0)),
                    kast_cnt,
                    acs,
                    kills,
                    deaths,
                    assists,
                    kpr,
                    dpr,
                    apr,
                    adr,
                    int(fk.get(puuid, 0)),
                    int(fd.get(puuid, 0)),
                    int(k1.get(puuid, 0)),
                    int(k2.get(puuid, 0)),
                    int(k3.get(puuid, 0)),
                    int(k4.get(puuid, 0)),
                    int(k5.get(puuid, 0)),
                    int(clutches.get(puuid, 0)),
                    imp_total,
                    imp_pr,
                    rtg,
                    raa,
                    filename,
                ],
            )
        )

    return out


# ---------- Escrita do Excel ----------

def _write_headers(ws, row_idx: int, headers: List[str] = HEADERS) -> None:
    header_font = Font(bold=True)
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=h)
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")


def _apply_number_formats(ws, start_row: int, end_row: int) -> None:
    # 4  KAST (%)
    # 9  ACS (2dp)
    # 13-16 KPR/DPR/APR/ADR (2dp)
    # 25-26 Impact (2dp)
    # 27 Rating antigo (3dp)
    # 28 rAAting 1.0 (3dp)
    for r in range(start_row, end_row + 1):
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=9).number_format = "0.00"
        for c in (13, 14, 15, 16):
            ws.cell(row=r, column=c).number_format = "0.00"
        for c in (25, 26):
            ws.cell(row=r, column=c).number_format = "0.00"
        ws.cell(row=r, column=27).number_format = "0.000"
        ws.cell(row=r, column=28).number_format = "0.000"


def _auto_width(ws) -> None:
    max_col = len(HEADERS)
    max_row = ws.max_row
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(10, max_len + 2), 45)


def _to_int(v: Any) -> int:
    try:
        if v is None or (isinstance(v, str) and not v.strip()):
            return 0
        return int(float(v))
    except Exception:
        return 0


def _to_float(v: Any) -> float:
    try:
        if v is None or (isinstance(v, str) and not v.strip()):
            return 0.0
        return float(v)
    except Exception:
        return 0.0


def compute_player_summary_cells(player_name: str, recs: List[RowRecord]) -> List[Any]:
    # índices 0-based (baseados em HEADERS atual)
    IDX_AGENT = 1
    IDX_ROUNDS = 4
    IDX_WINS = 5
    IDX_LOSSES = 6
    IDX_RKAST = 7
    IDX_ACS = 8
    IDX_KILLS = 9
    IDX_DEATHS = 10
    IDX_ASSISTS = 11
    IDX_ADR = 15
    IDX_FK = 16
    IDX_FD = 17
    IDX_1K = 18
    IDX_2K = 19
    IDX_3K = 20
    IDX_4K = 21
    IDX_5K = 22
    IDX_CLUTCH = 23
    IDX_IMP_TOTAL = 24

    match_count = len(recs)

    total_rounds = sum(_to_int(r.cells[IDX_ROUNDS]) for r in recs)
    total_wins = sum(_to_int(r.cells[IDX_WINS]) for r in recs)
    total_losses = sum(_to_int(r.cells[IDX_LOSSES]) for r in recs)
    total_rkast = sum(_to_int(r.cells[IDX_RKAST]) for r in recs)

    total_kills = sum(_to_int(r.cells[IDX_KILLS]) for r in recs)
    total_deaths = sum(_to_int(r.cells[IDX_DEATHS]) for r in recs)
    total_assists = sum(_to_int(r.cells[IDX_ASSISTS]) for r in recs)

    total_fk = sum(_to_int(r.cells[IDX_FK]) for r in recs)
    total_fd = sum(_to_int(r.cells[IDX_FD]) for r in recs)
    total_1k = sum(_to_int(r.cells[IDX_1K]) for r in recs)
    total_2k = sum(_to_int(r.cells[IDX_2K]) for r in recs)
    total_3k = sum(_to_int(r.cells[IDX_3K]) for r in recs)
    total_4k = sum(_to_int(r.cells[IDX_4K]) for r in recs)
    total_5k = sum(_to_int(r.cells[IDX_5K]) for r in recs)
    total_clutch = sum(_to_int(r.cells[IDX_CLUTCH]) for r in recs)

    kast_frac = safe_div(float(total_rkast), float(total_rounds))

    acs_vals = [_to_float(r.cells[IDX_ACS]) for r in recs]
    avg_acs = safe_div(sum(acs_vals), float(len(acs_vals))) if acs_vals else 0.0

    kpr = safe_div(float(total_kills), float(total_rounds))
    dpr = safe_div(float(total_deaths), float(total_rounds))
    apr = safe_div(float(total_assists), float(total_rounds))

    total_damage_est = 0.0
    for r in recs:
        adr_match = _to_float(r.cells[IDX_ADR])
        rounds_match = _to_int(r.cells[IDX_ROUNDS])
        total_damage_est += adr_match * rounds_match
    adr = safe_div(total_damage_est, float(total_rounds))

    # role mais jogada (por match), senão Flex
    role_counts: Dict[str, int] = defaultdict(int)
    for r in recs:
        agent = str(r.cells[IDX_AGENT] or "").strip()
        role = AGENT_TO_ROLE.get(agent, "")
        if role:
            role_counts[role] += 1

    if match_count > 0 and role_counts:
        top_role, top_count = max(role_counts.items(), key=lambda kv: kv[1])
        main_role = top_role if (top_count / match_count) > 0.5 else "Flex"
    else:
        main_role = "Flex"

    # Impacto total e por round (total_rounds)
    total_imp = sum(_to_float(r.cells[IDX_IMP_TOTAL]) for r in recs)
    imp_pr_total = safe_div(total_imp, float(total_rounds))

    rtg = rating_value(kast_frac, kpr, dpr, apr, adr)
    raa = rAAting_1_0_value(kast_frac, kpr, dpr, apr, adr, imp_pr_total)

    return [
        player_name,     # Player
        main_role,       # Agente (role)
        "-",             # Mapa (ou Equipe no sheet totals)
        kast_frac,
        total_rounds,
        total_wins,
        total_losses,
        total_rkast,
        avg_acs,
        total_kills,
        total_deaths,
        total_assists,
        kpr,
        dpr,
        apr,
        adr,
        total_fk,
        total_fd,
        total_1k,
        total_2k,
        total_3k,
        total_4k,
        total_5k,
        total_clutch,
        total_imp,       # Imp.Total (pp)
        imp_pr_total,    # Imp/Round (pp)
        rtg,             # Rating antigo
        raa,             # rAAting 1.0
        match_count,     # Partidas (count)
    ]


def write_all_matches_sheet(ws, records: List[RowRecord]) -> None:
    ws.title = "all_matches"
    _write_headers(ws, 1)
    for rec in records:
        ws.append([config_hub.texto_seguro(c) for c in rec.cells])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(records) + 1}"

    if records:
        _apply_number_formats(ws, start_row=2, end_row=len(records) + 1)
    _auto_width(ws)


def write_totals_all_players_sheet(ws, records: List[RowRecord]) -> None:
    ws.title = "stats_total_all_players"
    _write_headers(ws, 1, headers=TOTAL_HEADERS)

    by_player: Dict[str, List[RowRecord]] = defaultdict(list)
    for r in records:
        player = str(r.cells[0] or "").strip() or "SEM_NOME"
        by_player[player].append(r)

    players_sorted = sorted(by_player.keys(), key=lambda s: s.casefold())

    row_cursor = 2
    bold_font = Font(bold=True)

    for player in players_sorted:
        recs = by_player[player]
        team = infer_team_for_player(recs)

        summary = compute_player_summary_cells(player, recs)
        summary[2] = team  # coluna 3 agora é Equipe

        for col_idx, v in enumerate(summary, start=1):
            c = ws.cell(row=row_cursor, column=col_idx, value=v)
            c.font = bold_font

        row_cursor += 1

    if row_cursor > 2:
        _apply_number_formats(ws, start_row=2, end_row=row_cursor - 1)
    _auto_width(ws)


def write_team_sheet_per_player(ws, team_name: str, records: List[RowRecord]) -> None:
    ws.title = sheet_name_safe(team_name)

    by_player: Dict[str, List[RowRecord]] = defaultdict(list)
    for r in records:
        player = str(r.cells[0] or "").strip() or "SEM_NOME"
        by_player[player].append(r)

    players_sorted = sorted(by_player.keys(), key=lambda s: s.casefold())

    title_font = Font(bold=True, size=14)
    title_align = Alignment(horizontal="left", vertical="center")
    total_font = Font(bold=True)

    row_cursor = 1
    for player in players_sorted:
        recs = by_player[player]

        # título do jogador
        ws.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=len(HEADERS))
        tc = ws.cell(row=row_cursor, column=1, value=player)
        tc.font = title_font
        tc.alignment = title_align
        row_cursor += 1

        # headers
        _write_headers(ws, row_cursor)
        row_cursor += 1

        # dados
        start_data = row_cursor
        for rec in recs:
            for col_idx, v in enumerate(rec.cells, start=1):
                ws.cell(row=row_cursor, column=col_idx, value=v)
            row_cursor += 1
        end_data = row_cursor - 1

        if end_data >= start_data:
            _apply_number_formats(ws, start_row=start_data, end_row=end_data)

        # TOTAL
        summary_cells = compute_player_summary_cells(player, recs)
        for col_idx, v in enumerate(summary_cells, start=1):
            c = ws.cell(row=row_cursor, column=col_idx, value=v)
            c.font = total_font
        _apply_number_formats(ws, start_row=row_cursor, end_row=row_cursor)
        row_cursor += 1

        row_cursor += 2  # espaço entre tabelas

    _auto_width(ws)


def write_workbook(all_records: List[RowRecord], team_order: List[str], output_path: Path) -> None:
    wb = Workbook()

    # 1) all_matches
    ws_all = wb.active
    write_all_matches_sheet(ws_all, all_records)

    # 2) stats_total_all_players
    ws_tot = wb.create_sheet(title="stats_total_all_players")
    write_totals_all_players_sheet(ws_tot, all_records)

    # 3) Abas de times: cria SOMENTE para times que apareceram nos JSONs analisados.
    #
    # team_order vem de players.xlsx/teams.xlsx e pode conter todos os times cadastrados.
    # Se iterarmos direto nele, o Excel fica cheio de abas vazias.
    # Aqui filtramos pelo team_key presente em pelo menos uma linha gerada a partir dos JSONs.
    teams_with_records: Set[str] = {
        str(r.team_key or "").strip()
        for r in all_records
        if str(r.team_key or "").strip()
    }

    ordered_teams: List[str] = []
    seen_team_keys: Set[str] = set()

    # Mantém a ordem/base vinda dos cadastros quando o time apareceu.
    for team in team_order:
        clean_team = str(team or "").strip()
        if not clean_team or clean_team not in teams_with_records:
            continue
        if clean_team in seen_team_keys:
            continue
        ordered_teams.append(clean_team)
        seen_team_keys.add(clean_team)

    # Inclui qualquer time que apareceu nos JSONs, mas não estava em team_order.
    for team in sorted(teams_with_records - seen_team_keys, key=lambda s: s.casefold()):
        ordered_teams.append(team)

    for team in sorted(ordered_teams, key=lambda s: s.casefold()):
        team_recs = [r for r in all_records if str(r.team_key or "").strip() == team]
        if not team_recs:
            continue

        ws = wb.create_sheet(title=sheet_name_safe(team))
        write_team_sheet_per_player(ws, team, team_recs)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def open_file_after_save(path: Path) -> None:
    p = str(path)
    try:
        if sys.platform.startswith("win"):
            os.startfile(p)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", p], check=False)
        else:
            subprocess.run(["xdg-open", p], check=False)
    except Exception as e:
        print(f"Não foi possível abrir o arquivo automaticamente: {e}")


def main() -> int:
    input_dir = resolve_user_path(DEFAULT_INPUT_DIR)
    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"Pasta inválida em DEFAULT_INPUT_DIR: {input_dir}")

    if str(DEFAULT_OUTPUT_XLSX or "").strip():
        output_xlsx = resolve_user_path(DEFAULT_OUTPUT_XLSX)
    else:
        output_xlsx = build_default_output_path(input_dir).resolve()

    print("Pasta analisada:", input_dir)
    print("Recursivo:", bool(DEFAULT_RECURSIVE))

    teams_xlsx = resolve_user_path(DEFAULT_TEAMS_XLSX)
    players_xlsx = resolve_user_path(DEFAULT_PLAYERS_XLSX)
    state_xlsx = resolve_user_path(DEFAULT_STATE_WINRATES_XLSX)

    team_order, nick_to_info = load_rosters(teams_xlsx, players_xlsx)

    # Carrega winrates (XvY) do XLSX sempre atualizado.
    state_wr = load_state_winrates_xlsx(state_xlsx)

    json_files = list(iter_json_files(input_dir, recursive=DEFAULT_RECURSIVE))
    if not json_files:
        raise SystemExit(f"Nenhum .json encontrado em DEFAULT_INPUT_DIR: {input_dir}")

    print("JSONs encontrados:", len(json_files))

    all_records: List[RowRecord] = []
    invalid_files = 0
    for json_path in json_files:
        data = load_match(json_path)
        if not data:
            invalid_files += 1
            continue
        all_records.extend(
            build_records_for_match(
                data=data,
                filename=json_path.name,
                trade_window_ms=DEFAULT_TRADE_WINDOW_MS,
                nick_to_info=nick_to_info,
                state_wr=state_wr,
            )
        )

    write_workbook(all_records, team_order, output_xlsx)
    print(f"Arquivos inválidos/ignorados: {invalid_files}")
    print(f"OK: {len(all_records)} linhas geradas em: {output_xlsx}")

    open_file_after_save(output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
