import json
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_ROOT = ROOT / "dados_excel"
EVENTS_ROOT = ROOT / "campeonatos"
EVENT_META_FILES = {"campeonato.json", "event.json", "manifest.json", "tournament.json"}
ASSET_EXTENSIONS = (".png", ".jpg", ".jpeg", ".webp", ".svg")

KNOWN_BAD_TEAM_LOGOS = {
    "ufcg_pensaopet": {
        "https://univlr-bucket.s3.sa-east-1.amazonaws.com/teams/notapixel/notapixel.png",
    },
}

PHOTO_ALIASES = {
    "fillip1n": ["filip1n"],
    "pagode": ["pagod"],
    "gbzz": ["gbz"],
    "japinha99": ["jeyp"],
}


KNOWN_ACRONYMS = {
    "aaeu",
    "a2e",
    "caap",
    "ceub",
    "fei",
    "fatec",
    "gdu",
    "ibmec",
    "inatel",
    "pg",
    "pucgo",
    "pucc",
    "ufg",
    "uff",
    "ufmg",
    "ufpb",
    "unifesp",
    "ufu",
    "ufrj",
    "umc",
    "unirv",
    "xxii",
    "aoc",
    "cia",
    "jubs",
    "lpe",
    "sp",
    "uni",
}

EVENT_NAME_OVERRIDES = {
    "cia_2026": "CIA 2026",
    "pre_jubs": "Pré-JUBS",
    "pre_jubssp": "Pré-JUBS SP",
    "pre_jubs_sp": "Pré-JUBS SP",
    "rush_series_esquenta": "RUSH Series - Esquenta",
    "rivvalsgg": "RivvalsGG",
    "uni_kick_off_inters": "Uni Kick-OFF Inters",
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_public_path(value):
    path = clean(value).replace("\\", "/")
    if not path:
        return ""
    if path.startswith(("http://", "https://")):
        return path
    return path.lstrip("/")


def migrated_asset_path(value):
    path = normalize_public_path(value)
    if not path:
        return ""
    lower = path.lower()
    replacements = [
        ("logos/campeonatos/", "assets/tournament-icons/"),
        ("logos/", "assets/team-logos/"),
        ("static/agentes_icones/", "assets/agent-icons/"),
        ("static/mapas_fotos/", "assets/maps/"),
        ("agents_icons/", "assets/agent-icons/"),
        ("agentes_icones/", "assets/agent-icons/"),
        ("mapas_fotos/", "assets/maps/"),
        ("fotos_players/", "assets/player-photos/"),
    ]
    for old, new in replacements:
        if lower.startswith(old):
            return new + path[len(old):]
    return path


def public_asset(value):
    path = migrated_asset_path(value)
    if not path:
        return ""
    if path.startswith(("http://", "https://")):
        return path
    candidate = ROOT / path
    if candidate.exists() and candidate.is_file():
        return path
    if not candidate.suffix:
        for extension in ASSET_EXTENSIONS:
            with_extension = candidate.with_suffix(extension)
            if with_extension.exists() and with_extension.is_file():
                return normalize_public_path(with_extension.relative_to(ROOT).as_posix())
    return ""


def declared_asset(value):
    return public_asset(value) or migrated_asset_path(value)


def logo_for_team(team_id, provided_logo, logos_by_slug):
    normalized_logo = normalize_public_path(provided_logo)
    if normalized_logo in KNOWN_BAD_TEAM_LOGOS.get(team_id.lower(), set()):
        provided_logo = ""
    logo = public_asset(provided_logo)
    if logo:
        return logo
    team_key = team_id.lower()
    if team_key in logos_by_slug:
        return logos_by_slug[team_key]
    for stem, path in logos_by_slug.items():
        if team_key.startswith(f"{stem}_") or stem.startswith(f"{team_key}_"):
            return path
    return ""


def title_from_slug(slug):
    parts = clean(slug).split("_")
    return " ".join(part.upper() if part.lower() in KNOWN_ACRONYMS else part.capitalize() for part in parts if part)


def load_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [clean(item) for item in next(rows)]
    fields = [(index, header) for index, header in enumerate(headers) if header]

    for row in rows:
        item = {header: clean(row[index]) if index < len(row) else "" for index, header in fields}
        if any(item.values()):
            yield item


def local_logos():
    folder = ROOT / "assets" / "team-logos"
    if not folder.exists():
        return {}
    return {path.stem.lower(): normalize_public_path(path.relative_to(ROOT).as_posix()) for path in folder.glob("*") if path.is_file()}


# As fotos ficam em assets/player-photos/<pasta da equipe>/<nome do jogador>.
# O indice guarda a pasta de cada arquivo para nunca emprestar a foto de um
# jogador homonimo de outra equipe.
def player_photo_index():
    folder = ROOT / "assets" / "player-photos"
    index = {"scoped": {}, "bare": {}}
    if not folder.exists():
        return index
    for path in sorted(folder.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in ASSET_EXTENSIONS:
            continue
        stem_key = slug_key(path.stem)
        if not stem_key:
            continue
        public_path = normalize_public_path(path.relative_to(ROOT).as_posix())
        relative = path.relative_to(folder)
        folder_key = "/".join(part for part in (slug_key(item) for item in relative.parts[:-1]) if part)
        index["scoped"].setdefault(f"{folder_key}/{stem_key}", public_path)
        index["bare"].setdefault(stem_key, []).append((folder_key, public_path))
    return index


# O nome do arquivo segue o nome do jogador (nunca o nick in-game), com o
# primeiro token como alternativa para nomes compostos ("Queijo Coallho").
def player_photo_keys(name):
    keys = []

    def add_key(value):
        key = slug_key(value)
        if key and key not in keys:
            keys.append(key)
        if "_" in key:
            first_token = key.split("_", 1)[0]
            if first_token and first_token not in keys:
                keys.append(first_token)

    add_key(name)
    for alias in PHOTO_ALIASES.get(slug_key(name), []):
        add_key(alias)
    return keys


# Pastas por organizacao ("macklogic" para "macklogic_red") e por equipe antiga
# do jogador continuam valendo; pasta de outra equipe, nao.
def photo_folder_matches_team(folder_key, team_keys):
    if not folder_key:
        return True
    for team in team_keys:
        if not team:
            continue
        if folder_key == team or team.startswith(f"{folder_key}_") or folder_key.startswith(f"{team}_"):
            return True
    return False


def resolve_player_photo(index, name, current_team, team_history):
    team_keys = []
    for team in [current_team, *team_history]:
        if team and team not in team_keys:
            team_keys.append(team)
    keys = player_photo_keys(name)
    for key in keys:
        for team in team_keys:
            scoped = index["scoped"].get(f"{team}/{key}")
            if scoped:
                return scoped
    for key in keys:
        for folder_key, public_path in index["bare"].get(key, []):
            if photo_folder_matches_team(folder_key, team_keys):
                return public_path
    return ""


def slug_key(value):
    text = unicodedata.normalize("NFD", clean(value).lower())
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    output = []
    for char in text:
        if char.isalnum():
            output.append(char)
        elif output and output[-1] != "_":
            output.append("_")
    return "".join(output).strip("_")


def event_id_from_folder(name):
    text = clean(name).replace("Kick-OFF", "KickOFF").replace("kick-off", "kickoff")
    return slug_key(text).replace("_", "-")


def event_name_from_folder(folder):
    key = slug_key(folder.name)
    if key in EVENT_NAME_OVERRIDES:
        return EVENT_NAME_OVERRIDES[key]
    parts = clean(folder.name).replace("-", " ").split()
    return " ".join(part.upper() if part.lower() in KNOWN_ACRONYMS else part.capitalize() for part in parts)


def has_match_files(folder):
    return any(
        path.is_file() and path.suffix.lower() == ".json" and path.name.lower() not in EVENT_META_FILES
        for path in folder.iterdir()
    )


def has_event_metadata(folder):
    return any((folder / filename).exists() for filename in EVENT_META_FILES)


def read_event_metadata(folder):
    for filename in EVENT_META_FILES:
        path = folder / filename
        if not path.exists():
            continue
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return {}
    return {}


def event_lookup_keys(*values):
    keys = set()
    for value in values:
        text = clean(value)
        if not text:
            continue
        base = slug_key(text.replace("-", " "))
        if not base:
            continue
        keys.add(base)
        keys.add(base.replace("_", ""))
        keys.add(base.replace("_", "-"))
    return keys


def event_date_value(value):
    if not value:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    text = clean(value)
    if len(text) >= 19 and text[10] == " ":
        return text.replace(" ", "T", 1)
    return text


def read_tournament_infos():
    path = DATA_ROOT / "tournaments.xlsx"
    if not path.exists():
        return {}

    tournaments = {}
    for row in load_rows(path):
        name = row.get("name", "")
        keys = event_lookup_keys(name, row.get("id", ""))
        if not keys:
            continue
        item = {
            "sourceId": row.get("id", ""),
            "name": name,
            "organizer": row.get("organizer", ""),
            "startAt": event_date_value(row.get("start_date", "") or row.get("startAt", "")),
            "endAt": event_date_value(row.get("end_date", "") or row.get("endAt", "")),
            "logo": public_asset(row.get("logo", "")) or declared_asset(row.get("logo", "")),
        }
        for key in keys:
            tournaments[key] = item
    return tournaments


def build_data_sources():
    events = []
    if not EVENTS_ROOT.exists():
        return {"events": events}
    tournament_infos = read_tournament_infos()

    # Campeonatos ja disputados sao descobertos pelos JSONs de partida; campeonatos
    # futuros entram com uma pasta contendo apenas o campeonato.json (files vazio).
    event_folders = [
        folder
        for folder in EVENTS_ROOT.rglob("*")
        if folder.is_dir() and (has_match_files(folder) or has_event_metadata(folder))
    ]

    for folder in sorted(event_folders, key=lambda item: item.relative_to(EVENTS_ROOT).as_posix().lower()):
        meta = read_event_metadata(folder)
        files = [
            normalize_public_path(path.relative_to(ROOT).as_posix())
            for path in sorted(folder.glob("*.json"), key=lambda item: item.name.lower())
            if path.name.lower() not in EVENT_META_FILES
        ]
        if not files and not meta:
            continue

        relative_folder = folder.relative_to(EVENTS_ROOT)
        event_id = clean(meta.get("id")) or event_id_from_folder(relative_folder.as_posix())
        event_name = clean(meta.get("name")) or event_name_from_folder(folder)
        tournament = next(
            (
                tournament_infos[key]
                for key in event_lookup_keys(event_id, event_name, folder.name, relative_folder.as_posix())
                if key in tournament_infos
            ),
            {},
        )
        parent_tournament = {}
        for parent in relative_folder.parents:
            if parent == Path("."):
                continue
            parent_tournament = next(
                (
                    tournament_infos[key]
                    for key in event_lookup_keys(parent.name, parent.as_posix())
                    if key in tournament_infos
                ),
                {},
            )
            if parent_tournament:
                break
        event_item = {
            "id": event_id,
            "name": clean(meta.get("name")) or tournament.get("name") or event_name,
            "organizer": clean(meta.get("organizer")) or tournament.get("organizer") or "UNIVLR",
            "source": clean(meta.get("source")) or "Histórico",
            "sourceId": clean(meta.get("sourceId")) or tournament.get("sourceId", ""),
            "startAt": clean(meta.get("startAt")) or clean(meta.get("start_date")) or tournament.get("startAt", ""),
            "endAt": clean(meta.get("endAt")) or clean(meta.get("end_date")) or tournament.get("endAt", ""),
            "logo": public_asset(meta.get("logo", "")) or tournament.get("logo", "") or parent_tournament.get("logo", ""),
            "folder": normalize_public_path(folder.relative_to(ROOT).as_posix()),
            "files": files,
        }
        events.append(event_item)

    return {"events": events}


def read_states():
    path = DATA_ROOT / "estados.xlsx"
    if not path.exists():
        return []
    states = []
    for row in load_rows(path):
        sigla = row.get("sigla", "").upper()
        if not sigla:
            continue
        states.append(
            {
                "id": row.get("id", ""),
                "sigla": sigla,
                "name": row.get("nome", ""),
                "icon": public_asset(row.get("icone", "")),
                "region": row.get("regiao", ""),
            }
        )
    return states


def read_agents():
    path = DATA_ROOT / "agents.xlsx"
    if not path.exists():
        return []
    agents = []
    for row in load_rows(path):
        slug = row.get("slug", "").lower()
        if not slug or slug == "?":
            continue
        agents.append(
            {
                "id": row.get("id", ""),
                "slug": slug,
                "name": row.get("nome_agente", "") or title_from_slug(slug),
                "role": row.get("classe", ""),
                "icon": public_asset(f"assets/agent-icons/{slug}.png") or public_asset(row.get("icon", "")) or declared_asset(row.get("icon", "")),
            }
        )
    return agents


def read_maps():
    path = DATA_ROOT / "maps.xlsx"
    if not path.exists():
        return []
    maps = []
    for row in load_rows(path):
        slug = row.get("slug", "").lower()
        if not slug or slug == "?":
            continue
        local_icon = next(
            (
                normalize_public_path(candidate.relative_to(ROOT).as_posix())
                for extension in ASSET_EXTENSIONS
                for candidate in [ROOT / "assets" / "maps" / f"{slug}{extension}"]
                if candidate.exists() and candidate.is_file()
            ),
            "",
        )
        maps.append(
            {
                "id": row.get("id", ""),
                "slug": slug,
                "name": row.get("nome_mapa", "") or title_from_slug(slug),
                "icon": local_icon or public_asset(row.get("icon", "")) or declared_asset(row.get("icon", "")),
            }
        )
    return maps


def parse_win_percent(value):
    if value in (None, ""):
        return 0.5
    if isinstance(value, str):
        text = value.strip().replace("%", "").replace(",", ".")
        try:
            number = float(text)
        except ValueError:
            return 0.5
    else:
        number = float(value)
    return number / 100 if number > 1 else number


def read_state_winrates():
    path = DATA_ROOT / "round_state_winrates.xlsx"
    if not path.exists():
        return []

    rows = list(load_rows(path))
    output = []
    for row in rows:
        state = row.get("Situação") or row.get("Situacao") or ""
        if "v" not in state:
            continue
        left, right = state.lower().split("v", 1)
        try:
            allies = int(left.strip())
            enemies = int(right.strip())
        except ValueError:
            continue
        output.append(
            {
                "state": f"{allies}v{enemies}",
                "allies": allies,
                "enemies": enemies,
                "occurrences": row.get("Ocorrências", "") or row.get("Ocorrencias", ""),
                "wins": row.get("Vitórias", "") or row.get("Vitorias", ""),
                "winRate": parse_win_percent(row.get("Win%", "")),
            }
        )
    return output


def read_team_infos(states_by_sigla, logos_by_slug):
    path = DATA_ROOT / "teams_infos.xlsx"
    if not path.exists():
        return {}
    teams = {}
    for row in load_rows(path):
        slug = slug_key(row.get("slug", ""))
        if not slug:
            continue
        state_sigla = row.get("estado", "").upper()
        state = states_by_sigla.get(state_sigla, {})
        logo = logo_for_team(slug, row.get("logo", ""), logos_by_slug)
        teams[slug] = {
            "id": slug,
            "sourceId": row.get("id", ""),
            "slug": slug,
            "displayName": row.get("name", "") or title_from_slug(slug),
            "tag": row.get("tag", "") or slug,
            "org": row.get("org", ""),
            "orgTag": row.get("orgTag", ""),
            "logo": logo,
            "state": state_sigla,
            "stateName": state.get("name", ""),
            "stateFlag": state.get("icon", ""),
            "stateRegion": state.get("region", ""),
            "socials": {
                "instagram": row.get("instagram", ""),
            },
        }
    return teams


def read_team_lineups():
    path = DATA_ROOT / "teams.xlsx"
    if not path.exists():
        return {}

    teams = {}
    for row in load_rows(path):
        team_id = slug_key(row.get("team", ""))
        if not team_id:
            continue
        lineup = []
        for index in range(1, 20):
            player_name = row.get(f"player{index}", "")
            if player_name:
                lineup.append({"slot": index, "name": player_name})
        teams[team_id] = {"id": team_id, "lineup": lineup}
    return teams


def read_teams(states_by_sigla):
    logos_by_slug = local_logos()
    infos = read_team_infos(states_by_sigla, logos_by_slug)
    rosters = read_team_lineups()
    teams = []

    for team_id in sorted(set(infos) | set(rosters)):
        info = infos.get(team_id, {})
        roster = rosters.get(team_id, {})
        logo = info.get("logo", "") or logo_for_team(team_id, "", logos_by_slug)
        teams.append(
            {
                "id": team_id,
                "sourceId": info.get("sourceId", ""),
                "slug": team_id,
                "tag": info.get("tag", team_id),
                "displayName": info.get("displayName", title_from_slug(team_id)),
                "org": info.get("org", ""),
                "orgTag": info.get("orgTag", ""),
                "logo": logo,
                "state": info.get("state", ""),
                "stateName": info.get("stateName", ""),
                "stateFlag": info.get("stateFlag", ""),
                "stateRegion": info.get("stateRegion", ""),
                "socials": info.get("socials", {}),
                "lineup": roster.get("lineup", []),
            }
        )
    return teams


def read_players():
    path = DATA_ROOT / "players.xlsx"
    if not path.exists():
        return []
    players = []
    photos = player_photo_index()

    for row in load_rows(path):
        name = row.get("Jogador", "")
        if not name:
            continue
        team_history = [row[key] for key in sorted(row) if key.lower().startswith("team ") and row[key]]
        nick_history = [row[key] for key in sorted(row) if key.lower().startswith("nick ") and row[key]]
        puuid = row.get("puuid", "")
        current_team = slug_key(row.get("current_team", ""))
        team_history = [slug_key(team_id) for team_id in team_history]
        photo = resolve_player_photo(photos, name, current_team, team_history)
        players.append(
            {
                "id": puuid or name,
                "name": name,
                "puuid": puuid,
                "currentTeam": current_team,
                "teamHistory": team_history,
                "nickHistory": nick_history,
                "photo": photo,
            }
        )
    return players


def main():
    states = read_states()
    states_by_sigla = {state["sigla"]: state for state in states}
    metadata = {
        "sourceFiles": [
            "dados_excel/teams_infos.xlsx",
            "dados_excel/teams.xlsx",
            "dados_excel/tournaments.xlsx",
            "dados_excel/players.xlsx",
            "dados_excel/estados.xlsx",
            "dados_excel/agents.xlsx",
            "dados_excel/maps.xlsx",
            "dados_excel/round_state_winrates.xlsx",
            "assets/team-logos/",
            "assets/tournament-icons/",
            "assets/agent-icons/",
            "assets/maps/",
            "assets/player-photos/",
        ],
        "teams": read_teams(states_by_sigla),
        "players": read_players(),
        "states": states,
        "agents": read_agents(),
        "maps": read_maps(),
        "stateWinrates": read_state_winrates(),
    }
    output_path = ROOT / "metadata.json"
    output_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    data_sources = build_data_sources()
    sources_path = ROOT / "data-sources.json"
    sources_path.write_text(json.dumps(data_sources, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(output_path),
                "sources": str(sources_path),
                "events": len(data_sources["events"]),
                "matchFiles": sum(len(event["files"]) for event in data_sources["events"]),
                "teams": len(metadata["teams"]),
                "players": len(metadata["players"]),
                "states": len(metadata["states"]),
                "agents": len(metadata["agents"]),
                "maps": len(metadata["maps"]),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
