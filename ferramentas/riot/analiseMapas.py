import json
import os
import re
import subprocess
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ---- CONFIG PARA RODAR DIRETO NO VS CODE (sem argumentos) ----
import config_hub

# A pasta de trabalho fica fora do repo: e la que estao os JSONs brutos e as
# saidas de analise. Ver config_hub.
BASE_DIR = config_hub.TRABALHO_DIR
DEFAULT_INPUT_DIR = str(BASE_DIR / "Finalizados" / "Válidos" / "Univavá" / "Classificatórias 1")
DEFAULT_OUTPUT_XLSX = str(BASE_DIR / "analiseMapas.xlsx")
DEFAULT_RECURSIVE = True


# Aproximações de tempo (ms) para calcular duração média do round quando o JSON não traz duração pronta
ROUND_TIMER_EXPIRED_MS = 100_000  # round timer padrão (sem plant) ~100s
SPIKE_TIMER_MS = 45_000          # spike explode ~45s após plant


# ---- MAPAS: internalName/uuid -> displayName ----
MAP_INTERNAL_TO_DISPLAY = {
    "Infinity": "Abyss",
    "Ascent": "Ascent",
    "Duality": "Bind",
    "Foxtrot": "Breeze",
    "Rook": "Corrode",
    "Canyon": "Fracture",
    "Triad": "Haven",
    "Port": "Icebox",
    "Jam": "Lotus",
    "Pitt": "Pearl",
    "Bonsai": "Split",
    "Juliett": "Sunset",
}

MAP_UUID_TO_DISPLAY = {
    "224b0a95-48b9-f703-1bd8-67aca101a61f": "Abyss",
    "7eaecc1b-4337-bbf6-6ab9-04b8f06b3319": "Ascent",
    "2c9d57ec-4431-9c5e-2939-8f9ef6dd5cba": "Bind",
    "2fb9a4fd-47b8-4e7d-a969-74b4046ebd53": "Breeze",
    "1c18ab1f-420d-0d8b-71d0-77ad3c439115": "Corrode",
    "b529448b-4d60-346e-e89e-00a4c527a405": "Fracture",
    "2bee0dc9-4ffe-519b-1cbd-7fbe763a6047": "Haven",
    "e2ad5c54-4114-a870-9641-8ea21279579a": "Icebox",
    "2fe4ed3a-450a-948b-6d6b-e89a78e680a9": "Lotus",
    "fd267378-4d1d-484f-ff52-77821ed10dc2": "Pearl",
    "d960549e-485c-e861-8d71-aa9d1aed12a2": "Split",
    "92584fbe-486a-b1b2-9faa-39b0f486b498": "Sunset",
}


HEADERS = [
    "Mapa",
    "Vezes jogado",
    "Rounds jogados",
    "%winrate Ataque",
    "%winrate Defesa",
    "%vezes spike plantada",
    "%winrate retake",
    "Duração média do round (s)",
]


def safe_div(n: float, d: float) -> float:
    return (n / d) if d else 0.0


def map_name_from_map_id(map_id: Optional[str]) -> str:
    if not map_id:
        return ""
    s = str(map_id).strip()

    if s in MAP_UUID_TO_DISPLAY:
        return MAP_UUID_TO_DISPLAY[s]

    parts = [p for p in s.split("/") if p]
    internal = parts[-1] if parts else s
    return MAP_INTERNAL_TO_DISPLAY.get(internal, internal)


def iter_json_files(input_dir: Path, recursive: bool) -> Iterable[Path]:
    pattern = "**/*.json" if recursive else "*.json"
    yield from sorted(input_dir.glob(pattern))


def load_match(path: Path) -> Optional[Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return None
        if "matchInfo" not in data or "roundResults" not in data:
            return None
        return data
    except Exception:
        return None


def is_spike_planted(rr: Dict[str, Any]) -> bool:
    # No seu JSON, plantRoundTime > 0 quando houve plant.
    t = rr.get("plantRoundTime")
    try:
        return int(t or 0) > 0
    except Exception:
        return False


def _get_int(v: Any) -> int:
    try:
        return int(v)
    except Exception:
        return 0


def compute_round_duration_ms(rr: Dict[str, Any]) -> int:
    """
    Como o JSON não traz 'roundDurationMillis', fazemos uma aproximação:
    - Defuse: usa defuseRoundTime
    - Bomb exploded: usa plantRoundTime + 45s
    - Timer expired: usa 100s
    - Eliminations: usa o maior timeSinceRoundStartMillis das kills
    - Sempre pega o max() entre kills/plant/defuse para não subcontar em casos mistos.
    """
    result = str(rr.get("roundResult") or "").strip().lower()
    code = str(rr.get("roundResultCode") or "").strip().lower()

    plant_ms = _get_int(rr.get("plantRoundTime") or 0)
    defuse_ms = _get_int(rr.get("defuseRoundTime") or 0)

    kill_times: List[int] = []
    for ps in (rr.get("playerStats") or []):
        for k in (ps.get("kills") or []):
            t = k.get("timeSinceRoundStartMillis")
            if t is None:
                continue
            kill_times.append(_get_int(t))

    last_kill_ms = max(kill_times) if kill_times else 0

    # Defuse
    if defuse_ms > 0 or code == "defuse" or "defus" in result:
        return max(defuse_ms, plant_ms, last_kill_ms)

    # Bomb exploded (não apareceu nesse arquivo, mas pode existir em outros)
    if "explod" in result or code in {"detonate", "detonation", "bombdetonated", "explode"}:
        if plant_ms > 0:
            return max(plant_ms + SPIKE_TIMER_MS, last_kill_ms)
        # fallback
        return max(last_kill_ms, plant_ms)

    # Timer expired
    if "timer expired" in result or code == "timeout":
        return max(ROUND_TIMER_EXPIRED_MS, last_kill_ms, plant_ms)

    # Elimination / outros
    return max(last_kill_ms, plant_ms, defuse_ms)


@dataclass
class MapAgg:
    matches: int = 0
    rounds: int = 0
    attack_wins: int = 0
    defense_wins: int = 0
    plants: int = 0
    retake_wins: int = 0
    duration_sum_ms: int = 0


def analyze_maps(input_dir: Path, recursive: bool) -> Dict[str, MapAgg]:
    aggs: Dict[str, MapAgg] = defaultdict(MapAgg)

    for json_path in iter_json_files(input_dir, recursive=recursive):
        data = load_match(json_path)
        if not data:
            continue

        match_info = data.get("matchInfo") or {}
        map_name = map_name_from_map_id(match_info.get("mapId")) or "DESCONHECIDO"

        rr_list = data.get("roundResults") or []
        if not rr_list:
            continue

        agg = aggs[map_name]
        agg.matches += 1
        agg.rounds += len(rr_list)

        for rr in rr_list:
            role = str(rr.get("winningTeamRole") or "").strip()
            if role == "Attacker":
                agg.attack_wins += 1
            elif role == "Defender":
                agg.defense_wins += 1

            planted = is_spike_planted(rr)
            if planted:
                agg.plants += 1
                # retake = defesa ganha após plant
                if role == "Defender":
                    agg.retake_wins += 1

            agg.duration_sum_ms += compute_round_duration_ms(rr)

    return aggs


def _auto_width(ws) -> None:
    max_col = ws.max_column
    max_row = ws.max_row
    for col in range(1, max_col + 1):
        max_len = 0
        for row in range(1, max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 55)


def write_excel(aggs: Dict[str, MapAgg], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "map_stats"

    header_font = Font(bold=True)
    for c, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ordena por "vezes jogado" desc, depois nome asc
    rows = []
    for map_name, a in aggs.items():
        total_rounds = a.rounds

        atk_wr = safe_div(a.attack_wins, total_rounds)
        def_wr = safe_div(a.defense_wins, total_rounds)
        plant_pct = safe_div(a.plants, total_rounds)
        retake_wr = safe_div(a.retake_wins, a.plants) if a.plants else 0.0
        avg_dur_s = safe_div(a.duration_sum_ms, total_rounds) / 1000.0

        rows.append((
            map_name,
            a.matches,
            total_rounds,
            atk_wr,
            def_wr,
            plant_pct,
            retake_wr,
            avg_dur_s,
        ))

    rows.sort(key=lambda r: (-r[1], str(r[0]).casefold()))

    r0 = 2
    for row in rows:
        ws.append(list(row))

    # Formatos
    # col 4-7 percent, col 8 seconds
    for r in range(r0, ws.max_row + 1):
        for c in (4, 5, 6, 7):
            ws.cell(row=r, column=c).number_format = "0.0%"
        ws.cell(row=r, column=8).number_format = "0.0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    _auto_width(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def open_file_after_save(path: Path) -> None:
    p = str(path)
    try:
        if sys.platform.startswith("win"):
            os.startfile(p)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", p], check=False)
        else:
            subprocess.run(["xdg-open", p], check=False)
    except Exception as e:
        print(f"Não foi possível abrir o arquivo automaticamente: {e}")


def main() -> int:
    input_dir = Path(DEFAULT_INPUT_DIR).expanduser().resolve()
    output_xlsx = Path(DEFAULT_OUTPUT_XLSX).expanduser().resolve()

    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"Pasta inválida em DEFAULT_INPUT_DIR: {input_dir}")

    print("Pasta analisada:", input_dir)
    print("Recursivo:", bool(DEFAULT_RECURSIVE))

    aggs = analyze_maps(input_dir, recursive=DEFAULT_RECURSIVE)
    if not aggs:
        raise SystemExit("Nenhum .json válido encontrado para analisar em DEFAULT_INPUT_DIR.")

    write_excel(aggs, output_xlsx)
    print(f"OK: {len(aggs)} mapas em: {output_xlsx}")
    open_file_after_save(output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
