import json
import os
import subprocess
import sys
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ---- CONFIG (roda direto no VS Code) ----
import config_hub

# A pasta de trabalho fica fora do repo: e la que estao os JSONs brutos e as
# saidas de analise. Ver config_hub.
BASE_DIR = config_hub.TRABALHO_DIR
DEFAULT_INPUT_DIR = str(BASE_DIR / "Finalizados" / "Válidos")
DEFAULT_OUTPUT_XLSX = str(config_hub.ROUND_STATE_WINRATES_XLSX)
DEFAULT_RECURSIVE = True



@dataclass(frozen=True)
class KillEvent:
    t_ms: int
    killer: str
    victim: str


def iter_json_files(input_dir: Path, recursive: bool) -> Iterable[Path]:
    pattern = "**/*.json" if recursive else "*.json"
    yield from sorted(input_dir.glob(pattern))


def load_match(path: Path) -> Optional[Dict[str, Any]]:
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return None
        if "players" not in data or "roundResults" not in data:
            return None
        return data
    except Exception:
        return None


def extract_rosters(players: List[Dict[str, Any]]) -> Tuple[Set[str], Set[str], Dict[str, str]]:
    """
    Retorna:
      - roster_blue (puuids)
      - roster_red (puuids)
      - side_by_puuid (puuid -> "Blue"/"Red")
    """
    side_by_puuid: Dict[str, str] = {}
    roster_blue: Set[str] = set()
    roster_red: Set[str] = set()

    for p in players or []:
        if p.get("isObserver"):
            continue
        puuid = p.get("puuid")
        side = p.get("teamId")
        if not puuid or side not in ("Blue", "Red"):
            continue
        puuid = str(puuid)
        side = str(side)
        side_by_puuid[puuid] = side
        if side == "Blue":
            roster_blue.add(puuid)
        else:
            roster_red.add(puuid)

    return roster_blue, roster_red, side_by_puuid


def extract_round_kills(rr: Dict[str, Any], side_by_puuid: Dict[str, str]) -> List[KillEvent]:
    """
    Extrai kills do round e ordena por tempo.
    Deduplica por (t, killer, victim) para evitar duplicatas raras.
    """
    seen: Set[Tuple[int, str, str]] = set()
    events: List[KillEvent] = []

    for ps in (rr.get("playerStats") or []):
        for k in (ps.get("kills") or []):
            t = k.get("timeSinceRoundStartMillis")
            if t is None:
                continue
            killer = str(k.get("killer") or "")
            victim = str(k.get("victim") or "")
            if not victim or victim not in side_by_puuid:
                continue
            # killer pode vir vazio em casos estranhos; ainda assim o victim morreu
            if killer and killer not in side_by_puuid:
                # ignora killers não mapeados, mas mantém morte
                killer = ""

            key = (int(t), killer, victim)
            if key in seen:
                continue
            seen.add(key)
            events.append(KillEvent(int(t), killer, victim))

    events.sort(key=lambda e: e.t_ms)
    return events


def add_state_stat(
    stats: Dict[str, Dict[str, int]],
    x: int,
    y: int,
    winner_side: str,
    side_of_x: str,
) -> None:
    """
    Atualiza stats para o estado "xvy" (do ponto de vista do time que tem x vivos).
    Ex:
      se Blue tem 2 e Red tem 4 -> registra "2v4" com win se winner_side == "Blue"
      se Red tem 2 e Blue tem 4 -> registra "2v4" com win se winner_side == "Red"
    """
    if x < 1 or y < 1:
        return
    if x > 5 or y > 5:
        return

    key = f"{x}v{y}"
    stats[key]["occ"] += 1
    if winner_side == side_of_x:
        stats[key]["win"] += 1


def analyze_all_round_states(input_dir: Path, recursive: bool) -> Dict[str, Dict[str, int]]:
    """
    Para cada round, percorre a linha do tempo de kills e registra:
      5v5 (estado inicial) e todos os estados subsequentes após cada kill.
    Calcula win% do "lado com X vivos" no momento do estado.
    """
    stats: Dict[str, Dict[str, int]] = defaultdict(lambda: {"occ": 0, "win": 0})

    for jp in iter_json_files(input_dir, recursive=recursive):
        data = load_match(jp)
        if not data:
            continue

        roster_blue, roster_red, side_by_puuid = extract_rosters(data.get("players") or [])
        if not roster_blue or not roster_red:
            continue

        for rr in (data.get("roundResults") or []):
            winner = rr.get("winningTeam")
            if winner not in ("Blue", "Red"):
                continue

            alive_blue = set(roster_blue)
            alive_red = set(roster_red)

            # Estado inicial do round
            add_state_stat(stats, len(alive_blue), len(alive_red), winner, "Blue")
            add_state_stat(stats, len(alive_red), len(alive_blue), winner, "Red")

            # Processa kills
            events = extract_round_kills(rr, side_by_puuid)
            for e in events:
                v = e.victim
                if v in alive_blue:
                    alive_blue.remove(v)
                elif v in alive_red:
                    alive_red.remove(v)
                else:
                    continue  # vítima já estava morta (duplicata/ruído)

                # Após a kill, registra o novo estado (dos dois lados)
                add_state_stat(stats, len(alive_blue), len(alive_red), winner, "Blue")
                add_state_stat(stats, len(alive_red), len(alive_blue), winner, "Red")

    return stats


def write_excel(stats: Dict[str, Dict[str, int]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "round_state_winrates"

    headers = ["Situação", "Ocorrências", "Vitórias", "Win%"]
    header_font = Font(bold=True)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def sort_key(k: str) -> Tuple[int, int]:
        a, b = k.split("v")
        return (int(a), int(b))

    row = 2
    for k in sorted(stats.keys(), key=sort_key):
        occ = stats[k]["occ"]
        win = stats[k]["win"]
        wr = (win / occ) if occ else 0.0
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=occ)
        ws.cell(row=row, column=3, value=win)
        ws.cell(row=row, column=4, value=wr).number_format = "0.0%"
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    # Auto width
    for col in range(1, 5):
        max_len = 0
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 40)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def open_file_after_save(path: Path) -> None:
    p = str(path)
    try:
        if sys.platform.startswith("win"):
            os.startfile(p)  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.run(["open", p], check=False)
        else:
            subprocess.run(["xdg-open", p], check=False)
    except Exception as e:
        print(f"Não foi possível abrir o arquivo automaticamente: {e}")


def main() -> int:
    input_dir = Path(DEFAULT_INPUT_DIR).expanduser().resolve()
    output_xlsx = Path(DEFAULT_OUTPUT_XLSX).expanduser().resolve()

    if not input_dir.exists() or not input_dir.is_dir():
        raise SystemExit(f"Pasta inválida em DEFAULT_INPUT_DIR: {input_dir}")

    print("Pasta analisada:", input_dir)
    print("Recursivo:", bool(DEFAULT_RECURSIVE))

    stats = analyze_all_round_states(input_dir, recursive=DEFAULT_RECURSIVE)
    if not stats:
        raise SystemExit("Nenhum dado encontrado em DEFAULT_INPUT_DIR (jsons inválidos ou sem rounds).")

    write_excel(stats, output_xlsx)
    print(f"OK: gerado {output_xlsx}")
    open_file_after_save(output_xlsx)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
